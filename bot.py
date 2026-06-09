# GSA CAMBIOS BOT v6.0 — INTELIGENCIA DE MERCADO
"""
GSA CAMBIOS — BOT COMPLETO v6.0
Nuevas funciones:
- Alertas con precios reales top 2 (no promedios)
- Tasas sincronizadas al reloj (en punto y media)
- Historial de precios cada 5 min
- Análisis de patrones y proyecciones
- Gestor de capital inteligente
- Reportes diario y semanal automáticos
- Comandos: /mercado /patron /simular /umbral /clientes /resumen_corresponsal
"""

import os
import csv
import time
import sqlite3
import datetime
import requests
import threading
import json
from collections import defaultdict

# ══════════════════════════════════════════════════════════════════════
# CONFIGURACIÓN
# ══════════════════════════════════════════════════════════════════════
TELEGRAM_BOT_TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "")
TELEGRAM_CHAT_ID   = os.getenv("TELEGRAM_CHAT_ID", "")
DB_PATH            = os.getenv("GSA_DB_PATH", "gsa_cambios.db")
CSV_EXPORT_PATH    = os.getenv("GSA_CSV_PATH", "csv_export")

# Zona horaria UTC-4 (Chile invierno = Venezuela)
UTC_OFFSET = -4

# ══════════════════════════════════════════════════════════════════════
# SUPABASE
# ══════════════════════════════════════════════════════════════════════
SUPABASE_URL = os.getenv("SUPABASE_URL", "")
SUPABASE_KEY = os.getenv("SUPABASE_KEY", "")
USE_SUPABASE = bool(SUPABASE_URL and SUPABASE_KEY)
print(f"🔗 Supabase: {'ACTIVADO' if USE_SUPABASE else 'DESACTIVADO'}")

def supa_insert(tabla, datos):
    if not USE_SUPABASE: return None
    try:
        r = requests.post(
            f"{SUPABASE_URL}/rest/v1/{tabla}",
            headers={"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}",
                     "Content-Type": "application/json", "Prefer": "return=minimal"},
            json=datos, timeout=10)
        if r.status_code not in (200, 201):
            print(f"Supabase insert warn ({tabla}): {r.status_code} {r.text[:100]}")
        return r
    except Exception as e:
        print(f"Supabase insert error ({tabla}): {e}"); return None

def supa_select(tabla, query=""):
    if not USE_SUPABASE: return []
    try:
        r = requests.get(f"{SUPABASE_URL}/rest/v1/{tabla}?{query}",
            headers={"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}"}, timeout=10)
        return r.json() if r.status_code == 200 else []
    except Exception as e:
        print(f"Supabase select error ({tabla}): {e}"); return []

def supa_update(tabla, campo, valor, datos):
    if not USE_SUPABASE: return None
    try:
        r = requests.patch(f"{SUPABASE_URL}/rest/v1/{tabla}?{campo}=eq.{valor}",
            headers={"apikey": SUPABASE_KEY, "Authorization": f"Bearer {SUPABASE_KEY}",
                     "Content-Type": "application/json"},
            json=datos, timeout=10)
        return r
    except Exception as e:
        print(f"Supabase update error ({tabla}): {e}"); return None

# ══════════════════════════════════════════════════════════════════════
# FEES Y CONFIGURACIÓN DE NEGOCIO
# ══════════════════════════════════════════════════════════════════════
FEE_USDT_BS  = 0.0025
FEE_USDT_CLP = 0.002
FEE_WU       = 0.025
FEE_CLP_BS   = 0.055
FEE_BS_CLP   = 0.055
FEE_CLP_COP  = 0.075
FEE_COP_CLP  = 0.075
FEE_COP_BS   = 0.055
FEE_BS_COP   = 0.055
MARGEN_BS    = 30
SPREAD_CLP   = 50

# Umbrales de alertas de mercado
SPREAD_MIN_ALERTA   = 10    # Bs mínimo para alertar
SPREAD_CAMBIO_BS    = 2     # Bs de variación para nueva alerta
CAMBIO_MIN_CLP      = 5     # CLP de variación para nueva alerta
SPREAD_SILENCIO     = 7
SPREAD_MODERADO     = 10
SPREAD_BUENO        = 15
SPREAD_PREMIUM      = 20

# Intervalos
INTERVALO_MERCADO_SEG   = 300   # 5 min
INTERVALO_LIQUIDEZ_SEG  = 900   # 15 min
INTERVALO_CSV_SEG       = 3600  # 1 hora

PAUSA_SESION_MIN = 45

TIPOS_OP = [
    "CLP→BS","BS→CLP","CLP→COP","COP→CLP","COP→BS","BS→COP",
    "CLP→USDT","USDT→CLP","BS→USDT","USDT→BS","USD→CLP","CLP→USD",
    "USD→BS","BS→USD","GIRO INT","USDC→USDT","USDT→USDC","USDT→USDT",
]
CORRESPONSALES = [
    "Bancolombia C1","Bancolombia C2",
    "Nequi C1","Nequi C2","Nequi C3","Efectivo Orlando",
]
NOMBRES_CUENTAS = {
    "BS_BANESCO":"Banesco","BS_MERCANTIL":"Mercantil",
    "CLP_COPEC_PAY":"Copec Pay","CLP_BANCOESTADO":"BancoEstado",
    "COP_EFECTIVO_ORLANDO":"Efectivo Orlando",
    "COP_BANCOLOMBIA_C1":"Bancolombia C1","COP_BANCOLOMBIA_C2":"Bancolombia C2",
    "COP_NEQUI_C1":"Nequi C1","COP_NEQUI_C2":"Nequi C2","COP_NEQUI_C3":"Nequi C3",
    "USD_EFECTIVO":"USD Efectivo","USDT_BINANCE":"Binance USDT","USDC_AIRTM":"Airtm USDC",
}
DIAS_SEMANA = ["Lunes","Martes","Miércoles","Jueves","Viernes","Sábado","Domingo"]

# ══════════════════════════════════════════════════════════════════════
# UTILIDADES DE TIEMPO
# ══════════════════════════════════════════════════════════════════════
def now_local():
    """Hora actual en UTC-4."""
    return datetime.datetime.utcnow() + datetime.timedelta(hours=UTC_OFFSET)

def today_local():
    return now_local().date()

def hora_local():
    return now_local().strftime("%H:%M")

def dia_semana_local():
    return DIAS_SEMANA[now_local().weekday()]

# ══════════════════════════════════════════════════════════════════════
# BASE DE DATOS
# ══════════════════════════════════════════════════════════════════════
def get_conn():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn

def init_db():
    conn = get_conn(); c = conn.cursor()
    # Tablas existentes
    c.execute("""CREATE TABLE IF NOT EXISTS tasas (
        id INTEGER PRIMARY KEY AUTOINCREMENT, fecha_hora DATETIME DEFAULT CURRENT_TIMESTAMP,
        bcv_usd REAL, bcv_eur REAL,
        ban_bs_compra REAL, ban_bs_venta REAL, ban_bs_spread REAL,
        mer_bs_compra REAL, mer_bs_venta REAL, mer_bs_spread REAL,
        clp_compra REAL, clp_venta REAL, cop_compra REAL, cop_venta REAL,
        trm REAL, dolar_obs REAL, western REAL,
        limite_clp_bs REAL, limite_clp_cop REAL, limite_bs_cop REAL,
        tasa_gsa_clp_bs REAL, tasa_gsa_bs_clp REAL,
        tasa_gsa_clp_cop REAL, tasa_gsa_cop_clp REAL,
        tasa_gsa_cop_bs REAL, tasa_gsa_bs_cop REAL)""")
    c.execute("""CREATE TABLE IF NOT EXISTS operaciones (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha DATE, hora TIME, cliente TEXT, referente TEXT, tipo_op TEXT,
        origen_fondos TEXT, mon_entrada TEXT, monto_entrada REAL,
        mon_salida TEXT, monto_salida REAL, tasa_cliente REAL, tasa_referencia REAL,
        usdt_equiv REAL, diferencial REAL, metodo TEXT, corresponsal TEXT,
        traslado_bs REAL DEFAULT 0, encomienda_cop REAL DEFAULT 0, repartidor TEXT,
        financiador TEXT, estado TEXT DEFAULT 'Completada', observaciones TEXT,
        cxc_pendiente REAL DEFAULT 0, cxp_pendiente REAL DEFAULT 0,
        snap_pat_bs REAL, snap_dol_obs REAL, snap_trm REAL,
        usuario_telegram TEXT, fecha_registro DATETIME DEFAULT CURRENT_TIMESTAMP)""")
    c.execute("""CREATE TABLE IF NOT EXISTS saldos (
        id INTEGER PRIMARY KEY AUTOINCREMENT, cuenta TEXT UNIQUE,
        moneda TEXT, saldo REAL DEFAULT 0,
        ultima_actualizacion DATETIME DEFAULT CURRENT_TIMESTAMP)""")
    c.execute("""CREATE TABLE IF NOT EXISTS clientes (
        id INTEGER PRIMARY KEY AUTOINCREMENT, nombre TEXT UNIQUE,
        telefono TEXT, pais TEXT DEFAULT 'Venezuela',
        fecha_registro DATE DEFAULT CURRENT_DATE, ultima_operacion DATE,
        operaciones_total INTEGER DEFAULT 0, volumen_usdt REAL DEFAULT 0,
        ganancia_generada REAL DEFAULT 0)""")
    c.execute("""CREATE TABLE IF NOT EXISTS gastos (
        id INTEGER PRIMARY KEY AUTOINCREMENT, fecha DATE DEFAULT CURRENT_DATE,
        categoria TEXT, descripcion TEXT, monto REAL, moneda TEXT,
        usdt_equiv REAL, comprobante TEXT, usuario_telegram TEXT,
        fecha_registro DATETIME DEFAULT CURRENT_TIMESTAMP)""")
    c.execute("""CREATE TABLE IF NOT EXISTS tesoreria (
        id INTEGER PRIMARY KEY AUTOINCREMENT, fecha DATE DEFAULT CURRENT_DATE,
        tipo TEXT, persona TEXT, direccion TEXT, monto REAL, moneda TEXT,
        usdt_equiv REAL, estado TEXT DEFAULT 'Completado', usuario_telegram TEXT,
        fecha_registro DATETIME DEFAULT CURRENT_TIMESTAMP)""")
    c.execute("""CREATE TABLE IF NOT EXISTS cuentas_pendientes (
        id INTEGER PRIMARY KEY AUTOINCREMENT, tipo TEXT,
        fecha DATE DEFAULT CURRENT_DATE, contraparte TEXT, concepto TEXT,
        monto REAL, moneda TEXT, usdt_equiv REAL, vencimiento DATE,
        prioridad TEXT DEFAULT '🟡 Esta semana', estado TEXT DEFAULT 'Pendiente',
        op_origen_id INTEGER, fecha_registro DATETIME DEFAULT CURRENT_TIMESTAMP)""")
    c.execute("""CREATE TABLE IF NOT EXISTS auditoria (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha_hora DATETIME DEFAULT CURRENT_TIMESTAMP,
        usuario TEXT, accion TEXT, modulo TEXT, registro_id INTEGER,
        valor_anterior TEXT, valor_nuevo TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS saldos_iniciales (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cuenta TEXT UNIQUE, saldo REAL DEFAULT 0,
        fecha DATE DEFAULT CURRENT_DATE)""")
    # TABLA NUEVA: precios_historicos
    c.execute("""CREATE TABLE IF NOT EXISTS precios_historicos (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha DATE, hora TIME, dia_semana TEXT,
        bs_venta_1 REAL, bs_venta_2 REAL,
        bs_compra_1 REAL, bs_compra_2 REAL,
        bs_spread_real REAL,
        clp_compra_1 REAL, clp_compra_2 REAL,
        cop_compra_1 REAL, cop_compra_2 REAL,
        usuario_venta_1 TEXT, usuario_venta_2 TEXT,
        usuario_compra_1 TEXT, usuario_compra_2 TEXT,
        disp_venta_1 REAL, disp_venta_2 REAL,
        disp_compra_1 REAL, disp_compra_2 REAL,
        fecha_registro DATETIME DEFAULT CURRENT_TIMESTAMP)""")
    # TABLA NUEVA: oportunidades_perdidas
    c.execute("""CREATE TABLE IF NOT EXISTS oportunidades_perdidas (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha DATE, hora TIME, tipo TEXT,
        descripcion TEXT, monto_requerido REAL, moneda_requerida TEXT,
        ganancia_estimada_perdida REAL, razon TEXT,
        capital_disponible_usdt REAL, observaciones TEXT,
        fecha_registro DATETIME DEFAULT CURRENT_TIMESTAMP)""")
    # TABLA NUEVA: umbrales_liquidez
    c.execute("""CREATE TABLE IF NOT EXISTS umbrales_liquidez (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cuenta TEXT UNIQUE, umbral_minimo REAL, moneda TEXT,
        activo INTEGER DEFAULT 1)""")
    # TABLA NUEVA: config
    c.execute("""CREATE TABLE IF NOT EXISTS config (
        clave TEXT PRIMARY KEY, valor TEXT,
        actualizado DATETIME DEFAULT CURRENT_TIMESTAMP)""")
    conn.commit(); conn.close()
    print("✅ Base de datos v6.0 lista")
    init_inventario()

def init_saldos():
    cuentas = {
        "BS_BANESCO":"BS","BS_MERCANTIL":"BS",
        "CLP_COPEC_PAY":"CLP","CLP_BANCOESTADO":"CLP",
        "COP_EFECTIVO_ORLANDO":"COP","COP_BANCOLOMBIA_C1":"COP",
        "COP_BANCOLOMBIA_C2":"COP","COP_NEQUI_C1":"COP",
        "COP_NEQUI_C2":"COP","COP_NEQUI_C3":"COP",
        "USD_EFECTIVO":"USD","USDT_BINANCE":"USDT","USDC_AIRTM":"USDC",
    }
    conn = get_conn(); c = conn.cursor()
    for cuenta, moneda in cuentas.items():
        c.execute("INSERT OR IGNORE INTO saldos (cuenta,moneda,saldo) VALUES (?,?,0)", (cuenta, moneda))
    conn.commit(); conn.close()

def get_config(clave, default=""):
    conn = get_conn()
    row = conn.execute("SELECT valor FROM config WHERE clave=?", (clave,)).fetchone()
    conn.close()
    return row['valor'] if row else default

def set_config(clave, valor):
    conn = get_conn()
    conn.execute("INSERT INTO config (clave,valor,actualizado) VALUES (?,?,CURRENT_TIMESTAMP) ON CONFLICT(clave) DO UPDATE SET valor=excluded.valor, actualizado=CURRENT_TIMESTAMP",
                 (clave, str(valor)))
    conn.commit(); conn.close()

# ══════════════════════════════════════════════════════════════════════
# APIS BINANCE — PRECIOS REALES TOP 2
# ══════════════════════════════════════════════════════════════════════
def get_top_anuncios_bs(min_trans_ves=0):
    """Retorna top anuncios BS.
    min_trans_ves=0 para referencia (todos),
    min_trans_ves=1000 para Maker operable (acepta transacciones desde 1,000 Bs).
    El filtro es por monto mínimo de transacción en VES, no por USDT disponible.
    """
    url = "https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search"
    headers = {"Content-Type": "application/json"}

    def fetch_ads(side, pay_types=None):
        try:
            body = {
                "asset": "USDT", "fiat": "VES", "merchantCheck": False,
                "page": 1, "publisherType": None, "rows": 10,
                "tradeType": side, "payTypes": pay_types or []
            }
            # Si hay filtro de monto mínimo, lo pasamos como transAmount
            if min_trans_ves > 0:
                body["transAmount"] = str(min_trans_ves)
            r = requests.post(url, headers=headers, json=body, timeout=10)
            ads = r.json().get("data", [])
            result = []
            for a in ads[:2]:
                adv = a.get("adv", {})
                adv2 = a.get("advertiser", {})
                result.append({
                    "precio": float(adv.get("price", 0)),
                    "usuario": adv2.get("nickName", "—"),
                    "disponible": float(adv.get("surplusAmount", 0)),
                    "min_trans": float(adv.get("minSingleTransAmount", 0)),
                    "max_trans": float(adv.get("maxSingleTransAmount", 0)),
                })
            return result
        except: return []

    # SELL = ellos venden USDT, tú pagas Bs → solo Banesco/Mercantil
    # BUY  = ellos compran USDT, tú recibes Bs → Banesco/Mercantil/PagoMovil
    compras = fetch_ads("SELL", ["Banesco", "Mercantil"])
    ventas  = fetch_ads("BUY",  ["Banesco", "Mercantil", "PagoMovil"])
    return compras, ventas

def get_mercado_bs_completo():
    """Retorna referencia (todos) y Maker operable (acepta ≥1,000 Bs por transaccion)."""
    ref_compras, ref_ventas     = get_top_anuncios_bs(min_trans_ves=0)
    maker_compras, maker_ventas = get_top_anuncios_bs(min_trans_ves=1000)
    return {
        'ref_compras': ref_compras, 'ref_ventas': ref_ventas,
        'maker_compras': maker_compras, 'maker_ventas': maker_ventas,
    }

def get_top_anuncios_clp():
    """Retorna top 2 compradores Y vendedores de USDT en CLP."""
    url = "https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search"
    headers = {"Content-Type": "application/json"}

    def fetch(side):
        try:
            r = requests.post(url, headers=headers, json={
                "asset": "USDT", "fiat": "CLP", "merchantCheck": False,
                "page": 1, "publisherType": None, "rows": 5,
                "tradeType": side, "payTypes": []}, timeout=10)
            ads = r.json().get("data", [])
            result = []
            for a in ads[:2]:
                adv = a.get("adv", {})
                adv2 = a.get("advertiser", {})
                result.append({
                    "precio": float(adv.get("price", 0)),
                    "usuario": adv2.get("nickName", "—"),
                    "disponible": float(adv.get("surplusAmount", 0)),
                })
            return result
        except: return []

    # SELL = ellos venden USDT → tú compras USDT pagando CLP
    # BUY  = ellos compran USDT → tú vendes USDT recibiendo CLP
    compras_clp = fetch("SELL")  # precio más bajo = mejor para comprar
    ventas_clp  = fetch("BUY")   # precio más alto = mejor para vender
    return compras_clp, ventas_clp

def get_top_anuncios_cop():
    """Top 2 compradores COP."""
    url = "https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search"
    headers = {"Content-Type": "application/json"}
    try:
        r = requests.post(url, headers=headers, json={
            "asset": "USDT", "fiat": "COP", "merchantCheck": False,
            "page": 1, "publisherType": None, "rows": 5,
            "tradeType": "BUY", "payTypes": []}, timeout=10)
        ads = r.json().get("data", [])
        result = []
        for a in ads[:2]:
            adv = a.get("adv", {})
            adv2 = a.get("advertiser", {})
            result.append({
                "precio": float(adv.get("price", 0)),
                "usuario": adv2.get("nickName", "—"),
                "disponible": float(adv.get("surplusAmount", 0)),
            })
        return result
    except: return []

def get_binance_banco_promedio(banco):
    """Promedio de 3 anuncios que aceptan desde 1,000 Bs — para calcular tasas GSA y límites."""
    url = "https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search"
    headers = {"Content-Type": "application/json"}
    def fetch(side, pay):
        try:
            r = requests.post(url, headers=headers, json={
                "asset":"USDT","fiat":"VES","merchantCheck":False,"page":1,
                "publisherType":None,"rows":10,"tradeType":side,
                "payTypes":pay,"transAmount":"1000"}, timeout=10)
            ads = r.json()["data"][:3]
            prices = [float(a["adv"]["price"]) for a in ads]
            return round(sum(prices)/len(prices),2) if prices else None
        except: return None
    # Compra (SELL): solo Banesco o Mercantil según banco
    # Venta (BUY): banco + PagoMovil para capturar mejores precios de venta
    compra = fetch("SELL", [banco])
    venta  = fetch("BUY",  [banco, "PagoMovil"])
    spread = round(venta-compra,2) if venta and compra else 0
    return compra, venta, spread

def get_binance_fiat_promedio(fiat):
    url = "https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search"
    headers = {"Content-Type": "application/json"}
    def fetch(side):
        try:
            r = requests.post(url,headers=headers,json={
                "asset":"USDT","fiat":fiat,"merchantCheck":False,"page":1,
                "publisherType":None,"rows":10,"tradeType":side,"payTypes":[]},timeout=10)
            ads = r.json()["data"][:3]
            prices = [float(a["adv"]["price"]) for a in ads]
            return round(sum(prices)/len(prices),2) if prices else None
        except: return None
    return fetch("SELL"), fetch("BUY")

def get_dolar_observado():
    # Primero intentar mindicador.cl
    try:
        data = requests.get("https://mindicador.cl/api/dolar", timeout=10).json()
        valor = float(data["serie"][0]["valor"])
        if valor and valor > 500:  # Validar que sea un valor razonable
            set_config('ultimo_dolar_obs', str(valor))
            return valor
    except: pass

    # Fallback 1: usar último valor guardado en config
    try:
        ultimo = get_config('ultimo_dolar_obs', '')
        if ultimo:
            print(f"[dolar_obs] usando último valor guardado: {ultimo}")
            return float(ultimo)
    except: pass

    # Fallback 2: usar precio Binance CLP compra como referencia
    try:
        url = "https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search"
        r = requests.post(url, headers={"Content-Type": "application/json"}, json={
            "asset": "USDT", "fiat": "CLP", "merchantCheck": False,
            "page": 1, "publisherType": None, "rows": 3,
            "tradeType": "SELL", "payTypes": []}, timeout=10)
        ads = r.json().get("data", [])
        prices = [float(a["adv"]["price"]) for a in ads[:3]]
        if prices:
            precio_ref = round(sum(prices)/len(prices), 2)
            print(f"[dolar_obs] usando Binance CLP como referencia: {precio_ref}")
            return precio_ref
    except: pass

    return None

def get_trm():
    try:
        ayer = (datetime.date.today()-datetime.timedelta(days=5)).strftime("%Y-%m-%d")
        url = f"https://www.datos.gov.co/resource/32sa-8pi3.json?$where=vigenciadesde>='{ayer}T00:00:00.000'&$order=vigenciadesde DESC&$limit=1"
        data = requests.get(url,timeout=10).json()
        return float(data[0]["valor"]) if data else None
    except: return None

def get_bcv():
    try:
        data = requests.get("https://ve.dolarapi.com/v1/dolares/oficiales",timeout=10).json()
        usd = eur = None
        for item in data:
            m = item.get("moneda","").lower()
            if m=="usd": usd=float(item.get("promedio",0))
            elif m=="eur": eur=float(item.get("promedio",0))
        return usd, eur
    except: return None, None

# ══════════════════════════════════════════════════════════════════════
# GUARDAR HISTORIAL DE PRECIOS
# ══════════════════════════════════════════════════════════════════════
def guardar_precio_historico(compras_bs, ventas_bs, clp_ads, cop_ads):
    """Guarda snapshot de precios en precios_historicos."""
    ahora = now_local()
    fecha = str(ahora.date())
    hora  = ahora.strftime("%H:%M")
    dia   = DIAS_SEMANA[ahora.weekday()]

    bs_c1 = compras_bs[0]['precio'] if len(compras_bs) > 0 else None
    bs_c2 = compras_bs[1]['precio'] if len(compras_bs) > 1 else None
    bs_v1 = ventas_bs[0]['precio']  if len(ventas_bs)  > 0 else None
    bs_v2 = ventas_bs[1]['precio']  if len(ventas_bs)  > 1 else None
    spread = round(bs_v1 - bs_c1, 2) if bs_v1 and bs_c1 else None

    datos = {
        'fecha': fecha, 'hora': hora, 'dia_semana': dia,
        'bs_venta_1': bs_v1, 'bs_venta_2': bs_v2,
        'bs_compra_1': bs_c1, 'bs_compra_2': bs_c2,
        'bs_spread_real': spread,
        'clp_compra_1': clp_ads[0]['precio'] if len(clp_ads) > 0 else None,
        'clp_compra_2': clp_ads[1]['precio'] if len(clp_ads) > 1 else None,
        'cop_compra_1': cop_ads[0]['precio'] if len(cop_ads) > 0 else None,
        'cop_compra_2': cop_ads[1]['precio'] if len(cop_ads) > 1 else None,
        'usuario_venta_1': ventas_bs[0]['usuario']  if len(ventas_bs)  > 0 else None,
        'usuario_venta_2': ventas_bs[1]['usuario']  if len(ventas_bs)  > 1 else None,
        'usuario_compra_1': compras_bs[0]['usuario'] if len(compras_bs) > 0 else None,
        'usuario_compra_2': compras_bs[1]['usuario'] if len(compras_bs) > 1 else None,
        'disp_venta_1': ventas_bs[0]['disponible']  if len(ventas_bs)  > 0 else None,
        'disp_venta_2': ventas_bs[1]['disponible']  if len(ventas_bs)  > 1 else None,
        'disp_compra_1': compras_bs[0]['disponible'] if len(compras_bs) > 0 else None,
        'disp_compra_2': compras_bs[1]['disponible'] if len(compras_bs) > 1 else None,
    }

    conn = get_conn()
    conn.execute("""INSERT INTO precios_historicos
        (fecha,hora,dia_semana,bs_venta_1,bs_venta_2,bs_compra_1,bs_compra_2,
         bs_spread_real,clp_compra_1,clp_compra_2,cop_compra_1,cop_compra_2,
         usuario_venta_1,usuario_venta_2,usuario_compra_1,usuario_compra_2,
         disp_venta_1,disp_venta_2,disp_compra_1,disp_compra_2)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (datos['fecha'],datos['hora'],datos['dia_semana'],
         datos['bs_venta_1'],datos['bs_venta_2'],datos['bs_compra_1'],datos['bs_compra_2'],
         datos['bs_spread_real'],datos['clp_compra_1'],datos['clp_compra_2'],
         datos['cop_compra_1'],datos['cop_compra_2'],
         datos['usuario_venta_1'],datos['usuario_venta_2'],
         datos['usuario_compra_1'],datos['usuario_compra_2'],
         datos['disp_venta_1'],datos['disp_venta_2'],
         datos['disp_compra_1'],datos['disp_compra_2']))
    conn.commit(); conn.close()

    if USE_SUPABASE:
        supa_insert('precios_historicos', datos)

    return datos

# ══════════════════════════════════════════════════════════════════════
# ANÁLISIS DE PATRONES
# ══════════════════════════════════════════════════════════════════════
def analizar_patron_bs():
    """Analiza el historial para detectar patrones de precio BS."""
    conn = get_conn()
    rows = conn.execute("""
        SELECT hora, dia_semana, bs_venta_1, bs_compra_1, bs_spread_real
        FROM precios_historicos
        WHERE bs_venta_1 IS NOT NULL AND bs_compra_1 IS NOT NULL
        ORDER BY fecha DESC, hora
    """).fetchall()
    conn.close()

    if len(rows) < 20:
        return None, "Datos insuficientes. El bot necesita al menos 1 día de historial."

    # Agrupar por hora
    por_hora = defaultdict(list)
    for r in rows:
        hora_bloque = r['hora'][:2] + ":00"
        por_hora[hora_bloque].append({
            'venta': r['bs_venta_1'],
            'compra': r['bs_compra_1'],
            'spread': r['bs_spread_real'] or 0
        })

    resumen = {}
    for hora, datos in sorted(por_hora.items()):
        ventas = [d['venta'] for d in datos]
        spreads = [d['spread'] for d in datos]
        resumen[hora] = {
            'venta_prom': round(sum(ventas)/len(ventas), 2),
            'spread_prom': round(sum(spreads)/len(spreads), 2),
            'venta_max': max(ventas),
            'venta_min': min(ventas),
        }

    # Mejor y peor hora
    mejor_hora = max(resumen.items(), key=lambda x: x[1]['spread_prom'])
    peor_hora  = min(resumen.items(), key=lambda x: x[1]['spread_prom'])

    return resumen, mejor_hora, peor_hora

def proyeccion_manana():
    """Proyecta precios para mañana basado en mismo día de la semana."""
    manana = now_local() + datetime.timedelta(days=1)
    dia_manana = DIAS_SEMANA[manana.weekday()]

    conn = get_conn()
    rows = conn.execute("""
        SELECT hora, bs_venta_1, bs_spread_real
        FROM precios_historicos
        WHERE dia_semana=? AND bs_venta_1 IS NOT NULL
        ORDER BY fecha DESC, hora
    """, (dia_manana,)).fetchall()
    conn.close()

    if len(rows) < 5:
        return None, dia_manana

    por_hora = defaultdict(list)
    for r in rows:
        hora_bloque = r['hora'][:2] + ":00"
        por_hora[hora_bloque].append({
            'venta': r['bs_venta_1'],
            'spread': r['bs_spread_real'] or 0
        })

    proyeccion = {}
    for hora, datos in sorted(por_hora.items()):
        ventas = [d['venta'] for d in datos]
        spreads = [d['spread'] for d in datos]
        proyeccion[hora] = {
            'venta_esperada': round(sum(ventas)/len(ventas), 2),
            'spread_esperado': round(sum(spreads)/len(spreads), 2),
            'min': min(ventas),
            'max': max(ventas),
        }

    return proyeccion, dia_manana

# ══════════════════════════════════════════════════════════════════════
# GESTOR DE CAPITAL INTELIGENTE
# ══════════════════════════════════════════════════════════════════════
def analizar_capital():
    """Analiza el capital disponible vs el histórico y sugiere acciones."""
    saldos = get_saldos()
    t = get_ultima_tasa()
    ahora = now_local()
    hora_actual = ahora.hour
    dia_actual = DIAS_SEMANA[ahora.weekday()]

    pat_bs  = ((t.get('ban_bs_compra',0) or 0) + (t.get('ban_bs_venta',0) or 0)) / 2 or 1
    dol_obs = t.get('dolar_obs',1) or 1
    trm     = t.get('trm',1) or 1

    # Saldos actuales
    bs_total  = (saldos.get('BS_BANESCO',{}).get('saldo',0) or 0) + (saldos.get('BS_MERCANTIL',{}).get('saldo',0) or 0)
    clp_total = saldos.get('CLP_COPEC_PAY',{}).get('saldo',0) or 0
    cop_total = (saldos.get('COP_EFECTIVO_ORLANDO',{}).get('saldo',0) or 0)
    usdt_total = saldos.get('USDT_BINANCE',{}).get('saldo',0) or 0

    # Convertir todo a USDT
    bs_usdt   = bs_total / pat_bs if pat_bs else 0
    clp_usdt  = clp_total / dol_obs if dol_obs else 0
    cop_usdt  = cop_total / trm if trm else 0
    patrimonio_usdt = bs_usdt + clp_usdt + cop_usdt + usdt_total

    alertas = []

    # Revisar umbrales configurados
    conn = get_conn()
    umbrales = conn.execute("SELECT * FROM umbrales_liquidez WHERE activo=1").fetchall()
    conn.close()

    for u in umbrales:
        saldo_cuenta = saldos.get(u['cuenta'],{}).get('saldo',0) or 0
        if saldo_cuenta < u['umbral_minimo']:
            nombre = NOMBRES_CUENTAS.get(u['cuenta'], u['cuenta'])
            alertas.append({
                'tipo': 'umbral',
                'cuenta': nombre,
                'saldo': saldo_cuenta,
                'umbral': u['umbral_minimo'],
                'moneda': u['moneda'],
            })

    # Análisis de desbalance
    if patrimonio_usdt > 10:
        bs_pct   = bs_usdt / patrimonio_usdt * 100
        clp_pct  = clp_usdt / patrimonio_usdt * 100
        usdt_pct = usdt_total / patrimonio_usdt * 100

        if clp_pct < 10 and hora_actual >= 8 and hora_actual <= 20:
            alertas.append({
                'tipo': 'desbalance',
                'mensaje': f'Solo {clp_pct:.0f}% del capital en CLP. Operaciones CLP limitadas.',
                'sugerencia': f'Considera convertir USDT→CLP o BS→CLP'
            })
        if bs_pct < 10 and hora_actual >= 8 and hora_actual <= 20:
            alertas.append({
                'tipo': 'desbalance',
                'mensaje': f'Solo {bs_pct:.0f}% del capital en BS. Arbitraje BS limitado.',
                'sugerencia': f'Considera inyectar BS o convertir USDT→BS'
            })

    # Actividad esperada basada en historial
    conn = get_conn()
    ops_historico = conn.execute("""
        SELECT COUNT(*) as cnt FROM operaciones
        WHERE strftime('%H', hora) = ? AND dia_semana_local = ?
        AND estado='Completada'
    """, (str(hora_actual).zfill(2), dia_actual)).fetchone()
    conn.close()

    return {
        'bs_total': bs_total,
        'clp_total': clp_total,
        'cop_total': cop_total,
        'usdt_total': usdt_total,
        'patrimonio_usdt': patrimonio_usdt,
        'alertas': alertas,
    }

def msg_capital():
    """Genera mensaje del análisis de capital."""
    c = analizar_capital()
    t = get_ultima_tasa()
    pat_bs = ((t.get('ban_bs_compra',0) or 0) + (t.get('ban_bs_venta',0) or 0)) / 2 or 1
    dol_obs = t.get('dolar_obs',1) or 1

    m = f"💼 *ANÁLISIS DE CAPITAL*\n📅 {now_local().strftime('%d/%m %I:%M %p')}\n━━━━━━━━━━━━━━━━━━━━\n\n"
    m += f"🇻🇪 BS disponible: `{c['bs_total']:,.0f} Bs` (~`{c['bs_total']/pat_bs:.2f} USDT`)\n"
    m += f"🇨🇱 CLP disponible: `{c['clp_total']:,.0f} CLP` (~`{c['clp_total']/dol_obs:.2f} USDT`)\n"
    m += f"💎 USDT disponible: `{c['usdt_total']:.4f} USDT`\n"
    m += f"🏛️ Patrimonio total: `{c['patrimonio_usdt']:.2f} USDT`\n\n"

    if c['alertas']:
        m += "⚠️ *ALERTAS DE CAPITAL:*\n\n"
        for a in c['alertas']:
            if a['tipo'] == 'umbral':
                m += f"🔴 *{a['cuenta']}*: `{a['saldo']:,.2f}` (mín: `{a['umbral']:,.2f} {a['moneda']}`)\n"
            elif a['tipo'] == 'desbalance':
                m += f"🟡 {a['mensaje']}\n   _{a['sugerencia']}_\n"
        m += "\n"
    else:
        m += "✅ Capital balanceado\n\n"

    return m

# ══════════════════════════════════════════════════════════════════════
# SIMULADOR DE OPERACIONES
# ══════════════════════════════════════════════════════════════════════
def simular_operacion(moneda_origen, monto):
    """Simula cuánto ganas operando X monto ahora mismo."""
    t = get_ultima_tasa()
    compras_bs, ventas_bs = get_top_anuncios_bs()
    clp_ads = get_top_anuncios_clp()

    pat_bs = ((t.get('ban_bs_compra',0) or 0) + (t.get('ban_bs_venta',0) or 0)) / 2 or 1
    dol_obs = t.get('dolar_obs',1) or 1

    m = f"🧮 *SIMULADOR GSA*\n━━━━━━━━━━━━━━━━━━━━\n\n"
    m += f"Capital: `{monto:,.2f} {moneda_origen}`\n\n"

    if moneda_origen == 'CLP':
        precio_compra_clp = clp_ads[0]['precio'] if clp_ads else dol_obs
        usdt = monto / precio_compra_clp
        m += f"🔵 *CLP → USDT (Binance Maker)*\n"
        m += f"   Precio compra: `{precio_compra_clp:,.2f} CLP`\n"
        m += f"   Recibes: `{usdt:.4f} USDT`\n\n"

        if ventas_bs:
            precio_venta_bs = ventas_bs[0]['precio']
            bs_recibidos = usdt * precio_venta_bs * (1 - FEE_USDT_BS)
            clp_recuperado = bs_recibidos / pat_bs * dol_obs
            ganancia_clp = clp_recuperado - monto
            ganancia_usdt = ganancia_clp / dol_obs
            m += f"🔵 *USDT → BS → CLP (triangular)*\n"
            m += f"   Precio venta BS: `{precio_venta_bs:,.2f} Bs`\n"
            m += f"   BS recibidos: `{bs_recibidos:,.0f} Bs`\n"
            m += f"   CLP recuperados: `{clp_recuperado:,.0f} CLP`\n"
            rentable = "✅ RENTABLE" if ganancia_clp > 0 else "❌ NO RENTABLE"
            m += f"\n💰 *Ganancia: `{ganancia_clp:,.0f} CLP` (~`{ganancia_usdt:.2f} USDT`)* {rentable}\n"

    elif moneda_origen == 'BS':
        precio_compra_bs = compras_bs[0]['precio'] if compras_bs else pat_bs
        usdt = monto / precio_compra_bs * (1 - FEE_USDT_BS)
        m += f"🔵 *BS → USDT (Binance Maker)*\n"
        m += f"   Precio compra: `{precio_compra_bs:,.2f} Bs`\n"
        m += f"   Recibes: `{usdt:.4f} USDT`\n\n"

        if ventas_bs:
            precio_venta_bs = ventas_bs[0]['precio']
            bs_recuperados = usdt * precio_venta_bs
            ganancia_bs = bs_recuperados - monto
            ganancia_usdt = ganancia_bs / precio_compra_bs
            m += f"🔵 *USDT → BS (venta)*\n"
            m += f"   Precio venta: `{precio_venta_bs:,.2f} Bs`\n"
            m += f"   BS recuperados: `{bs_recuperados:,.0f} Bs`\n"
            rentable = "✅ RENTABLE" if ganancia_bs > 0 else "❌ NO RENTABLE"
            m += f"\n💰 *Ganancia: `{ganancia_bs:,.0f} Bs` (~`{ganancia_usdt:.4f} USDT`)* {rentable}\n"

    elif moneda_origen == 'USDT':
        m += f"*Con `{monto:.4f} USDT` puedes:*\n\n"
        if ventas_bs:
            bs = monto * ventas_bs[0]['precio'] * (1-FEE_USDT_BS)
            m += f"🇻🇪 Vender en BS: `{bs:,.0f} Bs` (@{ventas_bs[0]['precio']} Bs)\n"
        if clp_ads:
            clp = monto * clp_ads[0]['precio']
            m += f"🇨🇱 Vender en CLP: `{clp:,.0f} CLP` (@{clp_ads[0]['precio']} CLP)\n"

    m += f"\n_Precios en tiempo real — {now_local().strftime('%H:%M')}_"
    return m

# ══════════════════════════════════════════════════════════════════════
# MENSAJES DE MERCADO
# ══════════════════════════════════════════════════════════════════════
def msg_mercado():
    """Mensaje con top 2 precios reales BS y CLP."""
    compras_bs, ventas_bs = get_top_anuncios_bs()
    clp_ads = get_top_anuncios_clp()
    ahora = now_local().strftime("%d/%m %I:%M %p")

    m = f"📡 *MERCADO EN VIVO*\n📅 {ahora}\n━━━━━━━━━━━━━━━━━━━━\n\n"

    m += "🇻🇪 *BINANCE BS*\n\n"
    if compras_bs:
        m += "📥 *Compra (pagas Bs, recibes USDT):*\n"
        for i, a in enumerate(compras_bs, 1):
            m += f"  {i}️⃣ `{a['usuario']:15s}` `{a['precio']:,.2f} Bs` | `{a['disponible']:.2f} USDT` disp.\n"
    if ventas_bs:
        m += "\n📤 *Venta (vendes USDT, recibes Bs):*\n"
        for i, a in enumerate(ventas_bs, 1):
            m += f"  {i}️⃣ `{a['usuario']:15s}` `{a['precio']:,.2f} Bs` | `{a['disponible']:.2f} USDT` disp.\n"
    if compras_bs and ventas_bs:
        spread = ventas_bs[0]['precio'] - compras_bs[0]['precio']
        emoji = "🚀" if spread >= SPREAD_PREMIUM else "🟢" if spread >= SPREAD_BUENO else "🟡" if spread >= SPREAD_MODERADO else "🔴"
        m += f"\n  Spread real: `{spread:.2f} Bs` {emoji}\n"
        m += f"  💡 Tu precio de venta sugerido: `{ventas_bs[0]['precio']+1:.2f} Bs`\n"

    m += "\n━━━━━━━━━━━━━━━━━━━━\n"
    m += "🇨🇱 *BINANCE CLP*\n\n"
    
    compras_clp, ventas_clp = clp_ads if isinstance(clp_ads, tuple) else ([], clp_ads)
    
    if compras_clp:
        m += "📥 *Compra USDT (pagas CLP, recibes USDT):*\n"
        for i, a in enumerate(compras_clp, 1):
            m += f"  {i}️⃣ `{a['usuario']:15s}` `{a['precio']:,.2f} CLP` | `{a['disponible']:.2f} USDT` disp.\n"
        m += f"  💡 Mejor precio de compra: `{compras_clp[0]['precio']:,.2f} CLP`\n\n"

    if ventas_clp:
        m += "📤 *Venta USDT (entregas USDT, recibes CLP):*\n"
        for i, a in enumerate(ventas_clp, 1):
            m += f"  {i}️⃣ `{a['usuario']:15s}` `{a['precio']:,.2f} CLP` | `{a['disponible']:.2f} USDT` disp.\n"
        m += f"  💡 Para ser competitivo: publica a `{ventas_clp[0]['precio']+1:.2f} CLP`\n"

    return m

def msg_alerta_bs(compras_bs, ventas_bs, spread, precio_compra=0, precio_venta=0):
    """Mensaje de alerta cuando el spread BS supera el umbral."""
    if spread >= SPREAD_PREMIUM: emoji, nivel = "🚀", "PREMIUM"
    elif spread >= SPREAD_BUENO: emoji, nivel = "🟢", "BUENO"
    else: emoji, nivel = "🟡", "MODERADO"

    m = f"{emoji} *SEÑAL BS — {nivel}* | Spread: `{spread:.2f} Bs`\n"
    # Mostrar precios promedio del mercado (fuente del spread)
    if precio_compra and precio_venta:
        m += f"Compra promedio: `{precio_compra:.2f} Bs` | Venta promedio: `{precio_venta:.2f} Bs`\n\n"

    # Mostrar anuncios reales disponibles
    if compras_bs:
        m += "📥 *Compra (pagas Bs, recibes USDT):*\n"
        for i, a in enumerate(compras_bs, 1):
            m += f"  {i}️⃣ `{a['usuario']:12s}` `{a['precio']:,.2f} Bs` | `{a['disponible']:.1f} USDT`\n"
    if ventas_bs:
        m += "\n📤 *Venta (vendes USDT, recibes Bs):*\n"
        for i, a in enumerate(ventas_bs, 1):
            m += f"  {i}️⃣ `{a['usuario']:12s}` `{a['precio']:,.2f} Bs` | `{a['disponible']:.1f} USDT`\n"
        m += f"\n💡 Publica venta a: `{precio_venta+1:.2f} Bs`"
    return m

def msg_alerta_clp(compras_clp, ventas_clp):
    """Mensaje de alerta cuando el precio CLP cambia."""
    m = f"🇨🇱 *MERCADO CLP ACTUALIZADO*\n\n"
    if compras_clp:
        m += "📥 *Compra USDT (pagas CLP):*\n"
        for i, a in enumerate(compras_clp, 1):
            m += f"  {i}️⃣ `{a['usuario']:12s}` `{a['precio']:,.2f} CLP` | `{a['disponible']:.1f} USDT`\n"
        m += f"  💡 Mejor precio compra: `{compras_clp[0]['precio']:,.2f} CLP`\n\n"
    if ventas_clp:
        m += "📤 *Venta USDT (recibes CLP):*\n"
        for i, a in enumerate(ventas_clp, 1):
            m += f"  {i}️⃣ `{a['usuario']:12s}` `{a['precio']:,.2f} CLP` | `{a['disponible']:.1f} USDT`\n"
        m += f"  💡 Para ser competitivo: publica a `{ventas_clp[0]['precio']+1:.2f} CLP`"
    return m

def msg_alerta_triangular(monto_clp, clp_ads, ventas_bs, t):
    """Alerta cuando hay oportunidad de arbitraje triangular."""
    if not clp_ads or not ventas_bs: return None
    precio_clp = clp_ads[0]['precio']
    precio_venta_bs = ventas_bs[0]['precio']
    pat_bs = ((t.get('ban_bs_compra',0) or 0) + (t.get('ban_bs_venta',0) or 0)) / 2 or 1
    dol_obs = t.get('dolar_obs',1) or 1

    usdt = monto_clp / precio_clp
    bs_recibidos = usdt * precio_venta_bs * (1-FEE_USDT_BS)
    clp_recuperado = bs_recibidos / pat_bs * dol_obs
    ganancia = clp_recuperado - monto_clp
    pct = (ganancia / monto_clp) * 100

    if ganancia <= 0 or pct < 1: return None

    m = f"🔺 *ARBITRAJE TRIANGULAR DETECTADO*\n\n"
    m += f"`{monto_clp:,.0f} CLP → {usdt:.2f} USDT → {bs_recibidos:,.0f} Bs → {clp_recuperado:,.0f} CLP`\n\n"
    m += f"💰 Ganancia estimada: `{ganancia:,.0f} CLP` (`{pct:.1f}%`)\n"
    m += f"_Usa /simular CLP {monto_clp:.0f} para confirmar_"
    return m

# ══════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════
# GESTIÓN DE SESIÓN DE ARBITRAJE
# ══════════════════════════════════════════════════════════════════════
sesion_activa = {}
historial_spread_reciente = []

def iniciar_sesion_arbitraje(chat_id, moneda, capital):
    sesion_activa[chat_id] = {
        'moneda': moneda, 'capital_inicial': capital,
        'capital_disponible': capital, 'usdt_comprado': 0,
        'usdt_vendido': 0, 'bs_pagado': 0, 'bs_recibido': 0,
        'cpp': 0, 'ganancia_realizada': 0,
        'inicio': now_local(), 'activa': True,
    }
    return f"✅ Sesión iniciada\nMoneda: `{moneda}` | Capital: `{capital:,.0f}`\nUsa /sesion status para ver el estado."

def registrar_compra_sesion(chat_id, usdt, precio_bs):
    if chat_id not in sesion_activa: return "Sin sesión activa."
    s = sesion_activa[chat_id]
    bs_pagados = usdt * precio_bs
    total_usdt = s['usdt_comprado'] + usdt
    total_bs = (s['usdt_comprado'] * s['cpp']) + bs_pagados
    s['cpp'] = total_bs / total_usdt if total_usdt > 0 else precio_bs
    s['usdt_comprado'] = total_usdt
    s['bs_pagado'] += bs_pagados
    s['capital_disponible'] -= bs_pagados
    return f"✅ Compra registrada\n`{usdt:.4f} USDT` a `{precio_bs:.2f} Bs`\nNuevo CPP: `{s['cpp']:.2f} Bs`"

def registrar_venta_sesion(chat_id, usdt, precio_bs):
    if chat_id not in sesion_activa: return "Sin sesión activa."
    s = sesion_activa[chat_id]
    bs_recibidos = usdt * precio_bs
    ganancia = usdt * (precio_bs - s['cpp'])
    s['usdt_vendido'] += usdt
    s['bs_recibido'] += bs_recibidos
    s['ganancia_realizada'] += ganancia
    s['capital_disponible'] += bs_recibidos
    return f"✅ Venta registrada\n`{usdt:.4f} USDT` a `{precio_bs:.2f} Bs`\nGanancia esta venta: `{ganancia:.2f} Bs`"

def msg_sesion_status(chat_id):
    if chat_id not in sesion_activa:
        return "Sin sesión activa.\nUsa `/sesion BS 200000` para iniciar."
    s = sesion_activa[chat_id]
    compras_bs, ventas_bs = get_top_anuncios_bs(min_trans_ves=1000)
    duracion = int((now_local() - s['inicio']).total_seconds() / 60)
    spread_actual = (ventas_bs[0]['precio'] - compras_bs[0]['precio']) if compras_bs and ventas_bs else 0
    ganancia_latente = 0
    if s['usdt_comprado'] > s['usdt_vendido'] and s['cpp'] > 0 and ventas_bs:
        usdt_pend = s['usdt_comprado'] - s['usdt_vendido']
        ganancia_latente = usdt_pend * (ventas_bs[0]['precio'] - s['cpp'])
    m = f"🔴 *SESIÓN ACTIVA — {duracion} min*\n━━━━━━━━━━━━━━━━━━━━\n\n"
    m += f"Comprado: `{s['usdt_comprado']:.4f} USDT` | CPP: `{s['cpp']:.2f} Bs`\n"
    m += f"Vendido:  `{s['usdt_vendido']:.4f} USDT`\n"
    pend = s['usdt_comprado'] - s['usdt_vendido']
    if pend > 0: m += f"Pendiente: `{pend:.4f} USDT`\n"
    m += f"\nGanancia realizada: `{s['ganancia_realizada']:.2f} Bs`\n"
    if ganancia_latente: m += f"Ganancia latente: `~{ganancia_latente:.2f} Bs`\n"
    emoji_s = "🟢" if spread_actual >= SPREAD_MIN_ALERTA else "🔴"
    m += f"\nSpread Maker actual: `{spread_actual:.2f} Bs` {emoji_s}\n"
    if spread_actual < SPREAD_MIN_ALERTA and spread_actual > 0:
        m += f"\n⚠️ *Spread bajo umbral — considera pausar anuncio en Binance*"
    elif spread_actual <= 0:
        m += f"\n🚨 *Spread negativo — PAUSA el anuncio YA*"
    return m

def cerrar_sesion(chat_id):
    if chat_id not in sesion_activa: return "Sin sesión activa."
    s = sesion_activa[chat_id]
    duracion = int((now_local() - s['inicio']).total_seconds() / 60)
    m = f"✅ *SESIÓN CERRADA*\n━━━━━━━━━━━━━━━━━━━━\n"
    m += f"Duración: `{duracion} min`\n"
    m += f"USDT comprado: `{s['usdt_comprado']:.4f}`\n"
    m += f"USDT vendido: `{s['usdt_vendido']:.4f}`\n"
    m += f"CPP final: `{s['cpp']:.2f} Bs`\n"
    m += f"Ganancia total: `{s['ganancia_realizada']:.2f} Bs`\n"
    pend = s['usdt_comprado'] - s['usdt_vendido']
    if pend > 0.01: m += f"⚠️ USDT en inventario: `{pend:.4f} USDT`\n"
    del sesion_activa[chat_id]
    return m

# ══════════════════════════════════════════════════════════════════════
# DETECTOR DE TENDENCIA Y MOMENTO ÓPTIMO
# ══════════════════════════════════════════════════════════════════════
def analizar_tendencia_spread(spread_actual):
    global historial_spread_reciente
    historial_spread_reciente.append({'spread': spread_actual, 'hora': now_local()})
    if len(historial_spread_reciente) > 6:
        historial_spread_reciente = historial_spread_reciente[-6:]
    if len(historial_spread_reciente) < 3: return None
    ultimos = [h['spread'] for h in historial_spread_reciente[-3:]]
    if ultimos[0] < ultimos[1] < ultimos[2]: return 'subiendo'
    if ultimos[0] > ultimos[1] > ultimos[2]: return 'bajando'
    return 'estable'

def msg_momento_optimo(maker_compras, maker_ventas, spread, tendencia):
    if not maker_compras or not maker_ventas: return None
    if spread < SPREAD_MIN_ALERTA or tendencia != 'subiendo': return None
    ultimos = [h['spread'] for h in historial_spread_reciente[-3:]]
    precio_entrada = maker_compras[0]['precio'] + 1
    precio_min_venta = precio_entrada + 8
    m = f"⚡ *MOMENTO ÓPTIMO DETECTADO*\n\n"
    m += f"Spread Maker subiendo:\n"
    for i, s in enumerate(ultimos):
        m += f"  `{(len(ultimos)-1-i)*5:2d} min atrás → {s:.1f} Bs`\n"
    m += f"\n💡 *PRECIO SUGERIDO PARA ENTRAR:*\n"
    m += f"  Publica compra a: `{precio_entrada:.2f} Bs`\n"
    m += f"  Venta mínima rentable: `{precio_min_venta:.2f} Bs`\n"
    m += f"  Spread actual: `{spread:.1f} Bs` 🟢\n\n"
    m += f"_/sesion BS MONTO para activar seguimiento_"
    return m

def msg_alerta_spread_cayendo(spread_actual):
    if len(historial_spread_reciente) < 3: return None
    ultimos = [h['spread'] for h in historial_spread_reciente[-3:]]
    if not (ultimos[0] > ultimos[1] > ultimos[2]): return None
    if spread_actual >= SPREAD_MIN_ALERTA: return None
    m = f"⚠️ *SPREAD CAYENDO — CONSIDERA PAUSAR*\n\n"
    for i, s in enumerate(ultimos):
        m += f"  `{(len(ultimos)-1-i)*5:2d} min atrás → {s:.1f} Bs`\n"
    m += f"\n❌ Spread bajo umbral mínimo ({SPREAD_MIN_ALERTA} Bs)\n"
    m += f"💡 Pausa tu anuncio en Binance manualmente"
    return m

# ══════════════════════════════════════════════════════════════════════
# MERCADO COMPLETO CON MAKER FILTER
# ══════════════════════════════════════════════════════════════════════
def msg_mercado_completo():
    mercado = get_mercado_bs_completo()
    compras_clp, ventas_clp = get_top_anuncios_clp()
    ahora = now_local().strftime("%d/%m %I:%M %p")
    m = f"📡 *MERCADO EN VIVO*\n📅 {ahora}\n━━━━━━━━━━━━━━━━━━━━\n\n"
    m += "🇻🇪 *BINANCE BS*\n\n"
    ref_c = mercado['ref_compras']; ref_v = mercado['ref_ventas']
    if ref_c or ref_v:
        m += "📊 *Precio referencia:*\n"
        if ref_c: m += f"  Compra: `{ref_c[0]['precio']:,.2f} Bs` ({ref_c[0]['usuario']})\n"
        if ref_v: m += f"  Venta:  `{ref_v[0]['precio']:,.2f} Bs` ({ref_v[0]['usuario']})\n"
        if ref_c and ref_v:
            sr = ref_v[0]['precio'] - ref_c[0]['precio']
            m += f"  Spread: `{sr:.2f} Bs` {spread_emoji(sr)}\n"
        m += "\n"
    mk_c = mercado['maker_compras']; mk_v = mercado['maker_ventas']
    m += "💰 *Maker operable (≥1,000 USDT):*\n"
    if mk_c:
        m += "  📥 Compra:\n"
        for i, a in enumerate(mk_c, 1):
            min_t = f" | mín {a['min_trans']:,.0f} Bs" if a.get('min_trans') else ""
            m += f"    {i}️⃣ `{a['usuario']:15s}` `{a['precio']:,.2f} Bs` | `{a['disponible']:,.1f} USDT`{min_t}\n"
    else:
        m += "  📥 Sin anuncios compra con límite ≥1,000 Bs\n"
    if mk_v:
        m += "  📤 Venta:\n"
        for i, a in enumerate(mk_v, 1):
            min_t = f" | mín {a['min_trans']:,.0f} Bs" if a.get('min_trans') else ""
            m += f"    {i}️⃣ `{a['usuario']:15s}` `{a['precio']:,.2f} Bs` | `{a['disponible']:,.1f} USDT`{min_t}\n"
    else:
        m += "  📤 Sin anuncios venta con límite ≥1,000 Bs\n"
    if mk_c and mk_v:
        spread_m = mk_v[0]['precio'] - mk_c[0]['precio']
        emoji_m = "🚀" if spread_m >= SPREAD_PREMIUM else "🟢" if spread_m >= SPREAD_BUENO else "🟡" if spread_m >= SPREAD_MODERADO else "🔴"
        m += f"\n  Spread Maker: `{spread_m:.2f} Bs` {emoji_m}\n"
        if spread_m >= SPREAD_MIN_ALERTA:
            precio_entrada = mk_c[0]['precio'] + 1
            m += f"  💡 Entrada: publica compra a `{precio_entrada:.2f} Bs`\n"
            m += f"  💡 Venta mínima rentable: `{precio_entrada + 8:.2f} Bs`\n"
            m += f"  ✅ OPERABLE\n"
        else:
            m += f"  ❌ Spread insuficiente\n"
    m += "\n━━━━━━━━━━━━━━━━━━━━\n🇨🇱 *BINANCE CLP*\n\n"
    if compras_clp:
        m += "📥 *Compra USDT (pagas CLP):*\n"
        for i, a in enumerate(compras_clp, 1):
            m += f"  {i}️⃣ `{a['usuario']:15s}` `{a['precio']:,.2f} CLP` | `{a['disponible']:.1f} USDT`\n"
        m += f"  💡 Mejor compra: `{compras_clp[0]['precio']:,.2f} CLP`\n\n"
    if ventas_clp:
        m += "📤 *Venta USDT (recibes CLP):*\n"
        for i, a in enumerate(ventas_clp, 1):
            m += f"  {i}️⃣ `{a['usuario']:15s}` `{a['precio']:,.2f} CLP` | `{a['disponible']:.1f} USDT`\n"
        m += f"  💡 Competitivo: `{ventas_clp[0]['precio']+1:.2f} CLP`\n"
    return m

# REPORTES DIARIO Y SEMANAL
# ══════════════════════════════════════════════════════════════════════
def generar_reporte_diario():
    """Genera el reporte de análisis del día."""
    hoy = str(today_local())
    conn = get_conn()
    rows = conn.execute("""
        SELECT hora, bs_venta_1, bs_compra_1, bs_spread_real
        FROM precios_historicos
        WHERE fecha=? AND bs_venta_1 IS NOT NULL
        ORDER BY hora
    """, (hoy,)).fetchall()
    conn.close()

    if not rows:
        return "📊 Sin datos de mercado para analizar hoy."

    ventas = [(r['hora'], r['bs_venta_1']) for r in rows]
    spreads = [(r['hora'], r['bs_spread_real'] or 0) for r in rows]

    max_venta = max(ventas, key=lambda x: x[1])
    min_venta = min(ventas, key=lambda x: x[1])
    max_spread = max(spreads, key=lambda x: x[1])
    min_spread = min(spreads, key=lambda x: x[1])

    prom_venta = sum(v for _,v in ventas) / len(ventas)
    prom_spread = sum(s for _,s in spreads) / len(spreads)

    # Mejor ventana (hora con mayor spread promedio)
    por_hora = defaultdict(list)
    for r in rows:
        por_hora[r['hora'][:2]].append(r['bs_spread_real'] or 0)
    mejor_ventana = max(por_hora.items(), key=lambda x: sum(x[1])/len(x[1]))

    res_hoy = get_resultados_hoy()

    m = f"📊 *ANÁLISIS DIARIO — {now_local().strftime('%d/%m/%Y')}*\n"
    m += f"━━━━━━━━━━━━━━━━━━━━\n\n"
    m += f"📈 *PRECIO BS HOY*\n"
    m += f"  Venta más alta: `{max_venta[1]:.2f} Bs` a las `{max_venta[0]}`\n"
    m += f"  Venta más baja: `{min_venta[1]:.2f} Bs` a las `{min_venta[0]}`\n"
    m += f"  Promedio: `{prom_venta:.2f} Bs`\n\n"
    m += f"📊 *SPREAD BS HOY*\n"
    m += f"  Máximo: `{max_spread[1]:.2f} Bs` a las `{max_spread[0]}`\n"
    m += f"  Mínimo: `{min_spread[1]:.2f} Bs` a las `{min_spread[0]}`\n"
    m += f"  Promedio: `{prom_spread:.2f} Bs`\n\n"
    m += f"🏆 *Mejor ventana:* `{mejor_ventana[0]}:00 — {int(mejor_ventana[0])+1:02d}:00`\n\n"
    m += f"💼 *OPERACIONES HOY*\n"
    m += f"  Ops: `{res_hoy['ops']}` | Vol: `{res_hoy['volumen']:.4f} USDT`\n"
    m += f"  Ganancia neta: `{res_hoy['ganancia_neta']:.4f} USDT`\n\n"
    m += f"_Mañana: usa /patron para proyección_"
    return m

def generar_reporte_semanal():
    """Genera el reporte semanal con proyección."""
    conn = get_conn()
    rows = conn.execute("""
        SELECT dia_semana, hora, bs_venta_1, bs_spread_real
        FROM precios_historicos
        WHERE bs_venta_1 IS NOT NULL
        ORDER BY fecha DESC, hora
        LIMIT 2000
    """).fetchall()
    conn.close()

    if len(rows) < 50:
        return "📅 Datos insuficientes para reporte semanal. Se necesita al menos 1 semana de historial."

    por_dia = defaultdict(list)
    for r in rows:
        por_dia[r['dia_semana']].append({
            'venta': r['bs_venta_1'],
            'spread': r['bs_spread_real'] or 0,
            'hora': r['hora'][:2]
        })

    m = f"📅 *REPORTE SEMANAL BS*\n"
    m += f"━━━━━━━━━━━━━━━━━━━━\n\n"

    for dia in DIAS_SEMANA:
        if dia not in por_dia: continue
        datos = por_dia[dia]
        ventas = [d['venta'] for d in datos]
        spreads = [d['spread'] for d in datos]
        por_hora_dia = defaultdict(list)
        for d in datos: por_hora_dia[d['hora']].append(d['spread'])
        mejor_h = max(por_hora_dia.items(), key=lambda x: sum(x[1])/len(x[1])) if por_hora_dia else ("??", [0])
        m += f"*{dia}:* venta prom `{sum(ventas)/len(ventas):.2f} Bs` | spread prom `{sum(spreads)/len(spreads):.2f} Bs` | mejor hora `{mejor_h[0]}:00`\n"

    # Proyección
    proyeccion, dia_manana = proyeccion_manana()
    m += f"\n🔮 *PROYECCIÓN {dia_manana.upper()}*\n"
    if proyeccion:
        ventanas = [(h, d) for h, d in proyeccion.items() if d['spread_esperado'] >= SPREAD_MIN_ALERTA]
        if ventanas:
            m += "Ventanas con spread ≥ 10 Bs esperado:\n"
            for hora, d in sorted(ventanas, key=lambda x: x[1]['spread_esperado'], reverse=True)[:3]:
                m += f"  `{hora}` → spread ~`{d['spread_esperado']:.1f} Bs` | venta ~`{d['venta_esperada']:.2f} Bs`\n"
        else:
            m += "_Sin ventanas de alto spread proyectadas_\n"
    else:
        m += "_Sin suficiente historial para proyectar_\n"

    return m

# ══════════════════════════════════════════════════════════════════════
# RESUMEN DE CORRESPONSALES Y CLIENTES
# ══════════════════════════════════════════════════════════════════════

def registrar_oportunidad_perdida(tipo, descripcion, monto_requerido, moneda, ganancia_estimada, razon, observaciones=""):
    """Registra una oportunidad de ganancia que no se pudo aprovechar por falta de capital."""
    analisis = analizar_capital()
    ahora = now_local()
    datos = {
        'fecha': str(ahora.date()),
        'hora': ahora.strftime("%H:%M"),
        'tipo': tipo,
        'descripcion': descripcion,
        'monto_requerido': monto_requerido,
        'moneda_requerida': moneda,
        'ganancia_estimada_perdida': ganancia_estimada,
        'razon': razon,
        'capital_disponible_usdt': analisis['patrimonio_usdt'],
        'observaciones': observaciones,
    }
    conn = get_conn()
    conn.execute("""INSERT INTO oportunidades_perdidas
        (fecha,hora,tipo,descripcion,monto_requerido,moneda_requerida,
         ganancia_estimada_perdida,razon,capital_disponible_usdt,observaciones)
        VALUES (?,?,?,?,?,?,?,?,?,?)""",
        (datos['fecha'],datos['hora'],datos['tipo'],datos['descripcion'],
         datos['monto_requerido'],datos['moneda_requerida'],
         datos['ganancia_estimada_perdida'],datos['razon'],
         datos['capital_disponible_usdt'],datos['observaciones']))
    conn.commit(); conn.close()
    if USE_SUPABASE:
        supa_insert('oportunidades_perdidas', datos)

def msg_oportunidades(dias=7):
    """Resumen de oportunidades perdidas en los últimos N días."""
    conn = get_conn()
    rows = conn.execute("""
        SELECT fecha, tipo, descripcion, monto_requerido, moneda_requerida,
               ganancia_estimada_perdida, razon
        FROM oportunidades_perdidas
        WHERE fecha >= date('now', ?)
        ORDER BY fecha DESC, hora DESC
        LIMIT 20
    """, (f'-{dias} days',)).fetchall()
    
    total_perdido = conn.execute("""
        SELECT COALESCE(SUM(ganancia_estimada_perdida),0)
        FROM oportunidades_perdidas
        WHERE fecha >= date('now', ?)
    """, (f'-{dias} days',)).fetchone()[0]
    conn.close()

    if not rows:
        return f"✅ Sin oportunidades perdidas en los últimos {dias} días."

    m = f"📉 *OPORTUNIDADES PERDIDAS — últimos {dias} días*\n"
    m += f"━━━━━━━━━━━━━━━━━━━━\n\n"
    
    por_tipo = {}
    for r in rows:
        t = r['tipo'] if isinstance(r, dict) else r[1]
        por_tipo[t] = por_tipo.get(t, 0) + 1

    for r in rows[:10]:
        if isinstance(r, dict):
            fecha, tipo, desc, monto, moneda, ganancia, razon = r['fecha'],r['tipo'],r['descripcion'],r['monto_requerido'],r['moneda_requerida'],r['ganancia_estimada_perdida'],r['razon']
        else:
            fecha, tipo, desc, monto, moneda, ganancia, razon = r[0],r[1],r[2],r[3],r[4],r[5],r[6]
        m += f"🔴 `{fecha}` *{tipo}*\n"
        m += f"   {desc}\n"
        m += f"   Faltaba: `{monto:,.0f} {moneda}` | Perdiste: `{ganancia:.4f} USDT`\n"
        m += f"   Razón: _{razon}_\n\n"

    m += f"━━━━━━━━━━━━━━━━━━━━\n"
    m += f"💸 *Total ganancia perdida: `{total_perdido:.4f} USDT`*\n"
    
    causa_principal = max(por_tipo.items(), key=lambda x: x[1]) if por_tipo else ("N/A", 0)
    m += f"📊 Causa principal: *{causa_principal[0]}* ({causa_principal[1]} veces)\n"
    m += f"\n_Usa /oportunidades 30 para ver el mes completo_"
    return m


def msg_resumen_corresponsal(nombre):
    conn = get_conn()
    ops = conn.execute("""
        SELECT COUNT(*) as total, COALESCE(SUM(usdt_equiv),0) as vol
        FROM operaciones WHERE corresponsal LIKE ? AND estado='Completada'
    """, (f'%{nombre}%',)).fetchone()
    cxp = conn.execute("""
        SELECT COALESCE(SUM(monto),0) as total, COUNT(*) as cnt
        FROM cuentas_pendientes WHERE contraparte LIKE ? AND estado='Pendiente' AND tipo='CXP'
    """, (f'%{nombre}%',)).fetchone()
    conn.close()

    m = f"🏦 *CORRESPONSAL: {nombre.upper()}*\n━━━━━━━━━━━━━━━━━━━━\n\n"
    m += f"Operaciones: `{ops['total']}`\n"
    m += f"Volumen total: `{ops['vol']:.4f} USDT`\n"
    m += f"Comisión acumulada (2.5%): `{ops['vol']*0.025:.4f} USDT`\n\n"
    m += f"💰 *Pendiente por pagar:*\n"
    m += f"  `{cxp['total']:.4f} USDT` ({cxp['cnt']} registros)\n"
    return m

def msg_clientes_top():
    conn = get_conn()
    rows = conn.execute("""
        SELECT nombre, operaciones_total, volumen_usdt
        FROM clientes ORDER BY volumen_usdt DESC LIMIT 5
    """).fetchall()
    conn.close()
    if not rows: return "Sin clientes registrados aún."
    m = "🏆 *TOP CLIENTES*\n\n"
    for i, r in enumerate(rows, 1):
        m += f"{i}. `{r['nombre']}` — `{r['operaciones_total']} ops` | `{r['volumen_usdt']:.2f} USDT`\n"
    return m

def msg_clientes_riesgo():
    conn = get_conn()
    rows = conn.execute("""
        SELECT contraparte, COUNT(*) as cnt, SUM(monto) as total
        FROM cuentas_pendientes
        WHERE tipo='CXC' AND estado='Pendiente'
        GROUP BY contraparte ORDER BY total DESC LIMIT 5
    """).fetchall()
    conn.close()
    if not rows: return "✅ Sin clientes en riesgo (sin CXC pendiente)."
    m = "⚠️ *CLIENTES CON CXC PENDIENTE*\n\n"
    for r in rows:
        m += f"🔴 `{r['contraparte']}` — `{r['cnt']} deudas` | `{r['total']:,.2f}`\n"
    return m

# ══════════════════════════════════════════════════════════════════════
# BASE DE DATOS — FUNCIONES EXISTENTES
# ══════════════════════════════════════════════════════════════════════
def calcular_usdt_equiv(moneda, monto, pat_bs=0, dol_obs=0):
    if not monto: return 0.0
    if moneda in ('USDT','USD','USDC'): return float(monto)
    t = get_ultima_tasa()
    if moneda == 'BS':
        tasa = pat_bs or ((t.get('ban_bs_compra',0) or 0 + (t.get('ban_bs_venta',0) or 0))/2) or 1
        return round(float(monto)/tasa, 4) if tasa else 0
    if moneda == 'CLP':
        tasa = dol_obs or t.get('dolar_obs',0) or 1
        return round(float(monto)/tasa, 4) if tasa else 0
    if moneda == 'COP':
        tasa = t.get('trm',0) or 1
        return round(float(monto)/tasa, 4) if tasa else 0
    return float(monto)

def get_ultima_tasa():
    conn = get_conn()
    row = conn.execute("SELECT * FROM tasas ORDER BY fecha_hora DESC LIMIT 1").fetchone()
    conn.close()
    return dict(row) if row else {}

def calcular_limites(mejor_venta, dol_obs, western):
    limites = {}; tasas = {}
    if mejor_venta and dol_obs:
        limites['clp_bs'] = (mejor_venta*(1-FEE_USDT_BS))/(dol_obs*(1+FEE_USDT_CLP))
    if western:
        limites['clp_cop'] = western*(1-FEE_WU)
    if limites.get('clp_bs') and limites.get('clp_cop'):
        limites['bs_cop'] = limites['clp_cop']/limites['clp_bs']
    if limites.get('clp_bs'):
        tasas['clp_bs'] = round(limites['clp_bs']*(1-FEE_CLP_BS),6)
        tasas['bs_clp'] = round(limites['clp_bs']*(1+FEE_BS_CLP),6)
    if limites.get('clp_cop'):
        tasas['clp_cop'] = round(limites['clp_cop']*(1-FEE_CLP_COP),4)
        tasas['cop_clp'] = round(limites['clp_cop']*(1+FEE_COP_CLP),4)
    if limites.get('bs_cop'):
        tasas['cop_bs'] = round(limites['bs_cop']*(1-FEE_COP_BS),4)
        tasas['bs_cop'] = round(limites['bs_cop']*(1+FEE_BS_COP),4)
    return limites, tasas

def consultar_y_guardar(western_rate=None):
    ban_c,ban_v,ban_s = get_binance_banco_promedio("Banesco")
    mer_c,mer_v,mer_s = get_binance_banco_promedio("Mercantil")
    clp_c,clp_v = get_binance_fiat_promedio("CLP")
    cop_c,cop_v = get_binance_fiat_promedio("COP")
    dol_obs = get_dolar_observado()
    trm     = get_trm()
    bcv_usd,bcv_eur = get_bcv()
    mejor = "Mercantil" if (mer_s or 0) > (ban_s or 0) else "Banesco"
    mejor_venta = mer_v if mejor=="Mercantil" else ban_v
    limites,tasas = calcular_limites(mejor_venta, dol_obs, western_rate)
    datos = {
        'bcv_usd':bcv_usd,'bcv_eur':bcv_eur,
        'ban_bs_compra':ban_c,'ban_bs_venta':ban_v,'ban_bs_spread':ban_s,
        'mer_bs_compra':mer_c,'mer_bs_venta':mer_v,'mer_bs_spread':mer_s,
        'clp_compra':clp_c,'clp_venta':clp_v,'cop_compra':cop_c,'cop_venta':cop_v,
        'trm':trm,'dolar_obs':dol_obs,'western':western_rate,
        'limite_clp_bs':limites.get('clp_bs'),'limite_clp_cop':limites.get('clp_cop'),
        'limite_bs_cop':limites.get('bs_cop'),
        'tasa_gsa_clp_bs':tasas.get('clp_bs'),'tasa_gsa_bs_clp':tasas.get('bs_clp'),
        'tasa_gsa_clp_cop':tasas.get('clp_cop'),'tasa_gsa_cop_clp':tasas.get('cop_clp'),
        'tasa_gsa_cop_bs':tasas.get('cop_bs'),'tasa_gsa_bs_cop':tasas.get('bs_cop'),
        'mejor_banco':mejor,
    }
    # Guardar en DB
    if USE_SUPABASE:
        supa_insert('tasas', {k: v for k, v in datos.items() if k != 'mejor_banco' and v is not None})
    conn = get_conn(); c = conn.cursor()
    c.execute("""INSERT INTO tasas (
        bcv_usd,bcv_eur,ban_bs_compra,ban_bs_venta,ban_bs_spread,
        mer_bs_compra,mer_bs_venta,mer_bs_spread,clp_compra,clp_venta,
        cop_compra,cop_venta,trm,dolar_obs,western,
        limite_clp_bs,limite_clp_cop,limite_bs_cop,
        tasa_gsa_clp_bs,tasa_gsa_bs_clp,tasa_gsa_clp_cop,
        tasa_gsa_cop_clp,tasa_gsa_cop_bs,tasa_gsa_bs_cop
    ) VALUES (
        :bcv_usd,:bcv_eur,:ban_bs_compra,:ban_bs_venta,:ban_bs_spread,
        :mer_bs_compra,:mer_bs_venta,:mer_bs_spread,:clp_compra,:clp_venta,
        :cop_compra,:cop_venta,:trm,:dolar_obs,:western,
        :limite_clp_bs,:limite_clp_cop,:limite_bs_cop,
        :tasa_gsa_clp_bs,:tasa_gsa_bs_clp,:tasa_gsa_clp_cop,
        :tasa_gsa_cop_clp,:tasa_gsa_cop_bs,:tasa_gsa_bs_cop
    )""", datos)
    conn.commit(); conn.close()
    return datos

def guardar_operacion(datos):
    # Calcular ganancia de inventario si aplica
    tipo_op = datos.get('tipo_op', '')
    resultado_inv = None

    # Si la operación involucra venta de USDT (consume inventario)
    if tipo_op in ('USDT→BS', 'CLP→BS', 'COP→BS'):
        t = get_ultima_tasa()
        precio_venta_bs = datos.get('snap_pat_bs', 0) or t.get('ban_bs_venta', 0) or 0
        usdt_a_vender = datos.get('usdt_equiv', 0) or 0
        if usdt_a_vender > 0 and precio_venta_bs > 0:
            resultado_inv = consumir_inventario(usdt_a_vender, precio_venta_bs)
            datos['gan_financiera_bs'] = resultado_inv['ganancia_financiera_bs']
            datos['gan_financiera_usdt'] = resultado_inv['ganancia_financiera_usdt']
            datos['cpp_usado'] = resultado_inv['cpp_efectivo']
            datos['usdt_de_inventario'] = resultado_inv['usdt_de_inventario']
            datos['usdt_comprado_ahora'] = resultado_inv['usdt_comprado_ahora']

    # Si la operación compra USDT (agrega al inventario)
    elif tipo_op in ('BS→USDT', 'CLP→USDT'):
        precio_compra = datos.get('tasa_cliente', 0) or 0
        usdt_comprado = datos.get('monto_salida', 0) or 0
        if usdt_comprado > 0 and precio_compra > 0:
            actualizar_inventario_compra(usdt_comprado, precio_compra)

    # Calcular ganancia comercial
    t = get_ultima_tasa()
    gan_comercial = calcular_ganancia_comercial(
        tipo_op, datos.get('monto_entrada',0),
        datos.get('tasa_cliente',0), t)
    datos['gan_comercial_usdt'] = gan_comercial

    if USE_SUPABASE:
        supa_data = {k: str(v) if isinstance(v, (dict,list)) else v
                     for k, v in datos.items()
                     if k not in ('_tasa_sug','_mto_sal_sug','estado') and v is not None}
        supa_data['estado'] = datos.get('estado', 'Completada')
        supa_insert('operaciones', supa_data)
    conn = get_conn(); c = conn.cursor()
    datos['diferencial'] = (datos.get('tasa_cliente',0) or 0) - (datos.get('tasa_referencia',0) or 0)
    if not datos.get('usdt_equiv'):
        datos['usdt_equiv'] = calcular_usdt_equiv(
            datos.get('mon_entrada',''), datos.get('monto_entrada',0),
            datos.get('snap_pat_bs',0), datos.get('snap_dol_obs',0))
    keys = ['fecha','hora','cliente','referente','tipo_op','origen_fondos',
            'mon_entrada','monto_entrada','mon_salida','monto_salida',
            'tasa_cliente','tasa_referencia','usdt_equiv','diferencial',
            'metodo','corresponsal','traslado_bs','encomienda_cop',
            'repartidor','financiador','estado','observaciones',
            'cxc_pendiente','cxp_pendiente','snap_pat_bs','snap_dol_obs',
            'snap_trm','usuario_telegram']
    c.execute(f"""INSERT INTO operaciones ({','.join(keys)})
                  VALUES ({','.join([':'+k for k in keys])})""",
              {k: datos.get(k) for k in keys})
    op_id = c.lastrowid
    _actualizar_saldos_op(c, datos)
    _actualizar_cliente_db(c, datos)
    if datos.get('cxc_pendiente',0) > 0:
        c.execute("INSERT INTO cuentas_pendientes (tipo,contraparte,concepto,monto,moneda,vencimiento,op_origen_id) VALUES (?,?,?,?,?,date('now'),?)",
                  ('CXC',datos['cliente'],f"CXC Op#{op_id}",datos['cxc_pendiente'],datos['mon_entrada'],op_id))
    if datos.get('corresponsal') and datos.get('usdt_equiv',0) > 0:
        com = datos['usdt_equiv'] * 0.025
        c.execute("INSERT INTO cuentas_pendientes (tipo,contraparte,concepto,monto,moneda,vencimiento,op_origen_id) VALUES (?,?,?,?,'USDT',date('now','+7 days'),?)",
                  ('CXP',datos['corresponsal'],f"Comisión 2.5% Op#{op_id}",round(com,4),op_id))
    # Registrar costos en ledger
    costos = calcular_costos_operacion(datos)
    for costo in costos:
        registrar_costo_operacion(op_id, costo['tipo'],
            costo['descripcion'], costo['monto'], costo['moneda'])

    # Registrar en ledger
    ledger_insert('OPERACION_CLIENTE',
        datos.get('mon_sal','') or 'CAJA',
        datos.get('mon_entrada','') or 'CAJA',
        datos.get('mon_entrada','USDT'),
        datos.get('monto_entrada',0) or 0,
        ref_id=op_id, ref_tipo='operacion',
        descripcion=f"Op#{op_id} {datos.get('tipo_op','')} {datos.get('cliente','')}",
        usuario=datos.get('usuario_telegram','sistema'))

    conn.commit(); conn.close()
    return op_id

def _actualizar_saldos_op(c, datos):
    tipo = datos.get('tipo_op','')
    mto_ent = datos.get('monto_entrada',0) or 0
    mto_sal = datos.get('monto_salida',0) or 0
    mapa = {
        'CLP→BS':('CLP_COPEC_PAY','BS_BANESCO'),'BS→CLP':('BS_BANESCO','CLP_COPEC_PAY'),
        'CLP→COP':('CLP_COPEC_PAY','COP_EFECTIVO_ORLANDO'),'COP→CLP':('COP_EFECTIVO_ORLANDO','CLP_COPEC_PAY'),
        'COP→BS':('COP_EFECTIVO_ORLANDO','BS_BANESCO'),'BS→COP':('BS_BANESCO','COP_EFECTIVO_ORLANDO'),
        'CLP→USDT':('CLP_COPEC_PAY','USDT_BINANCE'),'USDT→CLP':('USDT_BINANCE','CLP_COPEC_PAY'),
        'BS→USDT':('BS_BANESCO','USDT_BINANCE'),'USDT→BS':('USDT_BINANCE','BS_BANESCO'),
        'USD→CLP':('USD_EFECTIVO','CLP_COPEC_PAY'),'CLP→USD':('CLP_COPEC_PAY','USD_EFECTIVO'),
        'USD→BS':('USD_EFECTIVO','BS_BANESCO'),'BS→USD':('BS_BANESCO','USD_EFECTIVO'),
    }
    if tipo in mapa:
        ce, cs = mapa[tipo]
        c.execute("UPDATE saldos SET saldo=saldo-?,ultima_actualizacion=CURRENT_TIMESTAMP WHERE cuenta=?", (mto_ent,ce))
        c.execute("UPDATE saldos SET saldo=saldo+?,ultima_actualizacion=CURRENT_TIMESTAMP WHERE cuenta=?", (mto_sal,cs))

def _actualizar_cliente_db(c, datos):
    nombre = datos.get('cliente','')
    if not nombre or 'Binance' in nombre or 'GSA' in nombre: return
    usdt = datos.get('usdt_equiv',0) or 0
    c.execute("""INSERT INTO clientes (nombre,ultima_operacion,operaciones_total,volumen_usdt)
                 VALUES (?,date('now'),1,?)
                 ON CONFLICT(nombre) DO UPDATE SET
                 ultima_operacion=date('now'),
                 operaciones_total=operaciones_total+1,
                 volumen_usdt=volumen_usdt+excluded.volumen_usdt""", (nombre,usdt))

def get_saldos():
    conn = get_conn()
    rows = conn.execute("SELECT cuenta,moneda,saldo FROM saldos").fetchall()
    conn.close()
    return {r['cuenta']:{'moneda':r['moneda'],'saldo':r['saldo']} for r in rows}

def set_saldo(cuenta, saldo):
    conn = get_conn()
    conn.execute("INSERT INTO saldos (cuenta,moneda,saldo) VALUES (?,'',:s) ON CONFLICT(cuenta) DO UPDATE SET saldo=:s,ultima_actualizacion=CURRENT_TIMESTAMP",
                 {'cuenta':cuenta,'s':saldo})
    conn.commit(); conn.close()

def set_saldo_inicial(cuenta, saldo):
    conn = get_conn(); c = conn.cursor()
    c.execute("""INSERT INTO saldos_iniciales (cuenta,saldo,fecha) VALUES (?,?,date('now'))
                 ON CONFLICT(cuenta) DO UPDATE SET saldo=excluded.saldo, fecha=excluded.fecha""", (cuenta, saldo))
    movimiento = _calcular_movimiento_cuenta(c, cuenta)
    c.execute("UPDATE saldos SET saldo=?,ultima_actualizacion=CURRENT_TIMESTAMP WHERE cuenta=?",
              (saldo + movimiento, cuenta))
    conn.commit(); conn.close()

def _calcular_movimiento_cuenta(c, cuenta):
    mapa_entrada = {
        'BS_BANESCO': ["USDT→BS","COP→BS","USD→BS"], 'BS_MERCANTIL': [],
        'CLP_COPEC_PAY': ["BS→CLP","USDT→CLP","USD→CLP"],
        'COP_EFECTIVO_ORLANDO': ["BS→COP","CLP→COP"],
        'USDT_BINANCE': ["BS→USDT","CLP→USDT"], 'USD_EFECTIVO': ["BS→USD","CLP→USD"],
    }
    mapa_salida = {
        'BS_BANESCO': ["BS→CLP","BS→COP","BS→USDT","BS→USD"],
        'CLP_COPEC_PAY': ["CLP→BS","CLP→COP","CLP→USDT","CLP→USD"],
        'COP_EFECTIVO_ORLANDO': ["COP→BS","COP→CLP"],
        'USDT_BINANCE': ["USDT→BS","USDT→CLP"], 'USD_EFECTIVO': ["USD→BS","USD→CLP"],
    }
    entradas = mapa_entrada.get(cuenta, [])
    salidas  = mapa_salida.get(cuenta, [])
    total = 0.0
    if entradas:
        ph = ",".join(["?"]*len(entradas))
        rows = c.execute(f"SELECT COALESCE(SUM(monto_salida),0) FROM operaciones WHERE tipo_op IN ({ph}) AND estado='Completada'", entradas).fetchone()
        total += rows[0] if rows else 0
    if salidas:
        ph = ",".join(["?"]*len(salidas))
        rows = c.execute(f"SELECT COALESCE(SUM(monto_entrada),0) FROM operaciones WHERE tipo_op IN ({ph}) AND estado='Completada'", salidas).fetchone()
        total -= rows[0] if rows else 0
    return total

def get_cuentas_pendientes(tipo=None):
    conn = get_conn()
    if tipo:
        rows = conn.execute("SELECT * FROM cuentas_pendientes WHERE estado='Pendiente' AND tipo=? ORDER BY fecha",(tipo,)).fetchall()
    else:
        rows = conn.execute("SELECT * FROM cuentas_pendientes WHERE estado='Pendiente' ORDER BY tipo,fecha").fetchall()
    conn.close()
    return [dict(r) for r in rows]

def marcar_pagado(cp_id):
    conn = get_conn()
    conn.execute("UPDATE cuentas_pendientes SET estado='Pagado' WHERE id=?",(cp_id,))
    conn.commit(); conn.close()

def get_resultados_hoy():
    conn = get_conn()
    hoy = str(today_local())
    row = conn.execute("SELECT COUNT(*) as ops,COALESCE(SUM(usdt_equiv),0) as vol,COALESCE(SUM(usdt_equiv*diferencial),0) as gan FROM operaciones WHERE fecha=? AND estado='Completada'", (hoy,)).fetchone()
    gas = conn.execute("SELECT COALESCE(SUM(usdt_equiv),0) as t FROM gastos WHERE fecha=?", (hoy,)).fetchone()
    conn.close()
    return {'ops':row['ops'],'volumen':row['vol'],'ganancia_operativa':row['gan'],'gastos':gas['t'],'ganancia_neta':row['gan']-gas['t']}

def get_resultados_mes():
    conn = get_conn()
    mes = now_local().strftime('%Y-%m')
    row = conn.execute("SELECT COUNT(*) as ops,COALESCE(SUM(usdt_equiv),0) as vol,COALESCE(SUM(usdt_equiv*diferencial),0) as gan FROM operaciones WHERE strftime('%Y-%m',fecha)=? AND estado='Completada'", (mes,)).fetchone()
    gas = conn.execute("SELECT COALESCE(SUM(usdt_equiv),0) as t FROM gastos WHERE strftime('%Y-%m',fecha)=?", (mes,)).fetchone()
    conn.close()
    return {'ops':row['ops'],'volumen':row['vol'],'ganancia_operativa':row['gan'],'gastos':gas['t'],'ganancia_neta':row['gan']-gas['t']}

def get_operaciones_hoy():
    conn = get_conn()
    hoy = str(today_local())
    rows = conn.execute("SELECT * FROM operaciones WHERE fecha=? ORDER BY hora", (hoy,)).fetchall()
    conn.close()
    return [dict(r) for r in rows]

# ══════════════════════════════════════════════════════════════════════
# MENSAJES TASAS Y SALDOS
# ══════════════════════════════════════════════════════════════════════
def fmt(v, d=2): return f"{v:,.{d}f}" if v else "N/D"

def spread_emoji(s):
    if s>=SPREAD_PREMIUM: return "🚀"
    if s>=SPREAD_BUENO:   return "🟢"
    if s>=SPREAD_MODERADO:return "🟡"
    return "🔴"

def construir_mensaje(d, es_especial=False):
    ahora = now_local().strftime("%d/%m/%Y — %I:%M %p")
    ban_s = d.get('ban_bs_spread',0) or 0
    mer_s = d.get('mer_bs_spread',0) or 0
    mejor = d.get('mejor_banco','—')
    prefijo = "🔔 *TASA DE REFERENCIA*\n" if es_especial else ""

    m  = f"{prefijo}📊 *RESUMEN DE TASAS*\n📅 {ahora}\n━━━━━━━━━━━━━━━━━━━━\n\n"

    # TASAS OFICIALES
    m += "🌎 *TASAS OFICIALES*\n\n"
    if d.get('trm'):         m += f"🇨🇴  *TRM*\n      `{fmt(d['trm'])} COP`\n\n"
    if d.get('bcv_usd'):     m += f"🏦  *USD/BCV*\n      `{fmt(d['bcv_usd'])} Bs`\n\n"
    if d.get('ban_bs_venta'):
        m += f"🏦  *Binance Banesco* {spread_emoji(ban_s)}\n      Compra: `{fmt(d['ban_bs_compra'])} Bs` | Venta: `{fmt(d['ban_bs_venta'])} Bs`\n      Spread: `{fmt(ban_s)} Bs`\n\n"
    if d.get('mer_bs_venta'):
        m += f"🏦  *Binance Mercantil* {spread_emoji(mer_s)}\n      Compra: `{fmt(d['mer_bs_compra'])} Bs` | Venta: `{fmt(d['mer_bs_venta'])} Bs`\n      Spread: `{fmt(mer_s)} Bs`\n\n"
    m += f"⭐ *Mejor opción: {mejor}*\n\n"
    if d.get('clp_venta'):   m += f"🔵  *Binance USDT/CLP*\n      Compra: `{fmt(d['clp_compra'])} CLP` | Venta: `{fmt(d['clp_venta'])} CLP`\n\n"
    if d.get('cop_venta'):   m += f"🔵  *Binance USDT/COP*\n      Compra: `{fmt(d['cop_compra'])} COP` | Venta: `{fmt(d['cop_venta'])} COP`\n\n"
    if d.get('dolar_obs'):   m += f"🇨🇱  *Dólar Observado*\n      `{fmt(d['dolar_obs'])} CLP`\n\n"
    if d.get('western'):     m += f"🌍  *Western Unión*\n      `{fmt(d['western'],4)} CLP/COP`\n\n"
    else:                    m += f"🌍  *Western Unión*\n      _Envía /western TASA_\n\n"

    # GSA CAMBIOS — organizado por categorías
    m += f"━━━━━━━━━━━━━━━━━━━━\n💱 *GSA CAMBIOS*\n_Calculado con {mejor}_\n\n"

    # Giros
    m += f"📌 *Giros*\n"
    if d.get('tasa_gsa_clp_bs'):
        m += f"🇨🇱➡️🇻🇪  CLP → Bs      `{fmt(d['tasa_gsa_clp_bs'],6)}`\n"
        m += f"🇻🇪➡️🇨🇱  Bs → CLP      `{fmt(d['tasa_gsa_bs_clp'],6)}`\n"
    if d.get('tasa_gsa_clp_cop'):
        m += f"🇨🇱➡️🇨🇴  CLP → COP     `{fmt(d['tasa_gsa_clp_cop'],4)}`\n"
        m += f"🇨🇴➡️🇨🇱  COP → CLP     `{fmt(d['tasa_gsa_cop_clp'],4)}`\n"
    if d.get('dolar_obs'):
        m += f"🇨🇱➡️🇺🇸  CLP → USD     `{fmt(d['dolar_obs']+SPREAD_CLP)} CLP`\n"
        m += f"🇺🇸➡️🇨🇱  USD → CLP     `{fmt(d['dolar_obs']-SPREAD_CLP)} CLP`\n"
    m += "\n"

    # Compra/Venta Bolívares
    m += f"📌 *Compra / Venta Bolívares*\n"
    if d.get('ban_bs_venta'):
        tasa_usd_bs = (d.get('ban_bs_venta',0) or 0) + MARGEN_BS
        tasa_bs_usd = (d.get('ban_bs_compra',0) or 0) - MARGEN_BS
        m += f"🇺🇸➡️🇻🇪  USD → Bs      `{fmt(tasa_usd_bs,2)} Bs`\n"
        m += f"🇻🇪➡️🇺🇸  Bs → USD      `{fmt(tasa_bs_usd,2)} Bs`\n"
    if d.get('tasa_gsa_cop_bs'):
        m += f"🇨🇴➡️🇻🇪  COP → Bs      `{fmt(d['tasa_gsa_cop_bs'],4)}`\n"
        m += f"🇻🇪➡️🇨🇴  Bs → COP      `{fmt(d['tasa_gsa_bs_cop'],4)}`\n"
    m += "\n"

    # Compra/Venta Pesos Colombianos
    if d.get('cop_venta'):
        m += f"📌 *Compra / Venta Pesos Colombianos*\n"
        m += f"🇺🇸➡️🇨🇴  USD → COP     `{fmt(d['cop_venta'],2)} COP`\n"
        m += f"🇨🇴➡️🇺🇸  COP → USD     `{fmt(d['cop_compra'],2)} COP`\n"
        m += "\n"

    # Límites
    m += f"━━━━━━━━━━━━━━━━━━━━\n📐 *LÍMITES OPERATIVOS*\n\n"
    if d.get('limite_clp_bs'):  m += f"🔴  *Límite CLP/Bs*\n      `{fmt(d['limite_clp_bs'],6)}`\n\n"
    if d.get('limite_clp_cop'): m += f"🔴  *Límite CLP/COP*\n      `{fmt(d['limite_clp_cop'],4)}`\n\n"
    if d.get('limite_bs_cop'):  m += f"🔴  *Límite Bs/COP*\n      `{fmt(d['limite_bs_cop'],4)}`\n\n"
    m += f"━━━━━━━━━━━━━━━━━━━━\n🏦 *Banco recomendado: {mejor}*"
    return m

GRUPOS_CUENTAS = {
    "🏦 CAJAS FÍSICAS":["COP_EFECTIVO_ORLANDO","USD_EFECTIVO"],
    "🇻🇪 BANCOS VENEZUELA":["BS_BANESCO","BS_MERCANTIL"],
    "🇨🇱 BANCOS CHILE":["CLP_COPEC_PAY","CLP_BANCOESTADO"],
    "🇨🇴 BANCOS COLOMBIA":["COP_BANCOLOMBIA_C1","COP_BANCOLOMBIA_C2","COP_NEQUI_C1","COP_NEQUI_C2","COP_NEQUI_C3"],
    "💎 WALLETS":["USDT_BINANCE","USDC_AIRTM"],
}

def get_moneda(cuenta):
    for m in ['BS','CLP','COP','USD','USDT','USDC']:
        if m in cuenta: return m
    return ''

def msg_saldos():
    saldos=get_saldos(); t=get_ultima_tasa()
    pat_bs=((t.get('ban_bs_compra',0) or 0)+(t.get('ban_bs_venta',0) or 0))/2 or 1
    dol_obs=t.get('dolar_obs',1) or 1
    ahora=now_local().strftime("%d/%m %I:%M %p")
    m=f"💰 *SALDOS GSA CAMBIOS*\n📅 {ahora}\n━━━━━━━━━━━━━━━━━━━━\n\n"
    patrimonio=0
    for grupo,cuentas in GRUPOS_CUENTAS.items():
        total_usdt=0; lineas=[]
        for cuenta in cuentas:
            info=saldos.get(cuenta,{}); saldo=info.get('saldo',0) or 0
            moneda=info.get('moneda','') or get_moneda(cuenta)
            usdt=calcular_usdt_equiv(moneda,saldo,pat_bs,dol_obs); total_usdt+=usdt
            nombre=NOMBRES_CUENTAS.get(cuenta,cuenta)
            if saldo!=0: lineas.append(f"  `{nombre:18s}` `{saldo:>12,.2f} {moneda}`")
        if lineas:
            m+=f"*{grupo}*\n"+"\n".join(lineas)+f"\n  Total: `{total_usdt:.2f} USDT`\n\n"
            patrimonio+=total_usdt
    m+=f"━━━━━━━━━━━━━━━━━━━━\n🏛️ *PATRIMONIO: `{patrimonio:.4f} USDT`*"
    return m

def msg_dashboard():
    hoy=get_resultados_hoy(); mes=get_resultados_mes()
    sald=get_saldos(); t=get_ultima_tasa()
    cxc=get_cuentas_pendientes('CXC'); cxp=get_cuentas_pendientes('CXP')
    pat_bs=((t.get('ban_bs_compra',0) or 0)+(t.get('ban_bs_venta',0) or 0))/2 or 1
    dol_obs=t.get('dolar_obs',1) or 1
    patrimonio=sum(calcular_usdt_equiv(info.get('moneda','') or get_moneda(k),info.get('saldo',0) or 0,pat_bs,dol_obs) for k,info in sald.items())
    ban_s=t.get('ban_bs_spread',0) or 0; mer_s=t.get('mer_bs_spread',0) or 0
    mejor_s=max(ban_s,mer_s)
    total_cxc=sum(c.get('monto',0) or 0 for c in cxc); total_cxp=sum(c.get('monto',0) or 0 for c in cxp)
    ahora=now_local().strftime("%d/%m/%Y %I:%M %p")
    m=f"📊 *GSA CAMBIOS — DASHBOARD*\n📅 {ahora}\n━━━━━━━━━━━━━━━━━━━━━━━━\n\n"
    m+=f"📈 *HOY*\n  Ops: `{hoy['ops']}` | Vol: `{hoy['volumen']:.4f} USDT`\n  Ganancia: `{hoy['ganancia_neta']:.4f} USDT`\n\n"
    m+=f"📆 *MES*\n  Ops: `{mes['ops']}` | Ganancia: `{mes['ganancia_neta']:.4f} USDT`\n\n"
    m+=f"🏛️ *PATRIMONIO: `{patrimonio:.4f} USDT`*\n\n"
    senal="🟢 OPERAR" if mejor_s>=10 else "🟡 PRECAUCIÓN" if mejor_s>=7 else "🔴 NO OPERAR"
    m+=f"💱 *MERCADO:* {senal} (spread {mejor_s} BS)\n\n"
    if total_cxc>0: m+=f"⚠️ CXC: `{total_cxc:,.2f}` ({len(cxc)} registros)\n"
    if total_cxp>0: m+=f"⚠️ CXP: `{total_cxp:,.4f}` ({len(cxp)} registros)\n"
    bs=(sald.get('BS_BANESCO',{}).get('saldo',0) or 0)+(sald.get('BS_MERCANTIL',{}).get('saldo',0) or 0)
    clp=sald.get('CLP_COPEC_PAY',{}).get('saldo',0) or 0
    usdt=sald.get('USDT_BINANCE',{}).get('saldo',0) or 0
    m+=f"\n💰 *SALDOS CLAVE*\n  BS: `{bs:,.2f}` | CLP: `{clp:,.2f}` | USDT: `{usdt:.4f}`\n\n"
    m+="_/capital para análisis completo_"
    return m

def msg_cxc():
    cxcs=get_cuentas_pendientes('CXC')
    if not cxcs: return "✅ *No hay cuentas por cobrar pendientes.*"
    m="📋 *CXC — POR COBRAR*\n\n"; total=0
    for c in cxcs:
        m+=f"*#{c['id']}* — {c['contraparte']}\n  `{c['monto']:,.2f} {c['moneda']}` — {c['concepto']}\n\n"
        total+=c.get('monto',0) or 0
    m+=f"*Total: `{total:,.2f}`*\n_/cobrado ID para marcar_"
    return m

def msg_cxp():
    cxps=get_cuentas_pendientes('CXP')
    if not cxps: return "✅ *No hay cuentas por pagar pendientes.*"
    m="📋 *CXP — POR PAGAR*\n\n"; total=0
    for c in cxps:
        m+=f"*#{c['id']}* — {c['contraparte']}\n  `{c['monto']:,.4f} {c['moneda']}` — {c['concepto']}\n\n"
        total+=c.get('monto',0) or 0
    m+=f"*Total: `{total:,.4f}`*\n_/pagado ID para marcar_"
    return m

def msg_saldos_iniciales():
    conn = get_conn()
    try:
        rows = conn.execute("SELECT cuenta,saldo,fecha FROM saldos_iniciales ORDER BY cuenta").fetchall()
        if not rows: return "No hay saldos iniciales.\n\nUsa: `/saldo_inicial BS_BANESCO 303581.42`"
        m = "📋 *SALDOS INICIALES*\n\n"
        for row in rows:
            nombre = NOMBRES_CUENTAS.get(row['cuenta'], row['cuenta'])
            m += f"`{nombre}`: `{row['saldo']:,.2f}` (desde {row['fecha']})\n"
        m += "\n_Usa /saldo_inicial CUENTA MONTO para corregir_"
        return m
    except: return "Tabla no encontrada."
    finally: conn.close()

# ══════════════════════════════════════════════════════════════════════
# IMPORTADOR BINANCE C2C
# ══════════════════════════════════════════════════════════════════════
def importar_c2c_inteligente(ruta_archivo, db_path, usuario='importacion'):
    import sqlite3 as _sq3
    from datetime import datetime as _dt

    _conn = _sq3.connect(db_path)
    _conn.execute("""CREATE TABLE IF NOT EXISTS debug_log
        (id INTEGER PRIMARY KEY AUTOINCREMENT, ts DATETIME DEFAULT CURRENT_TIMESTAMP, msg TEXT)""")
    _conn.execute("""CREATE TABLE IF NOT EXISTS binance_sesiones (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        sesion TEXT, fecha DATE, hora_inicio TEXT, hora_fin TEXT,
        compras INTEGER DEFAULT 0, ventas INTEGER DEFAULT 0,
        usdt_comprado REAL DEFAULT 0, bs_pagado REAL DEFAULT 0,
        usdt_vendido REAL DEFAULT 0, bs_recibido REAL DEFAULT 0,
        cpp_bs REAL DEFAULT 0, precio_venta_bs REAL DEFAULT 0,
        ganancia_bs REAL DEFAULT 0, ganancia_usdt REAL DEFAULT 0,
        usdt_pendiente REAL DEFAULT 0, fees_usdt REAL DEFAULT 0,
        estado TEXT DEFAULT 'Cerrado',
        fecha_registro DATETIME DEFAULT CURRENT_TIMESTAMP)""")
    _conn.commit()

    def _log(msg):
        _conn.execute("INSERT INTO debug_log (msg) VALUES (?)", (msg,))
        _conn.commit()

    def _f(v):
        if v is None: return 0.0
        s = str(v).strip().strip("'")
        try: return float(s.replace(',','')) if s else 0.0
        except: return 0.0

    try:
        from openpyxl import load_workbook as _lw
        _wb = _lw(ruta_archivo, data_only=True)
        _ws = _wb.active
        _log(f"file_opened: rows={_ws.max_row}")
    except Exception as e:
        _log(f"file_error: {e}"); _conn.close()
        return {'error': str(e), 'importadas_maker':0,'importadas_taker':0,
                'importadas_clp':0,'omitidas':0,'errores':0,'sesiones':[],
                'total_ganancia_bs':0,'total_ganancia_u':0,'usdt_pendiente':0,'fees_total':0}

    _header = 11
    for _i, _row in enumerate(_ws.iter_rows(values_only=True), 1):
        for _v in _row:
            if _v and 'Order Number' in str(_v):
                _header = _i + 1; break
        if _header != 11: break

    _ordenes = []
    _skip_status = 0; _skip_date = 0; _skip_nonum = 0
    for _row in _ws.iter_rows(min_row=_header, values_only=True):
        if not _row[2]: _skip_nonum+=1; continue
        _st = str(_row[13]).strip().strip("'") if _row[13] else ''
        if _st != 'Completed': _skip_status+=1; continue
        _cr = str(_row[14]).strip().strip("'") if _row[14] else ''
        try:
            _fs = '20'+_cr if _cr.startswith('26-') else _cr
            _dt2 = _dt.strptime(_fs[:16], '%Y-%m-%d %H:%M')
        except: _skip_date+=1; continue
        _tf = _f(_row[11]); _mf = _f(_row[10])
        _ordenes.append({
            'num': str(_row[2]).strip().strip("'"),
            'tipo': str(_row[3]).strip().strip("'"),
            'fiat': str(_row[5]).strip().strip("'"),
            'total':_f(_row[6]),'precio':_f(_row[7]),'cantidad':_f(_row[8]),
            'maker_fee':_mf,'taker_fee':_tf,
            'contra':str(_row[12]).strip().strip("'") if _row[12] else '',
            'dt':_dt2,'is_taker':_tf>0,'is_maker':_tf==0,
        })

    _maker = [o for o in _ordenes if o['fiat']=='VES' and o['is_maker']]
    _taker = [o for o in _ordenes if o['is_taker']]
    _clp   = [o for o in _ordenes if o['fiat']=='CLP' and not o['is_taker']]

    _sesiones_raw = []
    if _maker:
        _sa = [_maker[0]]
        for _o in _maker[1:]:
            _gap = (_o['dt']-_sa[-1]['dt']).total_seconds()/60
            if _gap > PAUSA_SESION_MIN: _sesiones_raw.append(_sa); _sa=[_o]
            else: _sa.append(_o)
        if _sa: _sesiones_raw.append(_sa)

    _imp=0; _omit=0; _err=0; _ses_saved=[]
    for _ns, _ords in enumerate(_sesiones_raw, 1):
        _comp=[o for o in _ords if o['tipo']=='Buy']
        _vent=[o for o in _ords if o['tipo']=='Sell']
        _bsp=sum(o['total'] for o in _comp); _uc=sum(o['cantidad'] for o in _comp)
        _bsr=sum(o['total'] for o in _vent); _uv=sum(o['cantidad'] for o in _vent)
        _fees=sum(o['maker_fee']+o['taker_fee'] for o in _ords)
        _cpp=_bsp/_uc if _uc else 0; _pv=_bsr/_uv if _uv else 0
        _gb=_bsr-(_uv*_cpp) if _uv and _cpp else 0
        _gu=_gb/_cpp if _cpp else 0; _pend=_uc-_uv
        _fd=_ords[0]['dt'].strftime('%Y-%m-%d'); _hi=_ords[0]['dt'].strftime('%H:%M')
        _hf=_ords[-1]['dt'].strftime('%H:%M'); _sl=f"S{_ns}"
        _est='Abierto' if _pend>0.01 else 'Cerrado'

        _ex=_conn.execute("SELECT id FROM binance_sesiones WHERE sesion=? AND fecha=?",(_sl,_fd)).fetchone()
        if _ex:
            _conn.execute("""UPDATE binance_sesiones SET
                compras=?,ventas=?,usdt_comprado=?,bs_pagado=?,usdt_vendido=?,bs_recibido=?,
                cpp_bs=?,precio_venta_bs=?,ganancia_bs=?,ganancia_usdt=?,usdt_pendiente=?,
                fees_usdt=?,estado=?,hora_fin=? WHERE sesion=? AND fecha=?""",
                (len(_comp),len(_vent),round(_uc,4),round(_bsp,2),round(_uv,4),round(_bsr,2),
                 round(_cpp,4),round(_pv,4),round(_gb,2),round(_gu,4),round(_pend,4),
                 round(_fees,4),_est,_hf,_sl,_fd))
        else:
            _conn.execute("""INSERT INTO binance_sesiones
                (sesion,fecha,hora_inicio,hora_fin,compras,ventas,usdt_comprado,bs_pagado,
                 usdt_vendido,bs_recibido,cpp_bs,precio_venta_bs,ganancia_bs,ganancia_usdt,
                 usdt_pendiente,fees_usdt,estado) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (_sl,_fd,_hi,_hf,len(_comp),len(_vent),round(_uc,4),round(_bsp,2),
                 round(_uv,4),round(_bsr,2),round(_cpp,4),round(_pv,4),round(_gb,2),
                 round(_gu,4),round(_pend,4),round(_fees,4),_est))
        _conn.commit()

        if USE_SUPABASE:
            ses_data = {'sesion':_sl,'fecha':_fd,'hora_inicio':_hi,'hora_fin':_hf,
                'compras':len(_comp),'ventas':len(_vent),'usdt_comprado':round(_uc,4),
                'bs_pagado':round(_bsp,2),'usdt_vendido':round(_uv,4),'bs_recibido':round(_bsr,2),
                'cpp_bs':round(_cpp,4),'precio_venta_bs':round(_pv,4),'ganancia_bs':round(_gb,2),
                'ganancia_usdt':round(_gu,4),'usdt_pendiente':round(_pend,4),
                'fees_usdt':round(_fees,4),'estado':_est}
            existing = supa_select('binance_sesiones',f'sesion=eq.{_sl}&fecha=eq.{_fd}')
            if existing: supa_update('binance_sesiones','sesion',_sl,ses_data)
            else: supa_insert('binance_sesiones',ses_data)

        for _o in _ords:
            _ex2=_conn.execute("SELECT id FROM operaciones WHERE observaciones LIKE ?",(f'%{_o["num"]}%',)).fetchone()
            if _ex2: _omit+=1; continue
            if _o['tipo']=='Buy':
                _top='BS→USDT';_me='BS';_ment=_o['total'];_ms='USDT';_msal=_o['cantidad']-_o['maker_fee']
                # Actualizar inventario: compramos USDT
                if _msal > 0 and _o['precio'] > 0:
                    actualizar_inventario_compra(_msal, _o['precio'])
            else:
                _top='USDT→BS';_me='USDT';_ment=_o['cantidad'];_ms='BS';_msal=_o['total']
                # Consumir inventario: vendemos USDT
                if _ment > 0 and _o['precio'] > 0:
                    consumir_inventario(_ment, _o['precio'])
            try:
                _conn.execute("""INSERT INTO operaciones
                    (fecha,hora,cliente,tipo_op,mon_entrada,monto_entrada,mon_salida,monto_salida,
                     tasa_cliente,tasa_referencia,usdt_equiv,diferencial,metodo,estado,observaciones,usuario_telegram)
                    VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (_o['dt'].strftime('%Y-%m-%d'),_o['dt'].strftime('%H:%M'),
                     f'Binance Maker ({_o["contra"]})',_top,_me,_ment,_ms,_msal,
                     _o['precio'],_o['precio'],_o['cantidad'],0,
                     f'Binance Maker — {_sl}','Completada',
                     f'Orden #{_o["num"]} | {_sl}',usuario))
                _imp+=1
                if USE_SUPABASE:
                    supa_insert('operaciones',{
                        'fecha':_o['dt'].strftime('%Y-%m-%d'),'hora':_o['dt'].strftime('%H:%M'),
                        'cliente':f'Binance Maker ({_o["contra"]})','tipo_op':_top,
                        'mon_entrada':_me,'monto_entrada':_ment,'mon_salida':_ms,'monto_salida':_msal,
                        'tasa_cliente':_o['precio'],'tasa_referencia':_o['precio'],
                        'usdt_equiv':_o['cantidad'],'diferencial':0,
                        'metodo':f'Binance Maker — {_sl}','estado':'Completada',
                        'observaciones':f'Orden #{_o["num"]} | {_sl}','usuario_telegram':usuario})
            except Exception as _ie: _err+=1; _log(f"insert_err: {_ie}")
        _conn.commit()
        _ses_saved.append({'sesion':_sl,'fecha':_fd,'hora_ini':_hi,'hora_fin':_hf,
            'compras':len(_comp),'ventas':len(_vent),'usdt_comp':_uc,'bs_pagado':_bsp,
            'usdt_vend':_uv,'bs_recibido':_bsr,'cpp':_cpp,'pv':_pv,
            'gan_bs':_gb,'gan_u':_gu,'pend':_pend,'fees':_fees,'estado':_est})

    # Taker
    _timp=0
    for _o in _taker:
        _ex=_conn.execute("SELECT id FROM operaciones WHERE observaciones LIKE ?",(f'%{_o["num"]}%',)).fetchone()
        if _ex: _omit+=1; continue
        _fiat=_o['fiat']
        if _fiat=='VES':
            if _o['tipo']=='Buy': _top='BS→USDT';_me='BS';_ment=_o['total'];_ms='USDT';_msal=_o['cantidad']-_o['taker_fee']
            else: _top='USDT→BS';_me='USDT';_ment=_o['cantidad'];_ms='BS';_msal=_o['total']
        elif _fiat=='CLP':
            if _o['tipo']=='Buy': _top='CLP→USDT';_me='CLP';_ment=_o['total'];_ms='USDT';_msal=_o['cantidad']-_o['taker_fee']
            else: _top='USDT→CLP';_me='USDT';_ment=_o['cantidad'];_ms='CLP';_msal=_o['total']
        else: continue
        try:
            _conn.execute("""INSERT INTO operaciones
                (fecha,hora,cliente,tipo_op,mon_entrada,monto_entrada,mon_salida,monto_salida,
                 tasa_cliente,tasa_referencia,usdt_equiv,diferencial,metodo,estado,observaciones,usuario_telegram)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (_o['dt'].strftime('%Y-%m-%d'),_o['dt'].strftime('%H:%M'),
                 f'Binance Taker ({_o["contra"]})',_top,_me,_ment,_ms,_msal,
                 _o['precio'],_o['precio'],_o['cantidad'],0,
                 'Binance P2P Taker','Completada',f'Orden #{_o["num"]} | Taker',usuario))
            _timp+=1
        except: pass

    # CLP
    _cimp=0
    for _o in _clp:
        _ex=_conn.execute("SELECT id FROM operaciones WHERE observaciones LIKE ?",(f'%{_o["num"]}%',)).fetchone()
        if _ex: continue
        if _o['tipo']=='Buy': _top='CLP→USDT';_me='CLP';_ment=_o['total'];_ms='USDT';_msal=_o['cantidad']-_o['maker_fee']
        else: _top='USDT→CLP';_me='USDT';_ment=_o['cantidad'];_ms='CLP';_msal=_o['total']
        try:
            _conn.execute("""INSERT INTO operaciones
                (fecha,hora,cliente,tipo_op,mon_entrada,monto_entrada,mon_salida,monto_salida,
                 tasa_cliente,tasa_referencia,usdt_equiv,diferencial,metodo,estado,observaciones,usuario_telegram)
                VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (_o['dt'].strftime('%Y-%m-%d'),_o['dt'].strftime('%H:%M'),
                 f'Binance CLP ({_o["contra"]})',_top,_me,_ment,_ms,_msal,
                 _o['precio'],_o['precio'],_o['cantidad'],0,
                 'Binance P2P CLP','Completada',f'Orden #{_o["num"]} | CLP',usuario))
            _cimp+=1
        except: pass

    _conn.commit(); _conn.close()
    _tgb=sum(s['gan_bs'] for s in _ses_saved)
    _tgu=sum(s['gan_u'] for s in _ses_saved)
    _tpend=sum(s['pend'] for s in _ses_saved if s['estado']=='Abierto')
    _tfees=sum(s['fees'] for s in _ses_saved)
    return {'sesiones':_ses_saved,'importadas_maker':_imp,'importadas_taker':_timp,
            'importadas_clp':_cimp,'omitidas':_omit,'errores':_err,
            'total_ganancia_bs':round(_tgb,2),'total_ganancia_u':round(_tgu,4),
            'usdt_pendiente':round(_tpend,4),'fees_total':round(_tfees,4)}

def formatear_resultado_inteligente(resultado):
    if 'error' in resultado: return f"❌ Error: {resultado['error']}"
    ses = resultado['sesiones']
    ses_mostrar = ses[-10:] if len(ses) > 10 else ses
    m = f"✅ *IMPORTACIÓN BINANCE C2C*\n\n"
    if len(ses) > 10: m += f"_{len(ses)} sesiones — mostrando últimas 10_\n\n"
    m += "📊 *SESIONES DE ARBITRAJE (Maker)*\n\n"
    for s in ses_mostrar:
        emoji = "🟡" if s['estado']=='Abierto' else "🟢"
        m += f"{emoji} *{s['sesion']}* — {s['fecha']} {s['hora_ini']}→{s['hora_fin']}\n"
        m += f"  Compras: `{s['compras']}` | Ventas: `{s['ventas']}`\n"
        m += f"  CPP: `{s['cpp']:.2f} BS` | Venta: `{s['pv']:.2f} BS`\n"
        if s['gan_bs'] != 0: m += f"  Ganancia: `{s['gan_bs']:.2f} BS` (`{s['gan_u']:.4f} USDT`)\n"
        if s['pend'] > 0.01: m += f"  ⚠️ Pendiente inventario: `{s['pend']:.4f} USDT`\n"
        m += "\n"
    m += "━━━━━━━━━━━━━━━━━━━━\n*TOTAL*\n"
    m += f"  Ganancia: `{resultado['total_ganancia_bs']:.2f} BS` (`{resultado['total_ganancia_u']:.4f} USDT`)\n"
    if resultado['usdt_pendiente'] > 0:
        m += f"  📦 USDT en inventario: `{resultado['usdt_pendiente']:.4f} USDT`\n"
    m += f"  Fees: `{resultado['fees_total']:.4f} USDT`\n\n"
    if resultado['importadas_taker'] > 0: m += f"🔄 Taker: `{resultado['importadas_taker']}`\n"
    if resultado['importadas_clp'] > 0: m += f"🇨🇱 CLP: `{resultado['importadas_clp']}`\n"
    total = resultado['importadas_maker']+resultado['importadas_taker']+resultado['importadas_clp']
    m += f"\n📥 Importadas: `{total}` | Omitidas: `{resultado['omitidas']}`"
    return m

# ══════════════════════════════════════════════════════════════════════
# CONVERSACIÓN /operacion
# ══════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════
# OPERACIÓN RÁPIDA — FORMULARIO DE UNA LÍNEA
# ══════════════════════════════════════════════════════════════════════
PLANTILLA_OP = """📋 NUEVA OPERACIÓN
━━━━━━━━━━━━━━━━━━━━
Fecha: hoy
Tipo: CLP-BS
Cliente: 
Monto entrada: 
Tasa: auto
Método: Copec Pay
Corresponsal: no
Entrega física: no
CXC pendiente: 0
━━━━━━━━━━━━━━━━━━━━
Tipos disponibles:
CLP-BS  BS-CLP  CLP-COP  COP-CLP
COP-BS  BS-COP  CLP-USDT  USDT-CLP
BS-USDT  USDT-BS  USD-CLP  CLP-USD
USD-BS  BS-USD  GIRO-INT"""

def normalizar_tipo(tipo_raw):
    """Convierte CLP-BS, CLP>BS, CLPBS etc a CLP→BS."""
    if not tipo_raw: return None
    t = tipo_raw.strip().upper()
    # Replace separators with arrow
    for sep in ['-', '>', ' A ', ' TO ']:
        t = t.replace(sep, '→')
    # Map common abbreviations
    alias = {
        'GIRO→INT': 'GIRO INT',
        'GIROINT': 'GIRO INT',
        'GIRO-INT': 'GIRO INT',
    }
    t = alias.get(t, t)
    # Check if valid
    if t in TIPOS_OP:
        return t
    # Try adding arrow if missing (e.g. CLPBS → CLP→BS)
    for op in TIPOS_OP:
        clean = op.replace('→', '')
        if t.replace('→','') == clean:
            return op
    return None

def parsear_operacion_rapida(texto):
    """Parsea el formulario de operación rápida y retorna datos o error."""
    lineas = texto.strip().split('\n')
    datos = {}
    errores = []

    for linea in lineas:
        if ':' not in linea: continue
        clave, _, valor = linea.partition(':')
        clave = clave.strip().lower()
        valor = valor.strip()
        if not valor or valor == '': continue

        if 'fecha' in clave:
            hoy = today_local()
            ayer = hoy - datetime.timedelta(days=1)
            if valor.lower() in ('hoy', 'today', 'h'):
                datos['fecha'] = str(hoy)
            elif valor.lower() in ('ayer', 'yesterday', 'a'):
                datos['fecha'] = str(ayer)
            else:
                for fmt_str in ('%d/%m/%Y','%Y-%m-%d','%d-%m-%Y'):
                    try:
                        dt = datetime.datetime.strptime(valor, fmt_str)
                        datos['fecha'] = dt.strftime('%Y-%m-%d')
                        break
                    except: pass
                if 'fecha' not in datos:
                    errores.append(f"Fecha inválida: `{valor}` — usa hoy, ayer o DD/MM/YYYY")

        elif 'tipo' in clave:
            tipo = normalizar_tipo(valor)
            if tipo:
                datos['tipo_op'] = tipo
                partes = tipo.replace('→','-').split('-')
                if len(partes) == 2:
                    datos['mon_entrada'] = partes[0]
                    datos['mon_salida'] = partes[1]
            else:
                errores.append(f"Tipo inválido: `{valor}` — ej: CLP-BS, BS-CLP")

        elif 'cliente' in clave:
            if valor:
                datos['cliente'] = valor
            else:
                errores.append("Cliente vacío")

        elif 'monto entrada' in clave or 'monto' in clave:
            try:
                datos['monto_entrada'] = float(valor.replace(',','.').replace(' ',''))
            except:
                errores.append(f"Monto inválido: `{valor}`")

        elif 'tasa' in clave:
            if valor.lower() in ('auto', 'automatica', 'automatico', 'a'):
                datos['tasa'] = 'auto'
            else:
                try:
                    datos['tasa'] = float(valor.replace(',','.'))
                except:
                    errores.append(f"Tasa inválida: `{valor}` — usa auto o número")

        elif 'método' in clave or 'metodo' in clave:
            datos['metodo'] = valor

        elif 'corresponsal' in clave:
            datos['corresponsal'] = '' if valor.lower() in ('no','ninguno','directo','n') else valor

        elif 'entrega' in clave or 'física' in clave or 'fisica' in clave:
            datos['entrega_fisica'] = valor.lower() in ('si','sí','s','yes')

        elif 'cxc' in clave:
            try:
                datos['cxc_pendiente'] = float(valor.replace(',','.')) if valor.lower() not in ('0','no','n') else 0
            except:
                datos['cxc_pendiente'] = 0

    return datos, errores

def procesar_op_rapida(chat_id, texto):
    """Procesa el formulario de operación rápida."""
    datos, errores = parsear_operacion_rapida(texto)

    # Validaciones básicas
    campos_req = ['fecha','tipo_op','cliente','monto_entrada']
    for campo in campos_req:
        if campo not in datos:
            if campo == 'fecha': errores.append("Falta: Fecha")
            elif campo == 'tipo_op': errores.append("Falta: Tipo")
            elif campo == 'cliente': errores.append("Falta: Cliente")
            elif campo == 'monto_entrada': errores.append("Falta: Monto entrada")

    if errores:
        m = "❌ *Errores en el formulario:*\n\n"
        for e in errores:
            m += f"• {e}\n"
        m += "\n_Corrige y envía de nuevo_"
        return m

    # Calcular tasa y monto salida
    t = get_ultima_tasa()
    tasa_ref = _tasa_sug(datos['tipo_op'], t)

    if datos.get('tasa') == 'auto' or 'tasa' not in datos:
        if tasa_ref:
            datos['tasa_cliente'] = tasa_ref
            datos['tasa_referencia'] = tasa_ref
            datos['monto_salida'] = _calc_sal(datos['tipo_op'], datos['monto_entrada'], tasa_ref)
            tasa_usada = f"{tasa_ref} (auto)"
        else:
            return "❌ No hay tasa automática disponible para este tipo.\nEscribe la tasa manualmente en el formulario."
    else:
        datos['tasa_cliente'] = datos['tasa']
        datos['tasa_referencia'] = tasa_ref or datos['tasa']
        datos['monto_salida'] = _calc_sal(datos['tipo_op'], datos['monto_entrada'], datos['tasa'])
        tasa_usada = str(datos['tasa'])

    # Snapshot tasas
    datos['snap_pat_bs'] = ((t.get('ban_bs_compra',0) or 0)+(t.get('ban_bs_venta',0) or 0))/2
    datos['snap_dol_obs'] = t.get('dolar_obs',0) or 0
    datos['snap_trm'] = t.get('trm',0) or 0
    datos['hora'] = hora_local()
    datos['usuario_telegram'] = str(chat_id)
    datos['estado'] = 'Completada'
    datos['traslado_bs'] = 0
    datos['encomienda_cop'] = 0
    datos['repartidor'] = 'Cristofer Ruiz'
    datos['usdt_equiv'] = calcular_usdt_equiv(
        datos.get('mon_entrada',''), datos.get('monto_entrada',0),
        datos['snap_pat_bs'], datos['snap_dol_obs'])

    if not datos.get('metodo'): datos['metodo'] = 'Copec Pay'
    if 'corresponsal' not in datos: datos['corresponsal'] = ''
    if 'cxc_pendiente' not in datos: datos['cxc_pendiente'] = 0

    # Verificar corresponsal antes de confirmar
    alerta_corresp = verificar_corresponsal(datos.get('corresponsal',''))
    if alerta_corresp:
        send(chat_id, alerta_corresp)

    # Verificar capital e impacto
    alertas_cap = verificar_capital_operacion(datos)
    msg_cap = msg_alerta_capital_op(alertas_cap, datos)

    # Guardar en conversaciones para confirmación
    conversaciones[chat_id] = {
        'paso': 'confirmar_rapido',
        'datos': datos,
        'tasa_usada': tasa_usada,
    }

    m = f"📋 *CONFIRMAR OPERACIÓN*\n━━━━━━━━━━━━━━━━━━━━\n"
    m += f"Fecha: `{datos['fecha']}`\n"
    m += f"Tipo: `{datos['tipo_op']}`\n"
    m += f"Cliente: `{datos['cliente']}`\n"
    m += f"Entrada: `{datos['monto_entrada']:,.2f} {datos.get('mon_entrada','')}`\n"
    m += f"Salida: `{datos['monto_salida']:,.2f} {datos.get('mon_salida','')}`\n"
    m += f"Tasa: `{tasa_usada}`\n"
    m += f"USDT equiv: `{datos['usdt_equiv']:.4f}`\n"
    m += f"Método: `{datos.get('metodo','')}`\n"
    if datos.get('corresponsal'): m += f"Corresponsal: `{datos['corresponsal']}`\n"
    if datos.get('cxc_pendiente',0) > 0: m += f"⚠️ CXC: `{datos['cxc_pendiente']:,.2f} {datos.get('mon_entrada','')}`\n"
    if msg_cap: m += f"\n{msg_cap}\n"
    m += f"━━━━━━━━━━━━━━━━━━━━\n*¿Confirmar?* `si` / `no`"
    return m


conversaciones = {}

def iniciar_operacion(chat_id):
    conversaciones[chat_id] = {'paso':-1,'datos':{
        'fecha':str(today_local()),
        'hora':hora_local(),
        'usuario_telegram':str(chat_id),'estado':'Completada',
        'traslado_bs':0,'encomienda_cop':0,'cxc_pendiente':0,'cxp_pendiente':0,
        'repartidor':'Cristofer Ruiz',
    }}
    tipos = "\n".join([f"  `{t}`" for t in TIPOS_OP])
    return f"💱 *NUEVA OPERACIÓN*\n\n¿Tipo de operación?\n\n{tipos}\n\n_/cancelar para salir_"

def procesar_conv(chat_id, texto):
    if chat_id not in conversaciones: return "No hay operación activa. Usa /op para iniciar."
    conv=conversaciones[chat_id]; paso=conv['paso']; datos=conv['datos']
    if texto.lower() in ('/cancelar','cancelar'):
        del conversaciones[chat_id]; return "❌ Operación cancelada."

    # Confirmación de operación rápida
    if paso == 'confirmar_rapido':
        if texto.lower() in ('si','sí','s','confirmar','ok'):
            op_id = guardar_operacion(datos)
            del conversaciones[chat_id]
            m  = f"✅ *Op #{op_id} registrada*\n"
            m += f"Cliente: `{datos['cliente']}`\n"
            m += f"Tipo: `{datos['tipo_op']}`\n"
            m += f"USDT: `{datos.get('usdt_equiv',0):.4f}`\n"
            if datos.get('cxc_pendiente',0)>0: m+=f"⚠️ CXC: `{datos['cxc_pendiente']:,.2f} {datos['mon_entrada']}`\n"
            # Mostrar desglose de ganancia si hay datos de inventario
            if datos.get('gan_comercial_usdt',0) > 0:
                m += f"\n💼 Ganancia comercial: `{datos['gan_comercial_usdt']:.4f} USDT`\n"
            if datos.get('gan_financiera_usdt',0) > 0:
                m += f"💰 Ganancia financiera: `{datos['gan_financiera_usdt']:.4f} USDT`\n"
                m += f"   (inventario CPP `{datos.get('cpp_usado',0):.2f} Bs`)\n"
            inv = get_inventario()
            if inv['cantidad'] > 0:
                gan_lat_bs, gan_lat_u = get_ganancia_latente()
                m += f"\n📦 Inventario: `{inv['cantidad']:.4f} USDT` | CPP: `{inv['cpp_bs']:.2f} Bs`\n"
                if gan_lat_u > 0:
                    m += f"📈 Ganancia latente: `{gan_lat_u:.4f} USDT`\n"
            m += "_Saldos actualizados_"
            return m
        del conversaciones[chat_id]; return "❌ Operación cancelada."

    if paso==-1:
        hoy = today_local(); ayer = hoy - datetime.timedelta(days=1)
        if texto.lower() in ('hoy','today','h'): datos['fecha'] = str(hoy)
        elif texto.lower() in ('ayer','yesterday','a'): datos['fecha'] = str(ayer)
        else:
            try:
                for fmt_str in ('%d/%m/%Y','%Y-%m-%d','%d-%m-%Y'):
                    try:
                        dt = datetime.datetime.strptime(texto, fmt_str)
                        datos['fecha'] = dt.strftime('%Y-%m-%d'); break
                    except: pass
                else: return "⚠️ Escribe *hoy*, *ayer*, o fecha como `05/06/2026`"
            except: return "⚠️ Fecha inválida."
        conv['paso'] = 0
        tipos = "\n".join([f"  `{t}`" for t in TIPOS_OP])
        return f"💱 *NUEVA OPERACIÓN*\nFecha: `{datos['fecha']}`\n\n¿Tipo de operación?\n\n{tipos}\n\n_/cancelar para salir_"

    if paso==0:
        if texto not in TIPOS_OP: return "⚠️ Selecciona un tipo válido de la lista."
        datos['tipo_op']=texto
        partes=texto.replace('→','-').split('-')
        if len(partes)==2: datos['mon_entrada']=partes[0]; datos['mon_salida']=partes[1]
        conv['paso']=1; return "👤 ¿Nombre del cliente?"
    elif paso==1:
        datos['cliente']=texto; conv['paso']=2
        return f"💰 ¿Cuánto {datos.get('mon_entrada','?')} entrega el cliente?"
    elif paso==2:
        try:
            datos['monto_entrada']=float(texto.replace(',','.'))
            t=get_ultima_tasa(); tasa_sug=_tasa_sug(datos['tipo_op'],t)
            if tasa_sug:
                datos['_tasa_sug']=tasa_sug
                mto_sal=_calc_sal(datos['tipo_op'],datos['monto_entrada'],tasa_sug)
                datos['_mto_sal_sug']=mto_sal; conv['paso']=3
                return f"📊 Tasa sugerida: `{tasa_sug}`\nCliente recibe: `{mto_sal:,.2f} {datos.get('mon_salida','?')}`\n\nEscribe *usar* para confirmar o ingresa otra tasa."
            conv['paso']=3; return f"💰 ¿Cuánto {datos.get('mon_salida','?')} recibe el cliente?"
        except: return "⚠️ Ingresa un número válido."
    elif paso==3:
        if texto.lower() in ('usar','si','sí','ok','s'):
            datos['tasa_cliente']=datos.get('_tasa_sug',0)
            datos['monto_salida']=datos.get('_mto_sal_sug',0)
        else:
            try:
                val=float(texto.replace(',','.')); datos['tasa_cliente']=val
                datos['monto_salida']=_calc_sal(datos['tipo_op'],datos['monto_entrada'],val)
            except: return "⚠️ Ingresa la tasa como número o escribe *usar*"
        t=get_ultima_tasa()
        datos['snap_pat_bs']=((t.get('ban_bs_compra',0) or 0)+(t.get('ban_bs_venta',0) or 0))/2
        datos['snap_dol_obs']=t.get('dolar_obs',0) or 0
        datos['snap_trm']=t.get('trm',0) or 0
        datos['usdt_equiv']=calcular_usdt_equiv(datos['mon_entrada'],datos['monto_entrada'],datos['snap_pat_bs'],datos['snap_dol_obs'])
        conv['paso']=4
        return "💳 ¿Método de pago?\n`Copec Pay` | `Bancolombia` | `Nequi` | `Efectivo` | `Airtm` | `Binance` | `Otro`"
    elif paso==4:
        datos['metodo']=texto; conv['paso']=5
        ops="\n".join([f"  `{c}`" for c in CORRESPONSALES])
        return f"🏦 ¿Corresponsal?\n{ops}\n\n_O escribe *no* si fue directo_"
    elif paso==5:
        datos['corresponsal']='' if texto.lower() in ('no','ninguno','directo') else texto
        conv['paso']=6; return "📦 ¿Hubo entrega física? (*si* o *no*)"
    elif paso==6:
        if texto.lower() in ('si','sí','s'):
            datos['entrega_fisica']='Sí'; conv['paso']=61
            return "💸 ¿Costo de traslado? (en COP o BS, o *0* si fue gratis)"
        datos['entrega_fisica']='No'; conv['paso']=7
        return "💳 ¿El cliente quedó debiendo? (*no*, *si*, o el monto)"
    elif paso==61:
        try:
            monto=float(texto.replace(',','.'))
            if monto>0:
                if monto>10000: datos['encomienda_cop']=monto
                else: datos['traslado_bs']=monto
            conv['paso']=7; return "💳 ¿El cliente quedó debiendo? (*no*, *si*, o el monto)"
        except: return "⚠️ Ingresa un número o *0*"
    elif paso==7:
        if texto.lower() in ('no','n','0'): datos['cxc_pendiente']=0
        elif texto.lower() in ('si','sí','s'):
            conv['paso']=71; return f"¿Cuánto {datos.get('mon_entrada','?')} debe el cliente?"
        else:
            try: datos['cxc_pendiente']=float(texto.replace(',','.'))
            except: return "Responde *si*, *no* o el monto."
        conv['paso']=8; return _resumen(datos)
    elif paso==71:
        try: datos['cxc_pendiente']=float(texto.replace(',','.')); conv['paso']=8; return _resumen(datos)
        except: return "⚠️ Ingresa el monto como número."
    elif paso==8:
        if texto.lower() in ('si','sí','s','confirmar','ok'):
            op_id=guardar_operacion(datos); del conversaciones[chat_id]
            m = f"✅ *Op #{op_id} registrada*\nCliente: `{datos['cliente']}`\nTipo: `{datos['tipo_op']}`\nUSDT: `{datos.get('usdt_equiv',0):.4f}`\n"
            if datos.get('cxc_pendiente',0)>0: m+=f"⚠️ CXC: `{datos['cxc_pendiente']:,.2f} {datos['mon_entrada']}`\n"
            m += "_Saldos actualizados_"; return m
        del conversaciones[chat_id]; return "❌ Operación cancelada."
    return "⚠️ Algo salió mal. Usa /operacion para reiniciar."

def _resumen(datos):
    m = f"📋 *CONFIRMAR*\n━━━━━━━━━━━━━━━━━━━━\n"
    m += f"Tipo: `{datos.get('tipo_op')}`\nCliente: `{datos.get('cliente')}`\n"
    m += f"Entrada: `{datos.get('monto_entrada',0):,.2f} {datos.get('mon_entrada')}`\n"
    m += f"Salida: `{datos.get('monto_salida',0):,.2f} {datos.get('mon_salida')}`\n"
    m += f"USDT: `{datos.get('usdt_equiv',0):.4f}`\n"
    if datos.get('cxc_pendiente',0)>0: m+=f"⚠️ CXC: `{datos['cxc_pendiente']:,.2f}`\n"
    m += "━━━━━━━━━━━━━━━━━━━━\n¿Confirmar? *si* o *no*"
    return m

def _tasa_sug(tipo, t):
    mapa={'CLP→BS':t.get('tasa_gsa_clp_bs'),'BS→CLP':t.get('tasa_gsa_bs_clp'),
          'CLP→COP':t.get('tasa_gsa_clp_cop'),'COP→CLP':t.get('tasa_gsa_cop_clp'),
          'COP→BS':t.get('tasa_gsa_cop_bs'),'BS→COP':t.get('tasa_gsa_bs_cop'),
          'CLP→USDT':t.get('clp_compra'),'USDT→CLP':t.get('clp_venta'),
          'BS→USDT':t.get('ban_bs_compra'),'USDT→BS':t.get('ban_bs_venta'),
          'USD→CLP':t.get('dolar_obs'),'USD→BS':t.get('ban_bs_venta')}
    return mapa.get(tipo)

def _calc_sal(tipo, mto, tasa):
    if not tasa: return 0
    if tipo in ('CLP→BS','COP→BS','USD→BS','CLP→COP','BS→CLP','BS→COP','BS→USD','USDT→CLP','USDT→BS'): return round(mto*tasa,2)
    if tipo in ('CLP→USDT','BS→USDT'): return round(mto/tasa,4)
    if tipo in ('COP→CLP',): return round(mto/tasa,2)
    return round(mto*tasa,2)

def hay_conv_activa(chat_id): return chat_id in conversaciones

# ══════════════════════════════════════════════════════════════════════
# CSV EXPORT
# ══════════════════════════════════════════════════════════════════════
def exportar_csv():
    os.makedirs(CSV_EXPORT_PATH, exist_ok=True)
    conn = sqlite3.connect(DB_PATH); conn.row_factory = sqlite3.Row
    for tabla in ['tasas','operaciones','saldos','clientes','gastos','tesoreria',
                  'cuentas_pendientes','precios_historicos','binance_sesiones']:
        try:
            rows=conn.execute(f"SELECT * FROM {tabla}").fetchall()
            if not rows: continue
            ruta=os.path.join(CSV_EXPORT_PATH,f"{tabla}.csv")
            with open(ruta,'w',newline='',encoding='utf-8-sig') as f:
                w=csv.DictWriter(f,fieldnames=rows[0].keys()); w.writeheader()
                w.writerows([dict(r) for r in rows])
        except: pass
    conn.close()
    print(f"✅ CSV exportados — {now_local().strftime('%H:%M:%S')}")

# ══════════════════════════════════════════════════════════════════════
# TELEGRAM
# ══════════════════════════════════════════════════════════════════════
BASE_URL = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}"

def send(chat_id, text, parse_mode="Markdown"):
    try:
        requests.post(f"{BASE_URL}/sendMessage", json={
            "chat_id":chat_id,"text":text,
            "parse_mode":parse_mode,"disable_web_page_preview":True}, timeout=10)
    except Exception as e: print(f"Error send: {e}")

def get_updates(offset=0):
    try:
        r=requests.get(f"{BASE_URL}/getUpdates",params={"offset":offset,"timeout":30},timeout=35)
        return r.json().get("result",[])
    except: return []

def download_file(file_id):
    try:
        r = requests.get(f"{BASE_URL}/getFile", params={"file_id": file_id}, timeout=10)
        file_path = r.json()["result"]["file_path"]
        r2 = requests.get(f"https://api.telegram.org/file/bot{TELEGRAM_BOT_TOKEN}/{file_path}", timeout=30)
        return r2.content
    except Exception as e:
        print(f"Error descargando archivo: {e}"); return None

# ══════════════════════════════════════════════════════════════════════
# PROCESADOR DE COMANDOS
# ══════════════════════════════════════════════════════════════════════
ultimo_offset=0
western_rate=None
ultimo_datos={}
ultimo_precio_bs_venta=0
ultimo_precio_clp=0
esperando_importar={}

def procesar(chat_id, texto):
    global western_rate, ultimo_datos

    if hay_conv_activa(chat_id) and not texto.startswith('/'):
        send(chat_id, procesar_conv(chat_id, texto)); return

    # Detectar formulario de operación rápida (contiene "Fecha:" y "Tipo:" y "Cliente:")
    if 'Fecha:' in texto and 'Tipo:' in texto and 'Cliente:' in texto and not texto.startswith('/'):
        send(chat_id, procesar_op_rapida(chat_id, texto)); return

    if chat_id in esperando_importar and not texto.startswith('/'):
        archivo=texto.strip()
        ruta=os.path.abspath(os.path.join(os.path.dirname(DB_PATH), archivo))
        send(chat_id, f"⏳ Procesando `{archivo}`...")
        resultado=importar_c2c_inteligente(ruta, DB_PATH, str(chat_id))
        send(chat_id, formatear_resultado_inteligente(resultado))
        del esperando_importar[chat_id]; return

    partes=texto.split(); cmd=partes[0].lower() if partes else ''

    # ── TASAS ──
    if cmd=='/tasas':
        send(chat_id,"⏳ Consultando tasas...")
        datos=consultar_y_guardar(western_rate); ultimo_datos=datos
        send(chat_id, construir_mensaje(datos))

    elif cmd=='/western':
        if len(partes)>=2:
            try:
                western_rate=float(partes[1].replace(',','.'))
                set_config('western_actualizado_hoy', str(today_local()))
                send(chat_id,f"✅ Western: `{western_rate}` — Recordatorio cancelado para hoy.")
            except: send(chat_id,"Uso: /western 0.0042")
        else: send(chat_id,"Uso: `/western 0.0042`")

    elif cmd=='/limites':
        t=ultimo_datos or {}
        m=f"📐 *LÍMITES*\nCLP/BS: `{t.get('limite_clp_bs','N/D')}`\nCLP/COP: `{t.get('limite_clp_cop','N/D')}`"
        send(chat_id,m)

    # ── MERCADO ──
    elif cmd=='/mercado':
        send(chat_id,"⏳ Consultando mercado en vivo...")
        send(chat_id, msg_mercado_completo())

    # ── PATRÓN ──
    elif cmd=='/patron':
        sub = partes[1].lower() if len(partes)>1 else 'bs'
        if sub == 'bs':
            resultado = analizar_patron_bs()
            if resultado[0] is None:
                send(chat_id, resultado[1]); return
            resumen, mejor, peor = resultado
            m = "📊 *PATRÓN HISTÓRICO BS*\n━━━━━━━━━━━━━━━━━━━━\n\n"
            for hora, d in list(resumen.items())[:12]:
                bar = "🟢" if d['spread_prom'] >= SPREAD_MIN_ALERTA else "🔴"
                m += f"`{hora}` {bar} spread `{d['spread_prom']:.1f}` | venta `{d['venta_prom']:.2f} Bs`\n"
            m += f"\n🏆 Mejor hora: `{mejor[0]}` (spread prom `{mejor[1]['spread_prom']:.1f} Bs`)\n"
            m += f"📉 Peor hora: `{peor[0]}` (spread prom `{peor[1]['spread_prom']:.1f} Bs`)\n"
            send(chat_id, m)
        elif sub == 'semana':
            send(chat_id, generar_reporte_semanal())
        elif sub == 'hoy':
            send(chat_id, generar_reporte_diario())
        else:
            send(chat_id, "Uso: `/patron bs` | `/patron hoy` | `/patron semana`")

    # ── SIMULAR ──
    elif cmd=='/simular':
        if len(partes) >= 3:
            try:
                moneda = partes[1].upper()
                monto = float(partes[2].replace(',','.'))
                if moneda not in ('CLP','BS','USDT','BS'):
                    send(chat_id, "Monedas: CLP, BS, USDT"); return
                send(chat_id, "⏳ Calculando...")
                send(chat_id, simular_operacion(moneda, monto))
            except Exception as e:
                send(chat_id, f"Uso: `/simular CLP 500000`\nError: {e}")
        else:
            send(chat_id, "Uso: `/simular CLP 500000` o `/simular BS 300000`")

    # ── CAPITAL ──
    elif cmd=='/capital':
        send(chat_id, msg_capital())
        ocioso = calcular_capital_ocioso()
        if ocioso['pct_ocioso'] > 30:
            m_oc = f"\n💤 *CAPITAL OCIOSO: `{ocioso['pct_ocioso']:.0f}%`*\n"
            m_oc += f"`{ocioso['capital_ocioso']:.2f} USDT` sin mover hoy\n"
            if ocioso['pct_ocioso'] > 60:
                m_oc += "⚠️ Más del 60% inactivo — considera operar más\n"
            m_oc += "_/kpi para análisis completo_"
            send(chat_id, m_oc)

    # ── UMBRAL ──
    elif cmd=='/umbral':
        if len(partes) >= 3:
            try:
                cuenta = partes[1].upper()
                monto = float(partes[2].replace(',','.'))
                if cuenta not in NOMBRES_CUENTAS:
                    send(chat_id, f"Cuenta no válida. Opciones:\n" + "\n".join([f"`{k}`" for k in NOMBRES_CUENTAS.keys()])); return
                moneda = get_moneda(cuenta)
                conn = get_conn()
                conn.execute("""INSERT INTO umbrales_liquidez (cuenta,umbral_minimo,moneda)
                    VALUES (?,?,?) ON CONFLICT(cuenta) DO UPDATE SET umbral_minimo=excluded.umbral_minimo""",
                    (cuenta, monto, moneda))
                conn.commit(); conn.close()
                nombre = NOMBRES_CUENTAS[cuenta]
                send(chat_id, f"✅ Umbral configurado\n`{nombre}`: mínimo `{monto:,.2f} {moneda}`")
            except: send(chat_id, "Uso: `/umbral CLP_COPEC_PAY 100000`")
        else:
            conn = get_conn()
            rows = conn.execute("SELECT * FROM umbrales_liquidez WHERE activo=1").fetchall()
            conn.close()
            if not rows: send(chat_id, "Sin umbrales configurados.\nUso: `/umbral CLP_COPEC_PAY 100000`"); return
            m = "⚙️ *UMBRALES CONFIGURADOS*\n\n"
            for r in rows:
                nombre = NOMBRES_CUENTAS.get(r['cuenta'], r['cuenta'])
                m += f"`{nombre}`: mín `{r['umbral_minimo']:,.2f} {r['moneda']}`\n"
            send(chat_id, m)

    # ── CLIENTES ──
    elif cmd=='/clientes':
        sub = partes[1].lower() if len(partes)>1 else 'top'
        if sub == 'top': send(chat_id, msg_clientes_top())
        elif sub == 'rentabilidad': send(chat_id, msg_clientes_top_rentabilidad())
        elif sub == 'riesgo': send(chat_id, msg_clientes_riesgo())
        else: send(chat_id, "Uso: `/clientes top` | `/clientes rentabilidad` | `/clientes riesgo`")

    elif cmd=='/oportunidades':
        dias = int(partes[1]) if len(partes) > 1 and partes[1].isdigit() else 7
        send(chat_id, msg_oportunidades(dias))

    elif cmd=='/resumen_corresponsal':
        if len(partes) >= 2:
            nombre = ' '.join(partes[1:])
            send(chat_id, msg_resumen_corresponsal(nombre))
        else:
            ops_lista = "\n".join([f"  `{c}`" for c in CORRESPONSALES])
            send(chat_id, f"Uso: `/resumen_corresponsal Bancolombia C1`\n\nCorresponsales:\n{ops_lista}")

    # ── ANUNCIO GSA ──
    elif cmd=='/mianuncio':
        send(chat_id, "⏳ Buscando tu anuncio en Binance...")
        send(chat_id, msg_estado_anuncio_gsa())

    # ── INVENTARIO ──
    elif cmd=='/inventario':
        inv = get_inventario()
        gan_lat_bs, gan_lat_u = get_ganancia_latente()
        t = get_ultima_tasa()
        precio_actual = t.get('ban_bs_venta', 0) or 0
        m = f"📦 *INVENTARIO USDT*\n━━━━━━━━━━━━━━━━━━━━\n\n"
        if inv['cantidad'] > 0:
            m += f"Cantidad: `{inv['cantidad']:.4f} USDT`\n"
            m += f"CPP: `{inv['cpp_bs']:.2f} Bs` por USDT\n"
            m += f"Costo total: `{inv['cantidad'] * inv['cpp_bs']:,.2f} Bs`\n\n"
            if precio_actual:
                m += f"Precio mercado actual: `{precio_actual:.2f} Bs`\n"
                if gan_lat_u > 0:
                    m += f"📈 Ganancia latente: `{gan_lat_bs:,.2f} Bs` (`{gan_lat_u:.4f} USDT`)\n"
                elif gan_lat_u < 0:
                    m += f"📉 Pérdida latente: `{abs(gan_lat_bs):,.2f} Bs` (`{abs(gan_lat_u):.4f} USDT`)\n"
                else:
                    m += f"⚪ Sin ganancia latente aún\n"
        else:
            m += f"Sin inventario USDT actualmente\n"
            m += f"_Las compras en Binance se registran aquí_"
        send(chat_id, m)

    # ── LEDGER v7 ──
    elif cmd=='/cierredia':
        sub = partes[1].lower() if len(partes) > 1 else ''
        forzar = sub == 'forzar'
        send(chat_id, "⏳ Ejecutando cierre diario...")
        ok, msg = ejecutar_cierre_diario(usuario=str(chat_id), forzar=forzar)
        send(chat_id, msg)

    elif cmd=='/traslado':
        if len(partes) >= 4:
            try:
                origen = partes[1].upper()
                destino = partes[2].upper()
                monto = float(partes[3].replace(',','.'))
                desc = ' '.join(partes[4:]) if len(partes) > 4 else f"Traslado {origen} a {destino}"
                if origen not in NOMBRES_CUENTAS or destino not in NOMBRES_CUENTAS:
                    send(chat_id, "Cuenta no valida."); return
                lid = registrar_traslado(origen, destino, monto, desc, str(chat_id))
                send(chat_id, f"✅ *Traslado registrado*\n`{NOMBRES_CUENTAS[origen]}` → `{NOMBRES_CUENTAS[destino]}`\nMonto: `{monto:,.2f}`\nLedger ID: `{lid}`")
            except Exception as e:
                send(chat_id, f"Uso: /traslado ORIGEN DESTINO MONTO descripcion\nError: {e}")
        else:
            send(chat_id, "Uso: /traslado BS_BANESCO BS_MERCANTIL 50000 descripcion")

    elif cmd=='/ajuste':
        if len(partes) >= 4:
            try:
                cuenta = partes[1].upper()
                monto = float(partes[2].replace(',','.'))
                motivo = ' '.join(partes[3:])
                if cuenta not in NOMBRES_CUENTAS:
                    send(chat_id, "Cuenta no valida."); return
                conn = get_conn()
                saldo_ant = conn.execute("SELECT saldo FROM saldos WHERE cuenta=?", (cuenta,)).fetchone()
                saldo_anterior = saldo_ant['saldo'] if saldo_ant else 0
                conn.execute("UPDATE saldos SET saldo=saldo+?,ultima_actualizacion=CURRENT_TIMESTAMP WHERE cuenta=?", (monto, cuenta))
                conn.commit(); conn.close()
                moneda = get_moneda(cuenta)
                lid = ledger_insert('AJUSTE_AUDITADO', cuenta if monto>0 else 'DIFERENCIAS',
                    'DIFERENCIAS' if monto>0 else cuenta, moneda, abs(monto),
                    descripcion=f"Ajuste: {motivo} (anterior: {saldo_anterior:,.2f})", usuario=str(chat_id))
                send(chat_id, f"✅ *Ajuste registrado*\n`{NOMBRES_CUENTAS[cuenta]}`: `{monto:+,.2f}`\nMotivo: _{motivo}_\nLedger ID: `{lid}`")
            except Exception as e:
                send(chat_id, f"Uso: /ajuste BS_BANESCO -500 Fee blockchain\nError: {e}")
        else:
            send(chat_id, "Uso: /ajuste CUENTA MONTO motivo obligatorio")

    elif cmd=='/reconstruir':
        if len(partes) >= 2:
            try:
                import datetime as dt2
                fecha_iso = dt2.datetime.strptime(partes[1], '%d/%m/%Y').strftime('%Y-%m-%d')
                send(chat_id, reconstruir_dia(fecha_iso))
            except:
                send(chat_id, "Uso: /reconstruir 08/06/2026")
        else:
            send(chat_id, "Uso: /reconstruir DD/MM/YYYY")

    elif cmd=='/patrimonio':
        if len(partes) >= 2 and partes[1].lower() == 'inicial':
            saldos_actuales = get_saldos()
            registrar_patrimonio_inicial(saldos_actuales)
            send(chat_id, f"✅ *Patrimonio inicial registrado*\nFecha: `{today_local()}`")
        else:
            conn2 = get_conn()
            rows2 = conn2.execute(
                "SELECT cuenta, moneda, saldo FROM patrimonio_inicial ORDER BY cuenta"
            ).fetchall()
            conn2.close()
            if not rows2:
                send(chat_id, "Sin patrimonio inicial.\nUsa `/patrimonio inicial` para registrar.")
            else:
                m2 = "🏛️ *PATRIMONIO INICIAL*\n\n"
                for r2 in rows2:
                    nombre2 = NOMBRES_CUENTAS.get(r2['cuenta'], r2['cuenta'])
                    m2 += f"`{nombre2}`: `{r2['saldo']:,.2f} {r2['moneda']}`\n"
                send(chat_id, m2)

    elif cmd=='/eliminar_operaciones':
        if len(partes) < 2:
            send(chat_id, "Uso:\n/eliminar_operaciones 01/06/2026\n/eliminar_operaciones todo")
            return

        sub = partes[1].lower()

        # Pedir confirmación si no viene "confirmar"
        confirmar = len(partes) >= 3 and partes[2].lower() == 'confirmar'

        if not confirmar:
            if sub == 'todo':
                send(chat_id,
                    "⚠️ *ELIMINAR TODAS LAS OPERACIONES*\n\n"
                    "Esto borrará:\n"
                    "• Todas las operaciones\n"
                    "• Todos los cierres diarios\n"
                    "• Todo el inventario\n"
                    "• Todos los saldos\n\n"
                    "El ledger mantendrá el registro de esta acción.\n\n"
                    "Para confirmar escribe:\n"
                    "`/eliminar_operaciones todo confirmar`")
            else:
                try:
                    import datetime as dt2
                    fecha_iso = dt2.datetime.strptime(sub, '%d/%m/%Y').strftime('%Y-%m-%d')
                    conn = get_conn()
                    cnt = conn.execute("SELECT COUNT(*) as c FROM operaciones WHERE fecha=?", (fecha_iso,)).fetchone()['c']
                    conn.close()
                    send(chat_id,
                        f"⚠️ *ELIMINAR OPERACIONES DEL {sub}*\n\n"
                        f"Se eliminarán `{cnt}` operaciones.\n"
                        f"El ledger registrará esta acción.\n\n"
                        f"Para confirmar escribe:\n"
                        f"`/eliminar_operaciones {sub} confirmar`")
                except:
                    send(chat_id, "Fecha inválida. Usa formato DD/MM/YYYY")
            return

        # Ejecutar eliminación
        conn = get_conn()
        if sub == 'todo':
            # Contar todo
            cnt_ops = conn.execute("SELECT COUNT(*) as c FROM operaciones").fetchone()['c']
            cnt_ses = conn.execute("SELECT COUNT(*) as c FROM binance_sesiones").fetchone()['c']

            # Registrar en ledger antes de borrar
            ledger_insert('AJUSTE_AUDITADO', 'APERTURA', 'APERTURA', 'USDT', 0,
                descripcion=f"ELIMINACION TOTAL: {cnt_ops} operaciones, {cnt_ses} sesiones Binance",
                usuario=str(chat_id))

            # Borrar todo
            conn.execute("DELETE FROM operaciones")
            conn.execute("DELETE FROM binance_sesiones")
            conn.execute("DELETE FROM cierres_diarios")
            conn.execute("DELETE FROM costos_operacion")
            conn.execute("DELETE FROM inventario_usdt")
            conn.execute("DELETE FROM saldos_iniciales")
            conn.execute("DELETE FROM patrimonio_inicial")
            conn.execute("UPDATE saldos SET saldo=0, ultima_actualizacion=CURRENT_TIMESTAMP")
            conn.commit(); conn.close()

            # Reset inventario
            init_inventario()

            send(chat_id,
                f"✅ *ELIMINACIÓN COMPLETA*\n\n"
                f"Eliminadas: `{cnt_ops}` operaciones\n"
                f"Sesiones Binance: `{cnt_ses}`\n"
                f"Saldos reseteados a 0\n\n"
                f"Registrado en ledger.\n"
                f"Usa /saldo_inicial para configurar saldos iniciales.")
        else:
            try:
                import datetime as dt2
                fecha_iso = dt2.datetime.strptime(sub, '%d/%m/%Y').strftime('%Y-%m-%d')
                cnt = conn.execute("SELECT COUNT(*) as c FROM operaciones WHERE fecha=?", (fecha_iso,)).fetchone()['c']

                # Registrar en ledger
                ledger_insert('AJUSTE_AUDITADO', 'APERTURA', 'APERTURA', 'USDT', 0,
                    descripcion=f"ELIMINACION fecha {fecha_iso}: {cnt} operaciones",
                    usuario=str(chat_id))

                # Borrar
                conn.execute("DELETE FROM operaciones WHERE fecha=?", (fecha_iso,))
                conn.execute("DELETE FROM costos_operacion WHERE fecha=?", (fecha_iso,))
                conn.commit(); conn.close()

                send(chat_id,
                    f"✅ *ELIMINADAS {cnt} OPERACIONES*\n"
                    f"Fecha: `{sub}`\n"
                    f"Registrado en ledger.\n\n"
                    f"_Los saldos pueden necesitar ajuste manual con /ajuste_")
            except Exception as e:
                conn.close()
                send(chat_id, f"Error: {e}")

    # ── KPIs ──
    elif cmd=='/kpi':
        send(chat_id, "⏳ Calculando KPIs...")
        send(chat_id, msg_kpi())

    elif cmd=='/profundidad':
        send(chat_id, "⏳ Analizando profundidad del mercado...")
        send(chat_id, msg_profundidad_mercado())

    # ── META DIARIA ──
    elif cmd=='/meta':
        if len(partes) >= 2:
            try:
                meta = float(partes[1].replace(',','.'))
                set_config(META_DIARIA_KEY, str(meta))
                send(chat_id, f"✅ Meta diaria configurada: `{meta:.2f} USDT`")
            except: send(chat_id, "Uso: /meta 10")
        else:
            send(chat_id, msg_meta_status())

    elif cmd=='/ranking':
        send(chat_id, "⏳ Calculando oportunidades...")
        send(chat_id, msg_prioridad_oportunidades())

    # ── SESION ARBITRAJE ──
    elif cmd=='/sesion':
        if len(partes) >= 2:
            sub = partes[1].lower()
            if sub == 'status':
                send(chat_id, msg_sesion_status(chat_id))
            elif sub == 'cerrar':
                send(chat_id, cerrar_sesion(chat_id))
            elif sub == 'compra' and len(partes) >= 4:
                try:
                    usdt = float(partes[2].replace(',','.'))
                    precio = float(partes[3].replace(',','.'))
                    send(chat_id, registrar_compra_sesion(chat_id, usdt, precio))
                except: send(chat_id, "Uso: /sesion compra 50 763.50")
            elif sub == 'venta' and len(partes) >= 4:
                try:
                    usdt = float(partes[2].replace(',','.'))
                    precio = float(partes[3].replace(',','.'))
                    send(chat_id, registrar_venta_sesion(chat_id, usdt, precio))
                except: send(chat_id, "Uso: /sesion venta 50 778.00")
            elif sub in ('bs','clp','usdt') and len(partes) >= 3:
                try:
                    capital = float(partes[2].replace(',','.'))
                    send(chat_id, iniciar_sesion_arbitraje(chat_id, sub.upper(), capital))
                except: send(chat_id, "Uso: /sesion BS 200000")
            else:
                m_ses = "*SESION DE ARBITRAJE*\n\n"
                m_ses += "Iniciar: /sesion BS 200000\n"
                m_ses += "Estado: /sesion status\n"
                m_ses += "Registrar compra: /sesion compra 50 763.50\n"
                m_ses += "Registrar venta: /sesion venta 50 778.00\n"
                m_ses += "Cerrar: /sesion cerrar"
                send(chat_id, m_ses)
        else:
            send(chat_id, "Uso: /sesion BS 200000 para iniciar")

    # ── OPERACIONES ──
    elif cmd in ('/op', '/operacion'):
        send(chat_id, PLANTILLA_OP)
    elif cmd=='/cancelar':
        if hay_conv_activa(chat_id): procesar_conv(chat_id,'/cancelar'); send(chat_id,"❌ Cancelado.")
        elif chat_id in esperando_importar: del esperando_importar[chat_id]; send(chat_id,"❌ Cancelado.")
        else: send(chat_id,"No hay operación activa.")

    elif cmd=='/operaciones':
        ops=get_operaciones_hoy()
        if not ops: send(chat_id,"Sin operaciones hoy."); return
        m=f"📋 *HOY ({len(ops)} ops)*\n\n"
        for op in ops[-10:]:
            m+=f"*#{op['id']}* {op['tipo_op']} — {op['cliente']}\n  `{op['monto_entrada']:,.2f} {op['mon_entrada']}` → `{op['monto_salida']:,.2f} {op['mon_salida']}`\n\n"
        send(chat_id,m)

    elif cmd=='/ultima':
        ops=get_operaciones_hoy()
        if not ops: send(chat_id,"Sin operaciones hoy."); return
        op=ops[-1]
        m=f"📋 *ÚLTIMA*\n#{op['id']} | {op['tipo_op']}\nCliente: {op['cliente']}\nEntrada: `{op['monto_entrada']:,.2f} {op['mon_entrada']}`\nSalida: `{op['monto_salida']:,.2f} {op['mon_salida']}`\nEstado: {op['estado']}"
        send(chat_id,m)

    # ── SALDOS ──
    elif cmd in ('/saldo','/caja','/posicion'): send(chat_id, msg_saldos())
    elif cmd=='/dashboard': send(chat_id, msg_dashboard())
    elif cmd=='/resultado':
        d=get_resultados_mes()
        send(chat_id,f"💵 *P&L MES*\nOps: `{d['ops']}` | Vol: `{d['volumen']:.4f}`\nGanancia: `{d['ganancia_neta']:.4f} USDT`")
    elif cmd=='/ganancia':
        d=get_resultados_hoy()
        send(chat_id,f"💵 *GANANCIA HOY*\nOps: `{d['ops']}` | Vol: `{d['volumen']:.4f}`\nGanancia: `{d['ganancia_neta']:.4f} USDT`")
    elif cmd=='/cxc': send(chat_id, msg_cxc())
    elif cmd=='/cxp': send(chat_id, msg_cxp())
    elif cmd=='/cobrado':
        if len(partes)>=2:
            try: marcar_pagado(int(partes[1])); send(chat_id,f"✅ CXC #{partes[1]} marcada cobrada.")
            except: send(chat_id,"Uso: /cobrado ID")
    elif cmd=='/pagado':
        if len(partes)>=2:
            try: marcar_pagado(int(partes[1])); send(chat_id,f"✅ CXP #{partes[1]} marcada pagada.")
            except: send(chat_id,"Uso: /pagado ID")

    elif cmd=='/setsaldo':
        if len(partes)>=3:
            try:
                cuenta=partes[1].upper(); monto=float(partes[2].replace(',','.'))
                if cuenta not in NOMBRES_CUENTAS:
                    send(chat_id,f"⚠️ Cuenta no válida."); return
                set_saldo(cuenta,monto); nombre=NOMBRES_CUENTAS[cuenta]
                send(chat_id,f"✅ `{nombre}`: `{monto:,.2f}`")
            except: send(chat_id,"Uso: `/setsaldo BS_BANESCO 125430`")
        else:
            send(chat_id,"Uso: `/setsaldo CUENTA MONTO`\n\n"+"\n".join([f"`{k}`" for k in NOMBRES_CUENTAS]))

    elif cmd=='/saldo_inicial':
        if len(partes)>=3:
            try:
                cuenta=partes[1].upper(); monto=float(partes[2].replace(',','.'))
                if cuenta not in NOMBRES_CUENTAS:
                    send(chat_id,"⚠️ Cuenta no válida."); return
                set_saldo_inicial(cuenta,monto); nombre=NOMBRES_CUENTAS[cuenta]
                send(chat_id,f"✅ Saldo inicial `{nombre}`: `{monto:,.2f}`")
            except: send(chat_id,"Uso: `/saldo_inicial BS_BANESCO 303581.42`")
        else: send(chat_id,"Uso: `/saldo_inicial CUENTA MONTO`")

    elif cmd=='/saldos_iniciales': send(chat_id, msg_saldos_iniciales())

    # ── SISTEMA ──
    elif cmd=='/importar':
        esperando_importar[chat_id]=True
        send(chat_id,"📂 *IMPORTAR BINANCE C2C*\n\nEscribe el nombre del archivo .xlsx o súbelo directamente.\n_/cancelar para salir_")

    elif cmd=='/sync':
        send(chat_id,"⏳ Exportando CSV..."); exportar_csv()
        send(chat_id,"✅ CSV exportados para Excel.")

    elif cmd=='/version':
        supa_st = "✅ Conectado" if USE_SUPABASE else "❌ Desactivado"
        send(chat_id,f"🤖 *GSA Cambios Bot v6.0*\nInteligencia de Mercado activa\nSupabase: {supa_st}\nZona horaria: UTC{UTC_OFFSET}")

    elif cmd=='/testsupabase':
        sb_url=os.getenv("SUPABASE_URL","NO_URL"); sb_key=os.getenv("SUPABASE_KEY","NO_KEY")
        send(chat_id,f"🔍 *DEBUG SUPABASE*\nURL: `{sb_url[:40]}`\nKEY: `{sb_key[:20]}...`\nUSE_SUPABASE: `{USE_SUPABASE}`")
        if USE_SUPABASE:
            try:
                r=requests.get(f"{SUPABASE_URL}/rest/v1/tasas?select=count",
                    headers={"apikey":SUPABASE_KEY,"Authorization":f"Bearer {SUPABASE_KEY}"},timeout=10)
                send(chat_id,f"✅ Supabase responde: `{r.status_code}` — `{r.text[:100]}`")
            except Exception as e: send(chat_id,f"❌ Error: `{e}`")
        else: send(chat_id,"❌ USE_SUPABASE es False")

    elif cmd in ('/ayuda','/start','/help'):
        send(chat_id,"""🤖 *GSA CAMBIOS v6.0 — COMANDOS*

*📊 TASAS*
/tasas | /western TASA | /limites

*📡 MERCADO EN VIVO*
/mercado | /patron bs | /patron hoy | /patron semana

*🧮 SIMULADOR*
/simular CLP 500000 | /simular BS 300000

*💼 CAPITAL*
/capital | /umbral CUENTA MONTO

*💱 OPERACIONES*
/op (operación rápida) | /operaciones | /ultima\n\n*📒 LEDGER v7*\n/cierredia | /traslado | /ajuste\n/reconstruir DD/MM/YYYY | /patrimonio

*📊 KPIs*
/kpi | /inventario | /profundidad
/clientes rentabilidad

*📡 ANUNCIO*
/mianuncio

*🎯 META Y OPORTUNIDADES*
/meta USDT | /ranking

*💰 SALDOS*
/saldo | /caja | /posicion

*📈 RESULTADOS*
/dashboard | /resultado | /ganancia

*📋 PENDIENTES*
/cxc | /cxp | /cobrado ID | /pagado ID

*👥 CLIENTES*
/clientes top | /clientes riesgo
/resumen_corresponsal NOMBRE
/oportunidades | /oportunidades 30

*⚙️ SISTEMA*
/importar | /sync | /saldo_inicial CUENTA MONTO
/setsaldo CUENTA MONTO | /testsupabase | /version""")

    elif hay_conv_activa(chat_id):
        send(chat_id, procesar_conv(chat_id, texto))

# ══════════════════════════════════════════════════════════════════════
# PROCESADOR DE DOCUMENTOS
# ══════════════════════════════════════════════════════════════════════
def procesar_documento(chat_id, file_id, nombre):
    nombre_lower = nombre.lower()
    if not (nombre_lower.endswith('.xlsx') or nombre_lower.endswith('.csv')):
        send(chat_id, f"⚠️ Solo acepto .xlsx o .csv\nRecibí: `{nombre}`"); return

    send(chat_id, f"⏳ Procesando `{nombre}`...")
    contenido = download_file(file_id)
    if not contenido:
        send(chat_id, "❌ No pude descargar el archivo."); return

    import tempfile
    if nombre_lower.endswith('.csv'):
        try:
            import csv as _csv, io
            from openpyxl import Workbook
            texto = contenido.decode('utf-8-sig', errors='replace')
            reader = _csv.reader(io.StringIO(texto))
            rows = list(reader)
            wb_tmp = Workbook(); ws_tmp = wb_tmp.active
            for _ in range(9): ws_tmp.append([])
            header_idx = 0
            for i, row in enumerate(rows):
                if row and 'Order Number' in str(row): header_idx = i; break
            if header_idx < len(rows):
                ws_tmp.append(['', ''] + rows[header_idx])
                for row in rows[header_idx+1:]:
                    if row: ws_tmp.append(['', ''] + row)
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
                wb_tmp.save(tmp.name); ruta_tmp = tmp.name
        except Exception as e:
            send(chat_id, f"❌ Error convirtiendo CSV: {e}"); return
    else:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            tmp.write(contenido); ruta_tmp = tmp.name

    try:
        resultado = importar_c2c_inteligente(ruta_tmp, DB_PATH, str(chat_id))
        send(chat_id, formatear_resultado_inteligente(resultado))
    except Exception as e:
        import traceback
        send(chat_id, f"❌ Error: `{e}`\n```{traceback.format_exc()[-300:]}```")
    finally:
        try: os.remove(ruta_tmp)
        except: pass

# ══════════════════════════════════════════════════════════════════════
# LOOPS AUTOMÁTICOS v6.0
# ══════════════════════════════════════════════════════════════════════

# ══════════════════════════════════════════════════════════════════════
# ESCENARIOS DE SESIÓN AVANZADOS
# ══════════════════════════════════════════════════════════════════════

# Estado del ranking para detectar cuando te superan
ultimo_ranking_compra = {}  # {precio, usuario}

def verificar_ranking_bs(mk_c, mk_v):
    """Detecta si alguien superó tu posición en el ranking."""
    global ultimo_ranking_compra
    if not mk_c: return None
    
    precio_actual_1 = mk_c[0]['precio']
    usuario_actual_1 = mk_c[0]['usuario']
    
    if not ultimo_ranking_compra:
        ultimo_ranking_compra = {'precio': precio_actual_1, 'usuario': usuario_actual_1}
        return None
    
    precio_ant = ultimo_ranking_compra['precio']
    
    # Si subió el precio #1 es porque alguien entró con más
    if precio_actual_1 > precio_ant + 0.5:
        ultimo_ranking_compra = {'precio': precio_actual_1, 'usuario': usuario_actual_1}
        # Solo alerta si hay sesión activa
        if sesion_activa:
            chat_id = list(sesion_activa.keys())[0]
            s = sesion_activa[chat_id]
            nuevo_precio_sug = precio_actual_1 + 1
            margen_nuevo = mk_v[0]['precio'] - nuevo_precio_sug if mk_v else 0
            m = f"⚠️ *TE SUPERARON EN COMPRA BS*\n\n"
            m += f"Nuevo #1: `{usuario_actual_1}` a `{precio_actual_1:.2f} Bs`\n"
            m += f"Tu entrada fue: `{s.get('precio_entrada', precio_ant):.2f} Bs`\n\n"
            if margen_nuevo >= 8:
                m += f"💡 Ajusta a `{nuevo_precio_sug:.2f} Bs` para recuperar #1\n"
                m += f"Nuevo margen si ajustas: `{margen_nuevo:.1f} Bs` ✅ rentable"
            else:
                m += f"⚠️ Si ajustas a `{nuevo_precio_sug:.2f} Bs`\n"
                m += f"Margen quedaría en `{margen_nuevo:.1f} Bs` — poco rentable"
            return m
    
    ultimo_ranking_compra = {'precio': precio_actual_1, 'usuario': usuario_actual_1}
    return None

def verificar_capital_operacion(datos_op):
    """Verifica si hay capital suficiente y el impacto de la operación."""
    saldos = get_saldos()
    t = get_ultima_tasa()
    mon_entrada = datos_op.get('mon_entrada','')
    monto = datos_op.get('monto_entrada', 0) or 0
    
    alertas = []
    
    # Verificar capital disponible
    if mon_entrada == 'CLP':
        disponible = saldos.get('CLP_COPEC_PAY', {}).get('saldo', 0) or 0
        restante = disponible - monto
        if restante < 50000:
            alertas.append({
                'tipo': 'capital_bajo',
                'msg': f"CLP restante tras operación: `{restante:,.0f} CLP` — casi agotado"
            })
        
        # Ver si hay spread activo que podría aprovechar
        mk_c, mk_v = get_top_anuncios_bs(min_trans_ves=1000)
        if mk_c and mk_v:
            spread_mk = mk_v[0]['precio'] - mk_c[0]['precio']
            if spread_mk >= SPREAD_MIN_ALERTA and monto > disponible * 0.7:
                dol_obs = t.get('dolar_obs', 895) or 895
                usdt_op = monto / dol_obs
                gan_cliente = usdt_op * 0.05  # margen estimado
                gan_spread = (disponible * 0.5 / mk_c[0]['precio']) * spread_mk / mk_c[0]['precio'] if mk_c[0]['precio'] else 0
                alertas.append({
                    'tipo': 'oportunidad_spread',
                    'spread': spread_mk,
                    'gan_cliente': gan_cliente,
                    'gan_spread': gan_spread,
                })
    
    elif mon_entrada == 'BS':
        disponible = (saldos.get('BS_BANESCO', {}).get('saldo', 0) or 0) +                      (saldos.get('BS_MERCANTIL', {}).get('saldo', 0) or 0)
        restante = disponible - monto
        if restante < 10000:
            alertas.append({
                'tipo': 'capital_bajo',
                'msg': f"BS restante tras operación: `{restante:,.0f} Bs` — casi agotado"
            })
    
    return alertas

def msg_alerta_capital_op(alertas, datos_op):
    """Genera mensaje de alerta de capital para una operación."""
    if not alertas: return None
    m = f"⚠️ *REVISA ANTES DE CONFIRMAR*\n\n"
    for a in alertas:
        if a['tipo'] == 'capital_bajo':
            m += f"🔴 {a['msg']}\n\n"
        elif a['tipo'] == 'oportunidad_spread':
            m += f"📊 *Spread BS activo: `{a['spread']:.1f} Bs` 🟢*\n"
            m += f"Esta operación usa ~70% de tu CLP\n"
            m += f"Ganancia cliente est: `~{a['gan_cliente']:.2f} USDT`\n"
            m += f"¿Vale la pena vs spread? Decide con cuidado\n\n"
    m += "_Escribe `si` para confirmar igual o `no` para cancelar_"
    return m

def verificar_corresponsal(nombre_corresponsal):
    """Verifica si el corresponsal tiene CXP pendiente que indique posible falta de fondos."""
    if not nombre_corresponsal: return None
    conn = get_conn()
    cxp = conn.execute("""
        SELECT COALESCE(SUM(monto),0) as total, COUNT(*) as cnt,
               MIN(fecha) as mas_antigua
        FROM cuentas_pendientes
        WHERE contraparte LIKE ? AND estado='Pendiente' AND tipo='CXP'
    """, (f'%{nombre_corresponsal}%',)).fetchone()
    conn.close()
    
    if not cxp or cxp['total'] == 0: return None
    
    dias = 0
    if cxp['mas_antigua']:
        try:
            from datetime import date
            fecha_ant = date.fromisoformat(cxp['mas_antigua'])
            dias = (today_local() - fecha_ant).days
        except: pass
    
    if cxp['total'] > 50 or dias > 5:
        m = f"⚠️ *CORRESPONSAL CON CXP PENDIENTE*\n\n"
        m += f"`{nombre_corresponsal}` tiene:\n"
        m += f"  Pendiente: `{cxp['total']:.4f} USDT` ({cxp['cnt']} registros)\n"
        if dias > 0: m += f"  Más antigua: `{dias} días` sin pagar\n"
        m += f"\n💡 Confirma disponibilidad antes de comprometer la operación"
        return m
    return None

def anticipar_trm():
    """Obtiene TRM del día siguiente si está disponible y anticipa impacto."""
    try:
        import datetime as dt
        manana = (dt.date.today() + dt.timedelta(days=1)).strftime("%Y-%m-%d")
        url = f"https://www.datos.gov.co/resource/32sa-8pi3.json?$where=vigenciadesde>='{manana}T00:00:00.000'&$order=vigenciadesde DESC&$limit=1"
        data = requests.get(url, timeout=10).json()
        if not data: return None
        trm_manana = float(data[0]["valor"])
        t = get_ultima_tasa()
        trm_hoy = t.get('trm', 0) or 0
        if not trm_hoy or not trm_manana: return None
        variacion = trm_manana - trm_hoy
        pct = (variacion / trm_hoy) * 100 if trm_hoy else 0
        if abs(pct) < 0.3: return None  # Cambio mínimo, no alertar
        
        dol_obs = t.get('dolar_obs', 895) or 895
        western = get_config('western_rate', '0')
        try: western_val = float(western)
        except: western_val = 0
        
        limite_clp_cop_nuevo = western_val * (1 - FEE_WU) if western_val else 0
        limite_clp_cop_hoy = t.get('limite_clp_cop', 0) or 0
        
        emoji = "📈" if variacion > 0 else "📉"
        m = f"{emoji} *TRM PUBLICADA PARA MAÑANA*\n\n"
        m += f"TRM hoy:    `{trm_hoy:,.2f} COP`\n"
        m += f"TRM mañana: `{trm_manana:,.2f} COP`\n"
        m += f"Variación:  `{variacion:+.2f} COP` (`{pct:+.1f}%`)\n\n"
        if limite_clp_cop_nuevo and limite_clp_cop_hoy:
            m += f"📐 *Impacto en límites mañana:*\n"
            m += f"Límite CLP/COP hoy:    `{limite_clp_cop_hoy:.4f}`\n"
            if variacion > 0:
                m += f"💡 Mañana puedes cotizar COP más caro\n"
            else:
                m += f"💡 Mañana debes cotizar COP más barato\n"
        return m
    except: return None



# ══════════════════════════════════════════════════════════════════════
# META DIARIA Y CALCULADORA DE RENTABILIDAD
# ══════════════════════════════════════════════════════════════════════

META_DIARIA_KEY = 'meta_diaria_usdt'
HISTORIAL_ROI = []  # lista de {tipo, usdt, minutos} para calcular ROI/hora

def get_meta_diaria():
    try: return float(get_config(META_DIARIA_KEY, '10'))
    except: return 10.0

def get_ganancia_hoy_usdt():
    res = get_resultados_hoy()
    return res.get('ganancia_neta', 0) or 0

def registrar_roi(tipo_op, usdt_ganado, minutos):
    """Registra el ROI de una operación para calcular ROI/hora."""
    HISTORIAL_ROI.append({
        'tipo': tipo_op,
        'usdt': usdt_ganado,
        'minutos': minutos,
        'hora': now_local()
    })
    if len(HISTORIAL_ROI) > 100:
        HISTORIAL_ROI.pop(0)

def calcular_roi_hora_por_tipo():
    """Calcula ROI/hora promedio por tipo de operación."""
    from collections import defaultdict
    agrupado = defaultdict(list)
    for r in HISTORIAL_ROI:
        if r['minutos'] > 0:
            roi_hora = (r['usdt'] / r['minutos']) * 60
            agrupado[r['tipo']].append(roi_hora)
    resultado = {}
    for tipo, valores in agrupado.items():
        resultado[tipo] = round(sum(valores)/len(valores), 4)
    return resultado

def msg_meta_status():
    """Estado de la meta diaria."""
    meta = get_meta_diaria()
    ganado = get_ganancia_hoy_usdt()
    falta = max(0, meta - ganado)
    pct = min(100, (ganado / meta * 100)) if meta > 0 else 0

    if pct >= 100: barra = "🟢🟢🟢🟢🟢"
    elif pct >= 75: barra = "🟢🟢🟢🟢⬜"
    elif pct >= 50: barra = "🟢🟢🟢⬜⬜"
    elif pct >= 25: barra = "🟢🟢⬜⬜⬜"
    else:           barra = "🟢⬜⬜⬜⬜"

    m = f"🎯 *META DIARIA*\n━━━━━━━━━━━━━━━━━━━━\n"
    m += f"Meta: `{meta:.2f} USDT`\n"
    m += f"Ganado hoy: `{ganado:.4f} USDT`\n"
    m += f"Progreso: {barra} `{pct:.0f}%`\n"

    if falta > 0:
        m += f"Falta: `{falta:.4f} USDT`\n\n"
        # Sugerir qué hacer para alcanzar la meta
        t = get_ultima_tasa()
        mk_c, mk_v = get_top_anuncios_bs(min_trans_ves=1000)
        spread_mk = (mk_v[0]['precio'] - mk_c[0]['precio']) if mk_c and mk_v else 0
        saldos = get_saldos()
        bs_disp = (saldos.get('BS_BANESCO',{}).get('saldo',0) or 0) +                   (saldos.get('BS_MERCANTIL',{}).get('saldo',0) or 0)
        clp_disp = saldos.get('CLP_COPEC_PAY',{}).get('saldo',0) or 0

        if spread_mk >= SPREAD_MIN_ALERTA and bs_disp > 0 and mk_c:
            gan_bs = (bs_disp / mk_c[0]['precio']) * spread_mk * 0.9975 / mk_c[0]['precio']
            m += f"💡 *Para alcanzar la meta ahora:*\n"
            m += f"Spread Maker activo: `{spread_mk:.1f} Bs`\n"
            m += f"Con `{bs_disp:,.0f} Bs` disponibles\n"
            m += f"Ganancia estimada sesión: `~{gan_bs:.4f} USDT`\n"
            if gan_bs >= falta:
                m += f"✅ Una sesión BS alcanza la meta"
            else:
                m += f"⚠️ Necesitas también operaciones con clientes"
        else:
            m += f"💡 Sin spread operable ahora\n"
            m += f"Espera alerta de mercado o atiende clientes"
    else:
        m += f"\n🎉 *Meta del día alcanzada*"

    return m

def msg_calculadora_con_meta(mk_c, mk_v, spread_maker):
    """Calcula la sesión óptima considerando la meta diaria."""
    if not mk_c or not mk_v or spread_maker < SPREAD_MIN_ALERTA:
        return None

    meta = get_meta_diaria()
    ganado = get_ganancia_hoy_usdt()
    falta = max(0, meta - ganado)

    saldos = get_saldos()
    bs_disp = (saldos.get('BS_BANESCO',{}).get('saldo',0) or 0) +               (saldos.get('BS_MERCANTIL',{}).get('saldo',0) or 0)

    if bs_disp <= 0: return None

    precio_compra = mk_c[0]['precio']
    precio_venta  = mk_v[0]['precio']

    # Ganancia estimada con todo el capital
    usdt_posible = bs_disp / precio_compra
    gan_total = usdt_posible * (precio_venta - precio_compra) * 0.9975

    # Capital mínimo para cubrir lo que falta
    if falta > 0 and gan_total > 0:
        pct_necesario = falta / gan_total
        bs_minimo = bs_disp * pct_necesario
        bs_sugerido = min(bs_disp, max(bs_minimo * 1.1, bs_disp * 0.6))
    else:
        bs_sugerido = bs_disp * 0.7

    gan_sugerida = (bs_sugerido / precio_compra) * (precio_venta - precio_compra) * 0.9975

    m = f"📐 *CALCULADORA DE SESIÓN*\n"
    m += f"Capital disponible: `{bs_disp:,.0f} Bs`\n"
    m += f"Compra sugerida: `{bs_sugerido:,.0f} Bs`\n"
    m += f"Ganancia estimada: `~{gan_sugerida:.4f} USDT`\n\n"
    m += f"🎯 Meta: `{meta:.2f} USDT` | Llevas: `{ganado:.4f} USDT`\n"
    if falta > 0:
        if gan_sugerida >= falta:
            m += f"✅ Esta sesión completa tu meta del día"
        else:
            m += f"⚠️ Esta sesión aporta `{gan_sugerida:.4f}` de los `{falta:.4f}` que faltan"
    else:
        m += f"🎉 Meta ya alcanzada — esta sesión es ganancia extra"

    return m

def msg_detector_ventana_excepcional(spread_actual):
    """Detecta si el spread está anormalmente alto vs el historial."""
    conn = get_conn()
    ahora = now_local()
    hora_bloque = str(ahora.hour).zfill(2)
    rows = conn.execute("""
        SELECT AVG(bs_spread_real) as prom, COUNT(*) as cnt
        FROM precios_historicos
        WHERE substr(hora,1,2)=? AND bs_spread_real IS NOT NULL
    """, (hora_bloque,)).fetchone()
    conn.close()

    if not rows or not rows['prom'] or rows['cnt'] < 5:
        return None

    prom = rows['prom']
    if prom <= 0: return None
    desviacion = ((spread_actual - prom) / prom) * 100

    if desviacion < 50: return None  # Solo alertar si está 50% por encima del promedio

    m = f"⚡ *VENTANA EXCEPCIONAL*\n\n"
    m += f"Spread ahora: `{spread_actual:.1f} Bs`\n"
    m += f"Promedio histórico a esta hora: `{prom:.1f} Bs`\n"
    m += f"Está `{desviacion:.0f}%` por encima del promedio\n\n"
    m += f"Esta ventana ocurre pocas veces\n"
    m += f"💡 *Prioriza esta operación sobre todo*"
    return m

def msg_prioridad_oportunidades():
    """Ranking de oportunidades en tiempo real."""
    t = get_ultima_tasa()
    mk_c, mk_v = get_top_anuncios_bs(min_trans_ves=1000)
    compras_clp, ventas_clp = get_top_anuncios_clp()
    saldos = get_saldos()

    pat_bs = ((t.get('ban_bs_compra',0) or 0)+(t.get('ban_bs_venta',0) or 0))/2 or 1
    dol_obs = t.get('dolar_obs',1) or 1

    bs_disp  = (saldos.get('BS_BANESCO',{}).get('saldo',0) or 0) +                (saldos.get('BS_MERCANTIL',{}).get('saldo',0) or 0)
    clp_disp = saldos.get('CLP_COPEC_PAY',{}).get('saldo',0) or 0
    usdt_disp = saldos.get('USDT_BINANCE',{}).get('saldo',0) or 0

    oportunidades = []

    # Arbitraje BS
    if mk_c and mk_v and bs_disp > 1000:
        spread_mk = mk_v[0]['precio'] - mk_c[0]['precio']
        if spread_mk >= SPREAD_MIN_ALERTA:
            usdt_pos = bs_disp / mk_c[0]['precio']
            gan = usdt_pos * spread_mk * 0.9975 / mk_c[0]['precio']
            roi = (gan / (bs_disp / pat_bs)) * 100 if bs_disp > 0 else 0
            roi_hora = roi * 4  # estimado 15 min por operacion
            oportunidades.append({
                'nombre': 'Arbitraje BS Maker',
                'gan_usdt': round(gan, 4),
                'roi': round(roi, 2),
                'roi_hora': round(roi_hora, 2),
                'capital': f'{bs_disp:,.0f} Bs',
                'disponible': True,
                'score': min(100, int(roi_hora * 10 + spread_mk * 2)),
            })

    # Triangular CLP→USDT→BS→CLP
    if compras_clp and mk_v and clp_disp > 10000:
        precio_clp = compras_clp[0]['precio']
        precio_bs_v = mk_v[0]['precio']
        usdt_tri = clp_disp / precio_clp
        bs_tri = usdt_tri * precio_bs_v * (1 - FEE_USDT_BS)
        clp_rec = bs_tri / pat_bs * dol_obs
        gan_clp = clp_rec - clp_disp
        gan_usdt = gan_clp / dol_obs if dol_obs else 0
        roi_tri = (gan_usdt / (clp_disp / dol_obs)) * 100 if clp_disp > 0 else 0
        if gan_usdt > 0:
            roi_hora_tri = roi_tri * 3  # ~20 min
            oportunidades.append({
                'nombre': 'Triangular CLP→USDT→BS',
                'gan_usdt': round(gan_usdt, 4),
                'roi': round(roi_tri, 2),
                'roi_hora': round(roi_hora_tri, 2),
                'capital': f'{clp_disp:,.0f} CLP',
                'disponible': True,
                'score': min(100, int(roi_hora_tri * 10)),
            })

    # Operacion cliente CLP→BS estimada
    if t.get('tasa_gsa_clp_bs') and clp_disp > 50000:
        gan_est = (clp_disp / dol_obs) * 0.045  # margen 4.5%
        roi_cli = 4.5
        roi_hora_cli = roi_cli * 4
        oportunidades.append({
            'nombre': 'Cliente CLP→BS',
            'gan_usdt': round(gan_est, 4),
            'roi': roi_cli,
            'roi_hora': round(roi_hora_cli, 2),
            'capital': f'{clp_disp:,.0f} CLP',
            'disponible': clp_disp > 50000,
            'score': 60,
        })

    if not oportunidades:
        return "❌ Sin oportunidades operables ahora.\nCapital insuficiente o spread bajo."

    # Ordenar por score
    oportunidades.sort(key=lambda x: x['score'], reverse=True)

    m = f"🔥 *RANKING DE OPORTUNIDADES*\n"
    m += f"📅 {now_local().strftime('%H:%M')}\n━━━━━━━━━━━━━━━━━━━━\n\n"

    for i, op in enumerate(oportunidades, 1):
        emoji = "🥇" if i==1 else "🥈" if i==2 else "🥉"
        m += f"{emoji} *{op['nombre']}*\n"
        m += f"  Ganancia est: `{op['gan_usdt']:.4f} USDT`\n"
        m += f"  ROI/hora: `{op['roi_hora']:.1f}%/h`\n"
        m += f"  Capital: `{op['capital']}`\n"
        m += f"  Score: `{op['score']}/100`\n\n"

    m += f"━━━━━━━━━━━━━━━━━━━━\n"
    m += f"💡 Ejecuta primero el de mayor score"
    return m



# ══════════════════════════════════════════════════════════════════════
# MONITOR DE ANUNCIO GSA_CAMBIOS EN BINANCE
# ══════════════════════════════════════════════════════════════════════
GSA_USUARIO = "GSA_Cambios"
ultimo_estado_gsa = {
    'clp': {'posicion': None, 'precio': None, 'activo': False},
    'bs':  {'posicion': None, 'precio': None, 'activo': False},
}

def buscar_anuncio_gsa(fiat, side, pay_types=None):
    """Busca el anuncio de GSA_Cambios en Binance y retorna posición y precio."""
    url = "https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search"
    headers = {"Content-Type": "application/json"}
    try:
        body = {
            "asset": "USDT", "fiat": fiat, "merchantCheck": False,
            "page": 1, "publisherType": None, "rows": 20,
            "tradeType": side, "payTypes": pay_types or [],
            "transAmount": "1000"
        }
        r = requests.post(url, headers=headers, json=body, timeout=10)
        ads = r.json().get("data", [])
        for pos, a in enumerate(ads, 1):
            adv2 = a.get("advertiser", {})
            if adv2.get("nickName", "") == GSA_USUARIO:
                adv = a.get("adv", {})
                return {
                    'encontrado': True,
                    'posicion': pos,
                    'precio': float(adv.get("price", 0)),
                    'disponible': float(adv.get("surplusAmount", 0)),
                    'min_trans': float(adv.get("minSingleTransAmount", 0)),
                }
        return {'encontrado': False}
    except: return {'encontrado': False}

def verificar_anuncio_gsa():
    """Verifica estado del anuncio GSA_Cambios y genera alertas."""
    global ultimo_estado_gsa
    alertas = []

    # Verificar anuncio CLP (BUY = GSA compra USDT pagando CLP)
    gsa_clp = buscar_anuncio_gsa("CLP", "BUY", [])
    estado_ant_clp = ultimo_estado_gsa['clp']

    if gsa_clp['encontrado']:
        pos = gsa_clp['posicion']
        precio = gsa_clp['precio']
        pos_ant = estado_ant_clp.get('posicion')
        precio_ant = estado_ant_clp.get('precio')

        # Escenario 1 — Te superaron (bajaste de posición)
        if pos_ant and pos > pos_ant:
            # Buscar quién está en #1
            top = buscar_top1_sin_gsa("CLP", "BUY")
            nuevo_precio_sug = top['precio'] + 1 if top else precio + 1
            margen_nuevo = nuevo_precio_sug  # CLP no tiene spread como BS
            alertas.append({
                'tipo': 'superado_clp',
                'msg': f"⚠️ *TE SUPERARON EN CLP*\n\n"
                       f"GSA_Cambios: posición #{pos} (antes #{pos_ant})\n"
                       f"Precio actual: `{precio:.2f} CLP`\n"
                       f"💡 Ajusta a `{nuevo_precio_sug:.2f} CLP` para subir"
            })

        # Escenario 8 — Bajaste mucho del ranking
        if pos > 3 and (not pos_ant or pos_ant <= 3):
            alertas.append({
                'tipo': 'ranking_bajo_clp',
                'msg': f"📉 *PERDISTE VISIBILIDAD EN CLP*\n\n"
                       f"GSA_Cambios ahora en posición #{pos}\n"
                       f"Precio: `{precio:.2f} CLP`\n"
                       f"💡 Revisa y ajusta precio para volver al top 3"
            })

        # Escenario 3 — Mercado cayó, tu precio quedó alto
        top1 = buscar_top1_sin_gsa("CLP", "BUY")
        if top1 and precio > top1['precio'] + 2:
            alertas.append({
                'tipo': 'precio_alto_clp',
                'msg': f"⚠️ *TU PRECIO CLP QUEDÓ ALTO*\n\n"
                       f"Tu precio: `{precio:.2f} CLP`\n"
                       f"Mejor del mercado: `{top1['precio']:.2f} CLP`\n"
                       f"Diferencia: `{precio - top1['precio']:.2f} CLP`\n"
                       f"💡 Nadie te venderá con esta diferencia"
            })

        ultimo_estado_gsa['clp'] = {
            'posicion': pos, 'precio': precio, 'activo': True
        }
    else:
        # No se encontró — anuncio pausado o agotado
        if estado_ant_clp.get('activo'):
            alertas.append({
                'tipo': 'anuncio_inactivo_clp',
                'msg': f"🔕 *ANUNCIO CLP INACTIVO*\n\n"
                       f"GSA_Cambios no aparece en el mercado CLP\n"
                       f"Puede estar pausado o agotado\n"
                       f"💡 Revisa tu anuncio en Binance"
            })
        ultimo_estado_gsa['clp'] = {'posicion': None, 'precio': None, 'activo': False}

    # Verificar anuncio BS (SELL = GSA vende USDT recibiendo BS)
    gsa_bs = buscar_anuncio_gsa("VES", "SELL", ["Banesco", "Mercantil"])
    estado_ant_bs = ultimo_estado_gsa['bs']

    if gsa_bs['encontrado']:
        pos = gsa_bs['posicion']
        precio = gsa_bs['precio']
        pos_ant = estado_ant_bs.get('posicion')

        # Escenario 1 BS — Te superaron
        if pos_ant and pos > pos_ant:
            top = buscar_top1_sin_gsa("VES", "SELL")
            nuevo_precio = top['precio'] + 1 if top else precio + 1
            alertas.append({
                'tipo': 'superado_bs',
                'msg': f"⚠️ *TE SUPERARON EN BS VENTA*\n\n"
                       f"GSA_Cambios: posición #{pos} (antes #{pos_ant})\n"
                       f"Tu precio: `{precio:.2f} Bs`\n"
                       f"💡 Ajusta a `{nuevo_precio:.2f} Bs` para recuperar #1"
            })

        # Escenario 7 — Spread alto pero tu anuncio BS no está visible
        mk_c, mk_v = get_top_anuncios_bs(min_trans_ves=1000)
        if mk_c and mk_v:
            spread_mk = mk_v[0]['precio'] - mk_c[0]['precio']
            if spread_mk >= SPREAD_MIN_ALERTA and pos > 3:
                alertas.append({
                    'tipo': 'spread_alto_ranking_bajo',
                    'msg': f"🔥 *SPREAD ALTO PERO PIERDES POSICIÓN*\n\n"
                           f"Spread Maker: `{spread_mk:.1f} Bs` 🟢\n"
                           f"GSA_Cambios en posición #{pos}\n"
                           f"💡 Ajusta precio para subir al top 3 y aprovechar"
                })

        ultimo_estado_gsa['bs'] = {
            'posicion': pos, 'precio': precio, 'activo': True
        }
    else:
        if estado_ant_bs.get('activo'):
            # Verificar si hay spread alto y anuncio inactivo
            mk_c, mk_v = get_top_anuncios_bs(min_trans_ves=1000)
            if mk_c and mk_v:
                spread_mk = mk_v[0]['precio'] - mk_c[0]['precio']
                if spread_mk >= SPREAD_MIN_ALERTA:
                    alertas.append({
                        'tipo': 'anuncio_inactivo_spread_alto',
                        'msg': f"🔥 *SPREAD ALTO Y ANUNCIO INACTIVO*\n\n"
                               f"Spread Maker: `{spread_mk:.1f} Bs` 🟢\n"
                               f"GSA_Cambios NO está en el mercado BS\n"
                               f"💡 Activa tu anuncio ahora para aprovechar"
                    })
                else:
                    alertas.append({
                        'tipo': 'anuncio_inactivo_bs',
                        'msg': f"🔕 *ANUNCIO BS INACTIVO*\n\n"
                               f"GSA_Cambios no aparece en mercado BS\n"
                               f"💡 Revisa tu anuncio en Binance"
                    })
        ultimo_estado_gsa['bs'] = {'posicion': None, 'precio': None, 'activo': False}

    return alertas

def buscar_top1_sin_gsa(fiat, side, pay_types=None):
    """Busca el #1 del mercado excluyendo GSA_Cambios."""
    url = "https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search"
    headers = {"Content-Type": "application/json"}
    try:
        r = requests.post(url, headers=headers, json={
            "asset": "USDT", "fiat": fiat, "merchantCheck": False,
            "page": 1, "publisherType": None, "rows": 5,
            "tradeType": side, "payTypes": pay_types or [],
            "transAmount": "1000"
        }, timeout=10)
        ads = r.json().get("data", [])
        for a in ads:
            adv2 = a.get("advertiser", {})
            if adv2.get("nickName", "") != GSA_USUARIO:
                adv = a.get("adv", {})
                return {
                    'precio': float(adv.get("price", 0)),
                    'usuario': adv2.get("nickName", "—"),
                }
        return None
    except: return None

def msg_estado_anuncio_gsa():
    """Muestra estado actual del anuncio GSA_Cambios en ambos mercados."""
    gsa_clp = buscar_anuncio_gsa("CLP", "BUY", [])
    gsa_bs  = buscar_anuncio_gsa("VES", "SELL", ["Banesco", "Mercantil"])

    m = f"📊 *ESTADO ANUNCIO GSA_Cambios*\n"
    m += f"📅 {now_local().strftime('%d/%m %I:%M %p')}\n"
    m += f"━━━━━━━━━━━━━━━━━━━━\n\n"

    # CLP
    m += "🇨🇱 *Mercado CLP:*\n"
    if gsa_clp['encontrado']:
        top1 = buscar_top1_sin_gsa("CLP", "BUY")
        dif = gsa_clp['precio'] - top1['precio'] if top1 else 0
        emoji_pos = "🥇" if gsa_clp['posicion'] == 1 else "🥈" if gsa_clp['posicion'] == 2 else "🥉" if gsa_clp['posicion'] == 3 else "📉"
        m += f"  {emoji_pos} Posición: #{gsa_clp['posicion']}\n"
        m += f"  Precio: `{gsa_clp['precio']:.2f} CLP`\n"
        m += f"  Disponible: `{gsa_clp['disponible']:.2f} USDT`\n"
        if top1:
            m += f"  Mejor competidor: `{top1['precio']:.2f} CLP` ({top1['usuario']})\n"
            if dif > 0:
                m += f"  ✅ Estás `{dif:.2f} CLP` por encima del mercado\n"
            elif dif < 0:
                m += f"  ⚠️ Estás `{abs(dif):.2f} CLP` por debajo — considera subir\n"
    else:
        m += f"  ❌ Anuncio no encontrado — pausado o agotado\n"

    m += "\n🇻🇪 *Mercado BS:*\n"
    if gsa_bs['encontrado']:
        top1_bs = buscar_top1_sin_gsa("VES", "SELL", ["Banesco", "Mercantil"])
        emoji_pos = "🥇" if gsa_bs['posicion'] == 1 else "🥈" if gsa_bs['posicion'] == 2 else "🥉" if gsa_bs['posicion'] == 3 else "📉"
        m += f"  {emoji_pos} Posición: #{gsa_bs['posicion']}\n"
        m += f"  Precio: `{gsa_bs['precio']:.2f} Bs`\n"
        m += f"  Disponible: `{gsa_bs['disponible']:.2f} USDT`\n"
        if top1_bs:
            m += f"  Mejor competidor: `{top1_bs['precio']:.2f} Bs` ({top1_bs['usuario']})\n"
    else:
        m += f"  ❌ Anuncio no encontrado — pausado o agotado\n"

    return m



# ══════════════════════════════════════════════════════════════════════
# INVENTARIO USDT CON CPP PROMEDIO PONDERADO
# ══════════════════════════════════════════════════════════════════════

def init_inventario():
    """Crea tabla de inventario si no existe."""
    conn = get_conn()
    conn.execute("""CREATE TABLE IF NOT EXISTS inventario_usdt (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        cantidad REAL DEFAULT 0,
        cpp_bs REAL DEFAULT 0,
        ultima_compra_precio REAL DEFAULT 0,
        ultima_actualizacion DATETIME DEFAULT CURRENT_TIMESTAMP
    )""")
    # Insertar registro inicial si no existe
    row = conn.execute("SELECT id FROM inventario_usdt LIMIT 1").fetchone()
    if not row:
        conn.execute("INSERT INTO inventario_usdt (cantidad, cpp_bs) VALUES (0, 0)")
    conn.commit()
    conn.close()

def get_inventario():
    """Retorna el inventario actual de USDT."""
    conn = get_conn()
    row = conn.execute("SELECT * FROM inventario_usdt ORDER BY id DESC LIMIT 1").fetchone()
    conn.close()
    if row:
        return {'cantidad': row['cantidad'] or 0, 'cpp_bs': row['cpp_bs'] or 0,
                'ultima_compra': row['ultima_compra_precio'] or 0}
    return {'cantidad': 0, 'cpp_bs': 0, 'ultima_compra': 0}

def actualizar_inventario_compra(usdt_comprado, precio_compra_bs):
    """Actualiza inventario cuando se compra USDT. Recalcula CPP."""
    inv = get_inventario()
    cantidad_ant = inv['cantidad']
    cpp_ant = inv['cpp_bs']

    # Recalcular CPP promedio ponderado
    total_bs_ant = cantidad_ant * cpp_ant
    total_bs_nuevo = usdt_comprado * precio_compra_bs
    nueva_cantidad = cantidad_ant + usdt_comprado
    nuevo_cpp = (total_bs_ant + total_bs_nuevo) / nueva_cantidad if nueva_cantidad > 0 else precio_compra_bs

    conn = get_conn()
    conn.execute("""UPDATE inventario_usdt SET
        cantidad=?, cpp_bs=?, ultima_compra_precio=?,
        ultima_actualizacion=CURRENT_TIMESTAMP""",
        (round(nueva_cantidad, 4), round(nuevo_cpp, 4), precio_compra_bs))
    conn.commit()
    conn.close()

    # Registrar en Supabase
    if USE_SUPABASE:
        supa_insert('inventario_usdt', {
            'cantidad': round(nueva_cantidad, 4),
            'cpp_bs': round(nuevo_cpp, 4),
            'ultima_compra_precio': precio_compra_bs,
        })

    return {'cantidad': nueva_cantidad, 'cpp_bs': nuevo_cpp}

def consumir_inventario(usdt_requerido, precio_venta_bs):
    """
    Consume inventario para una operación.
    Si hay suficiente: usa CPP del inventario.
    Si no hay suficiente: usa existencia + compra el resto al precio actual.
    Retorna desglose de ganancia financiera.
    """
    inv = get_inventario()
    disponible = inv['cantidad']
    cpp = inv['cpp_bs']

    resultado = {
        'usdt_de_inventario': 0,
        'usdt_comprado_ahora': 0,
        'cpp_inventario': cpp,
        'precio_compra_ahora': precio_venta_bs,  # precio actual si hay que comprar
        'ganancia_financiera_bs': 0,
        'ganancia_financiera_usdt': 0,
        'cpp_efectivo': 0,
        'inventario_restante': 0,
    }

    if disponible >= usdt_requerido:
        # CASO A — Inventario suficiente
        resultado['usdt_de_inventario'] = usdt_requerido
        resultado['usdt_comprado_ahora'] = 0
        resultado['ganancia_financiera_bs'] = (precio_venta_bs - cpp) * usdt_requerido
        resultado['ganancia_financiera_usdt'] = resultado['ganancia_financiera_bs'] / precio_venta_bs if precio_venta_bs else 0
        resultado['cpp_efectivo'] = cpp
        nueva_cantidad = disponible - usdt_requerido

    elif disponible > 0:
        # CASO B — Inventario parcial
        faltante = usdt_requerido - disponible
        precio_actual = precio_venta_bs  # compra al precio actual del mercado

        resultado['usdt_de_inventario'] = disponible
        resultado['usdt_comprado_ahora'] = faltante
        resultado['precio_compra_ahora'] = precio_actual

        # Ganancia solo sobre lo que venía del inventario
        resultado['ganancia_financiera_bs'] = (precio_venta_bs - cpp) * disponible
        resultado['ganancia_financiera_usdt'] = resultado['ganancia_financiera_bs'] / precio_venta_bs if precio_venta_bs else 0

        # CPP efectivo ponderado
        total_bs = (disponible * cpp) + (faltante * precio_actual)
        resultado['cpp_efectivo'] = total_bs / usdt_requerido if usdt_requerido > 0 else precio_actual
        nueva_cantidad = 0

    else:
        # CASO C — Sin inventario
        resultado['usdt_de_inventario'] = 0
        resultado['usdt_comprado_ahora'] = usdt_requerido
        resultado['precio_compra_ahora'] = precio_venta_bs
        resultado['ganancia_financiera_bs'] = 0
        resultado['ganancia_financiera_usdt'] = 0
        resultado['cpp_efectivo'] = precio_venta_bs
        nueva_cantidad = 0

    # Actualizar inventario
    conn = get_conn()
    conn.execute("""UPDATE inventario_usdt SET
        cantidad=?, ultima_actualizacion=CURRENT_TIMESTAMP""",
        (round(max(0, nueva_cantidad), 4),))
    conn.commit()
    conn.close()

    resultado['inventario_restante'] = round(max(0, nueva_cantidad), 4)
    return resultado

def msg_inventario_op(resultado_consumo, usdt_requerido, precio_venta_bs):
    """Genera mensaje de desglose de inventario para una operación."""
    m = ""
    caso = ""

    if resultado_consumo['usdt_comprado_ahora'] == 0:
        caso = "A"
        m += f"📦 *Inventario suficiente*\n"
        m += f"  {usdt_requerido:.4f} USDT del inventario\n"
        m += f"  CPP: `{resultado_consumo['cpp_inventario']:.2f} Bs`\n"
    elif resultado_consumo['usdt_de_inventario'] > 0:
        caso = "B"
        m += f"📦 *Inventario parcial*\n"
        m += f"  {resultado_consumo['usdt_de_inventario']:.4f} USDT del inventario (CPP `{resultado_consumo['cpp_inventario']:.2f} Bs`)\n"
        m += f"  {resultado_consumo['usdt_comprado_ahora']:.4f} USDT al precio actual (`{resultado_consumo['precio_compra_ahora']:.2f} Bs`)\n"
    else:
        caso = "C"
        m += f"📦 *Sin inventario — compra al precio actual*\n"
        m += f"  {usdt_requerido:.4f} USDT a `{resultado_consumo['precio_compra_ahora']:.2f} Bs`\n"

    if resultado_consumo['ganancia_financiera_bs'] > 0:
        m += f"\n💰 Ganancia financiera: `{resultado_consumo['ganancia_financiera_bs']:.2f} Bs` (`{resultado_consumo['ganancia_financiera_usdt']:.4f} USDT`)\n"
    elif caso == "C":
        m += f"\n⚪ Sin ganancia financiera (sin inventario previo)\n"

    m += f"📊 Inventario restante: `{resultado_consumo['inventario_restante']:.4f} USDT`"
    return m

def calcular_ganancia_comercial(tipo_op, monto_entrada, tasa_cliente, t):
    """Calcula la ganancia comercial (margen sobre tasa límite)."""
    limite = None
    if tipo_op in ('CLP→BS', 'BS→CLP'):
        limite = t.get('limite_clp_bs', 0) or 0
    elif tipo_op in ('CLP→COP', 'COP→CLP'):
        limite = t.get('limite_clp_cop', 0) or 0
    elif tipo_op in ('COP→BS', 'BS→COP'):
        limite = t.get('limite_bs_cop', 0) or 0

    if not limite or not tasa_cliente: return 0

    dol_obs = t.get('dolar_obs', 1) or 1
    usdt_equiv = monto_entrada / dol_obs if 'CLP' in tipo_op else monto_entrada / (t.get('ban_bs_compra',1) or 1)

    margen = abs(tasa_cliente - limite) / limite if limite else 0
    return round(usdt_equiv * margen, 4)

def get_ganancia_latente():
    """Calcula la ganancia latente del inventario actual."""
    inv = get_inventario()
    if inv['cantidad'] <= 0 or inv['cpp_bs'] <= 0:
        return 0, 0

    t = get_ultima_tasa()
    precio_actual = (t.get('ban_bs_venta', 0) or 0)
    if not precio_actual: return 0, 0

    ganancia_bs = (precio_actual - inv['cpp_bs']) * inv['cantidad']
    ganancia_usdt = ganancia_bs / precio_actual if precio_actual else 0
    return round(ganancia_bs, 2), round(ganancia_usdt, 4)



# ══════════════════════════════════════════════════════════════════════
# KPIs — ROIC, CAPITAL OCIOSO, RENTABILIDAD POR CLIENTE
# ══════════════════════════════════════════════════════════════════════

def calcular_roic_diario():
    """Calcula el ROIC diario (ganancia / capital invertido)."""
    # Ganancia del día
    res = get_resultados_hoy()
    ganancia_hoy = res.get('ganancia_neta', 0) or 0

    # Capital total en USDT
    saldos = get_saldos()
    t = get_ultima_tasa()
    pat_bs  = ((t.get('ban_bs_compra',0) or 0)+(t.get('ban_bs_venta',0) or 0))/2 or 1
    dol_obs = t.get('dolar_obs',1) or 1
    trm     = t.get('trm',1) or 1

    capital_total = 0
    for cuenta, info in saldos.items():
        saldo = info.get('saldo', 0) or 0
        moneda = info.get('moneda', '') or ''
        capital_total += calcular_usdt_equiv(moneda, saldo, pat_bs, dol_obs)

    roic = (ganancia_hoy / capital_total * 100) if capital_total > 0 else 0
    return {
        'ganancia_hoy': ganancia_hoy,
        'capital_total': capital_total,
        'roic_diario': round(roic, 4),
        'roic_anualizado': round(roic * 365, 2),
    }

def calcular_capital_ocioso():
    """Detecta capital que no está siendo utilizado."""
    saldos = get_saldos()
    t = get_ultima_tasa()
    pat_bs  = ((t.get('ban_bs_compra',0) or 0)+(t.get('ban_bs_venta',0) or 0))/2 or 1
    dol_obs = t.get('dolar_obs',1) or 1

    # Capital total
    capital_total_usdt = 0
    for cuenta, info in saldos.items():
        saldo = info.get('saldo', 0) or 0
        moneda = info.get('moneda', '') or ''
        capital_total_usdt += calcular_usdt_equiv(moneda, saldo, pat_bs, dol_obs)

    # Capital activo hoy (en operaciones del día)
    conn = get_conn()
    hoy = str(today_local())
    vol_hoy = conn.execute("""
        SELECT COALESCE(SUM(usdt_equiv),0) as vol
        FROM operaciones WHERE fecha=? AND estado='Completada'
    """, (hoy,)).fetchone()['vol'] or 0
    conn.close()

    # Inventario activo
    inv = get_inventario()
    capital_inventario = (inv['cantidad'] * inv['cpp_bs'] / pat_bs) if pat_bs else 0

    capital_activo = vol_hoy + capital_inventario
    capital_ocioso = max(0, capital_total_usdt - capital_activo)
    pct_ocioso = (capital_ocioso / capital_total_usdt * 100) if capital_total_usdt > 0 else 0

    return {
        'capital_total': round(capital_total_usdt, 4),
        'capital_activo': round(capital_activo, 4),
        'capital_ocioso': round(capital_ocioso, 4),
        'pct_ocioso': round(pct_ocioso, 1),
        'vol_hoy': round(vol_hoy, 4),
        'capital_inventario': round(capital_inventario, 4),
    }

def msg_kpi():
    """Dashboard de KPIs del negocio."""
    roic = calcular_roic_diario()
    ocioso = calcular_capital_ocioso()
    inv = get_inventario()
    gan_lat_bs, gan_lat_u = get_ganancia_latente()
    res_mes = get_resultados_mes()
    ahora = now_local().strftime("%d/%m %I:%M %p")

    m = f"📊 *KPIs GSA CAMBIOS*\n📅 {ahora}\n━━━━━━━━━━━━━━━━━━━━\n\n"

    # ROIC
    m += f"🎯 *RETORNO SOBRE CAPITAL (ROIC)*\n"
    if roic['roic_diario'] > 0:
        emoji_r = "🚀" if roic['roic_diario'] > 1 else "🟢" if roic['roic_diario'] > 0.5 else "🟡"
    else:
        emoji_r = "🔴"
    m += f"  Ganancia hoy: `{roic['ganancia_hoy']:.4f} USDT`\n"
    m += f"  Capital total: `{roic['capital_total']:.2f} USDT`\n"
    m += f"  ROIC diario: `{roic['roic_diario']:.4f}%` {emoji_r}\n"
    m += f"  ROIC anualizado: `{roic['roic_anualizado']:.2f}%`\n\n"

    # Capital ocioso
    m += f"💤 *CAPITAL OCIOSO*\n"
    emoji_o = "🔴" if ocioso['pct_ocioso'] > 60 else "🟡" if ocioso['pct_ocioso'] > 30 else "🟢"
    m += f"  Capital total: `{ocioso['capital_total']:.2f} USDT`\n"
    m += f"  Activo hoy: `{ocioso['capital_activo']:.2f} USDT`\n"
    m += f"  Ocioso: `{ocioso['capital_ocioso']:.2f} USDT` (`{ocioso['pct_ocioso']:.0f}%`) {emoji_o}\n"
    if ocioso['pct_ocioso'] > 50:
        m += f"  ⚠️ Más del 50% del capital sin mover\n"
    m += "\n"

    # Inventario y ganancia latente
    m += f"📦 *INVENTARIO USDT*\n"
    if inv['cantidad'] > 0:
        m += f"  Cantidad: `{inv['cantidad']:.4f} USDT` | CPP: `{inv['cpp_bs']:.2f} Bs`\n"
        if gan_lat_u != 0:
            emoji_g = "📈" if gan_lat_u > 0 else "📉"
            m += f"  {emoji_g} Ganancia latente: `{gan_lat_u:.4f} USDT`\n"
    else:
        m += f"  Sin inventario actualmente\n"
    m += "\n"

    # P&L del mes
    m += f"📆 *P&L DEL MES*\n"
    m += f"  Ops: `{res_mes['ops']}` | Vol: `{res_mes['volumen']:.4f} USDT`\n"
    m += f"  Ganancia neta: `{res_mes['ganancia_neta']:.4f} USDT`\n"
    if res_mes['ops'] > 0:
        gan_por_op = res_mes['ganancia_neta'] / res_mes['ops']
        m += f"  Ganancia/op: `{gan_por_op:.4f} USDT`\n"

    return m

def msg_clientes_top_rentabilidad():
    """Top clientes por ganancia generada, no solo volumen."""
    conn = get_conn()
    rows = conn.execute("""
        SELECT nombre, operaciones_total, volumen_usdt, ganancia_generada
        FROM clientes
        ORDER BY ganancia_generada DESC LIMIT 10
    """).fetchall()
    conn.close()

    if not rows:
        return "Sin clientes registrados aún."

    m = "🏆 *TOP CLIENTES POR RENTABILIDAD*\n━━━━━━━━━━━━━━━━━━━━\n\n"
    for i, r in enumerate(rows, 1):
        vol = r['volumen_usdt'] or 0
        gan = r['ganancia_generada'] or 0
        roi_cli = (gan / vol * 100) if vol > 0 else 0
        emoji = "🥇" if i==1 else "🥈" if i==2 else "🥉" if i==3 else f"{i}."
        m += f"{emoji} `{r['nombre']}`\n"
        m += f"   Ops: `{r['operaciones_total']}` | Vol: `{vol:.2f} USDT`\n"
        m += f"   Ganancia: `{gan:.4f} USDT` | ROI: `{roi_cli:.2f}%`\n\n"
    return m



# ══════════════════════════════════════════════════════════════════════
# PROFUNDIDAD REAL DE MERCADO BINANCE
# ══════════════════════════════════════════════════════════════════════

def get_profundidad_bs(side, montos=[100, 250, 500, 1000, 2000]):
    """Calcula el precio promedio ponderado para diferentes montos en USDT."""
    url = "https://p2p.binance.com/bapi/c2c/v2/friendly/c2c/adv/search"
    headers = {"Content-Type": "application/json"}
    pay = ["Banesco", "Mercantil"] if side == "SELL" else ["Banesco", "Mercantil", "PagoMovil"]

    try:
        r = requests.post(url, headers=headers, json={
            "asset": "USDT", "fiat": "VES", "merchantCheck": False,
            "page": 1, "publisherType": None, "rows": 20,
            "tradeType": side, "payTypes": pay
        }, timeout=10)
        ads = r.json().get("data", [])
    except: return {}

    # Construir libro de órdenes
    libro = []
    for a in ads:
        adv = a.get("adv", {})
        disponible = float(adv.get("surplusAmount", 0))
        precio = float(adv.get("price", 0))
        if disponible > 0 and precio > 0:
            libro.append({'precio': precio, 'disponible': disponible})

    # Calcular precio promedio ponderado para cada monto
    resultado = {}
    for monto in montos:
        restante = monto
        costo_total = 0
        cubierto = 0
        for orden in libro:
            if restante <= 0: break
            tomar = min(restante, orden['disponible'])
            costo_total += tomar * orden['precio']
            cubierto += tomar
            restante -= tomar
        if cubierto >= monto * 0.95:  # Al menos 95% cubierto
            resultado[monto] = round(costo_total / cubierto, 2)
        else:
            resultado[monto] = None  # No hay suficiente liquidez
    return resultado

def msg_profundidad_mercado():
    """Muestra la profundidad real del mercado BS para diferentes montos."""
    ahora = now_local().strftime("%d/%m %I:%M %p")
    m = f"📊 *PROFUNDIDAD MERCADO BS*\n📅 {ahora}\n━━━━━━━━━━━━━━━━━━━━\n\n"

    compra = get_profundidad_bs("SELL")  # Tú compras USDT pagando Bs
    venta  = get_profundidad_bs("BUY")   # Tú vendes USDT recibiendo Bs

    montos = [100, 250, 500, 1000, 2000]

    m += "📥 *Precio real si COMPRAS X USDT:*\n"
    m += "`Monto    Precio/USDT  Spread real`\n"
    for monto in montos:
        pc = compra.get(monto)
        pv = venta.get(monto)
        if pc and pv:
            spread = pv - pc
            emoji = "🟢" if spread >= 10 else "🟡" if spread >= 5 else "🔴"
            m += f"`{monto:5d} USDT  {pc:,.2f} Bs   {spread:.2f} Bs {emoji}`\n"
        elif pc:
            m += f"`{monto:5d} USDT  {pc:,.2f} Bs   S/D`\n"
        else:
            m += f"`{monto:5d} USDT  Sin liquidez`\n"

    m += "\n📤 *Precio real si VENDES X USDT:*\n"
    for monto in montos:
        pv = venta.get(monto)
        if pv:
            m += f"`{monto:5d} USDT  {pv:,.2f} Bs`\n"
        else:
            m += f"`{monto:5d} USDT  Sin liquidez`\n"

    m += "\n_Precios promedio ponderados por volumen real_"
    return m



# ══════════════════════════════════════════════════════════════════════
# LEDGER — LIBRO MAYOR FINANCIERO v7
# ══════════════════════════════════════════════════════════════════════

TIPOS_MOVIMIENTO = [
    'OPERACION_CLIENTE',   # operación con cliente
    'ARBITRAJE_BINANCE',   # compra/venta en Binance
    'TRASLADO_INTERNO',    # entre cuentas propias
    'DEPOSITO_CAPITAL',    # inyección de capital propio
    'RETIRO_PROPIETARIO',  # retiro personal
    'GASTO_OPERATIVO',     # fee, delivery, suscripción
    'CORRESPONSAL_PAGO',   # pago a corresponsal
    'AJUSTE_AUDITADO',     # corrección con motivo
    'CONCILIACION',        # diferencia de conciliación
    'CIERRE_DIARIO',       # snapshot fin de día
    'PATRIMONIO_INICIAL',  # punto de partida
]

CUENTAS_ACTIVO = [
    'BS_BANESCO', 'BS_MERCANTIL',
    'CLP_COPEC_PAY', 'CLP_BANCOESTADO',
    'COP_EFECTIVO_ORLANDO', 'COP_BANCOLOMBIA_C1',
    'COP_BANCOLOMBIA_C2', 'COP_NEQUI_C1',
    'COP_NEQUI_C2', 'COP_NEQUI_C3',
    'USD_EFECTIVO', 'USDT_BINANCE', 'USDC_AIRTM',
]
CUENTAS_PASIVO = ['CXP_CORRESPONSALES', 'CXP_CLIENTES']
CUENTAS_PATRIMONIO = ['CAPITAL_LENDYS', 'UTILIDADES_RETENIDAS', 'RETIROS_PROPIETARIO']
CUENTAS_INGRESO = ['SPREAD_BS', 'COMISION_CAMBIO', 'COMISION_REFERIDO']
CUENTAS_GASTO = [
    'FEE_BINANCE', 'FEE_BANCARIO', 'DELIVERY_ENCOMIENDA',
    'SUSCRIPCIONES', 'TELEFONIA', 'FEE_RETIRO_CUCUTA',
    'GASTOS_OPERATIVOS',
]
CUENTAS_SISTEMA = ['APERTURA', 'DIFERENCIAS']

def init_ledger():
    """Crea tablas del ledger si no existen."""
    conn = get_conn()

    # Tabla principal del ledger
    conn.execute("""CREATE TABLE IF NOT EXISTS ledger (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha DATE NOT NULL,
        hora TIME NOT NULL,
        tipo_movimiento TEXT NOT NULL,
        cuenta_debe TEXT NOT NULL,
        cuenta_haber TEXT NOT NULL,
        moneda TEXT NOT NULL,
        monto REAL NOT NULL,
        monto_usdt REAL DEFAULT 0,
        tasa_conversion REAL DEFAULT 1,
        referencia_id INTEGER DEFAULT 0,
        referencia_tipo TEXT DEFAULT '',
        descripcion TEXT,
        usuario TEXT,
        fecha_registro DATETIME DEFAULT CURRENT_TIMESTAMP
    )""")

    # Tabla de patrimonio inicial
    conn.execute("""CREATE TABLE IF NOT EXISTS patrimonio_inicial (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha_corte DATE NOT NULL,
        cuenta TEXT NOT NULL,
        moneda TEXT NOT NULL,
        saldo REAL NOT NULL,
        saldo_usdt REAL DEFAULT 0,
        tasa_conversion REAL DEFAULT 1,
        descripcion TEXT,
        fecha_registro DATETIME DEFAULT CURRENT_TIMESTAMP
    )""")

    # Tabla de cierres diarios
    conn.execute("""CREATE TABLE IF NOT EXISTS cierres_diarios (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha DATE UNIQUE NOT NULL,
        estado TEXT DEFAULT 'PENDIENTE',
        patrimonio_usdt REAL DEFAULT 0,
        ganancia_comercial_usdt REAL DEFAULT 0,
        ganancia_financiera_usdt REAL DEFAULT 0,
        ganancia_total_usdt REAL DEFAULT 0,
        gastos_usdt REAL DEFAULT 0,
        utilidad_neta_usdt REAL DEFAULT 0,
        cxc_total REAL DEFAULT 0,
        cxp_total REAL DEFAULT 0,
        inventario_usdt REAL DEFAULT 0,
        cpp_bs REAL DEFAULT 0,
        saldos_json TEXT DEFAULT '{}',
        tasas_cierre_json TEXT DEFAULT '{}',
        ops_total INTEGER DEFAULT 0,
        volumen_usdt REAL DEFAULT 0,
        usuario_cierre TEXT DEFAULT 'sistema',
        fecha_cierre DATETIME DEFAULT CURRENT_TIMESTAMP
    )""")

    # Tabla de costos por operación
    conn.execute("""CREATE TABLE IF NOT EXISTS costos_operacion (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        op_id INTEGER NOT NULL,
        tipo_costo TEXT NOT NULL,
        descripcion TEXT,
        monto REAL NOT NULL,
        moneda TEXT NOT NULL,
        monto_usdt REAL DEFAULT 0,
        fecha DATE DEFAULT CURRENT_DATE,
        fecha_registro DATETIME DEFAULT CURRENT_TIMESTAMP
    )""")

    # Tabla de traslados internos
    conn.execute("""CREATE TABLE IF NOT EXISTS traslados (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        fecha DATE DEFAULT CURRENT_DATE,
        hora TIME,
        cuenta_origen TEXT NOT NULL,
        cuenta_destino TEXT NOT NULL,
        moneda TEXT NOT NULL,
        monto REAL NOT NULL,
        monto_usdt REAL DEFAULT 0,
        descripcion TEXT,
        usuario TEXT,
        fecha_registro DATETIME DEFAULT CURRENT_TIMESTAMP
    )""")

    conn.commit()
    conn.close()
    print("✅ Ledger v7 inicializado")

def ledger_insert(tipo, cuenta_debe, cuenta_haber, moneda, monto,
                  tasa=1, ref_id=0, ref_tipo='', descripcion='', usuario='sistema'):
    """Inserta un movimiento en el ledger."""
    ahora = now_local()
    monto_usdt = calcular_usdt_equiv(moneda, monto, 0, 0)
    conn = get_conn()
    conn.execute("""INSERT INTO ledger
        (fecha, hora, tipo_movimiento, cuenta_debe, cuenta_haber,
         moneda, monto, monto_usdt, tasa_conversion,
         referencia_id, referencia_tipo, descripcion, usuario)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (str(ahora.date()), ahora.strftime("%H:%M"), tipo,
         cuenta_debe, cuenta_haber, moneda, monto, monto_usdt,
         tasa, ref_id, ref_tipo, descripcion, usuario))
    lid = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
    conn.commit(); conn.close()

    if USE_SUPABASE:
        supa_insert('ledger', {
            'fecha': str(ahora.date()), 'hora': ahora.strftime("%H:%M"),
            'tipo_movimiento': tipo, 'cuenta_debe': cuenta_debe,
            'cuenta_haber': cuenta_haber, 'moneda': moneda,
            'monto': monto, 'monto_usdt': monto_usdt,
            'tasa_conversion': tasa, 'referencia_id': ref_id,
            'referencia_tipo': ref_tipo, 'descripcion': descripcion,
            'usuario': usuario,
        })
    return lid

def registrar_patrimonio_inicial(saldos_dict):
    """Registra el patrimonio inicial como punto de partida del ledger."""
    conn = get_conn()
    fecha_corte = str(today_local())
    t = get_ultima_tasa()
    pat_bs = ((t.get('ban_bs_compra',0) or 0)+(t.get('ban_bs_venta',0) or 0))/2 or 1
    dol_obs = t.get('dolar_obs',1) or 1

    for cuenta, datos in saldos_dict.items():
        saldo = datos.get('saldo', 0) or 0
        moneda = datos.get('moneda', '') or ''
        if saldo == 0: continue

        tasa = pat_bs if moneda == 'BS' else dol_obs if moneda == 'CLP' else 1
        saldo_usdt = calcular_usdt_equiv(moneda, saldo, pat_bs, dol_obs)

        conn.execute("""INSERT INTO patrimonio_inicial
            (fecha_corte, cuenta, moneda, saldo, saldo_usdt, tasa_conversion, descripcion)
            VALUES (?,?,?,?,?,?,?)""",
            (fecha_corte, cuenta, moneda, saldo, saldo_usdt, tasa,
             f"Patrimonio inicial al {fecha_corte}"))

        # También insertar en ledger como PATRIMONIO_INICIAL
        ledger_insert(
            tipo='PATRIMONIO_INICIAL',
            cuenta_debe=cuenta,
            cuenta_haber='CAPITAL_LENDYS',
            moneda=moneda,
            monto=saldo,
            tasa=tasa,
            descripcion=f"Saldo inicial {cuenta} al {fecha_corte}",
        )

    conn.commit(); conn.close()
    print(f"✅ Patrimonio inicial registrado al {fecha_corte}")

def registrar_traslado(cuenta_origen, cuenta_destino, monto, descripcion='', usuario='sistema'):
    """Registra un traslado entre cuentas propias en el ledger."""
    conn = get_conn()
    moneda_origen = ''
    for m in ['BS','CLP','COP','USD','USDT','USDC']:
        if m in cuenta_origen: moneda_origen = m; break

    t = get_ultima_tasa()
    pat_bs = ((t.get('ban_bs_compra',0) or 0)+(t.get('ban_bs_venta',0) or 0))/2 or 1
    dol_obs = t.get('dolar_obs',1) or 1
    monto_usdt = calcular_usdt_equiv(moneda_origen, monto, pat_bs, dol_obs)

    # Actualizar saldos
    conn.execute("UPDATE saldos SET saldo=saldo-?,ultima_actualizacion=CURRENT_TIMESTAMP WHERE cuenta=?",
                 (monto, cuenta_origen))
    conn.execute("UPDATE saldos SET saldo=saldo+?,ultima_actualizacion=CURRENT_TIMESTAMP WHERE cuenta=?",
                 (monto, cuenta_destino))

    # Insertar traslado
    ahora = now_local()
    conn.execute("""INSERT INTO traslados
        (fecha, hora, cuenta_origen, cuenta_destino, moneda, monto, monto_usdt, descripcion, usuario)
        VALUES (?,?,?,?,?,?,?,?,?)""",
        (str(ahora.date()), ahora.strftime("%H:%M"),
         cuenta_origen, cuenta_destino, moneda_origen,
         monto, monto_usdt, descripcion, usuario))
    conn.commit(); conn.close()

    # Insertar en ledger
    lid = ledger_insert(
        tipo='TRASLADO_INTERNO',
        cuenta_debe=cuenta_destino,
        cuenta_haber=cuenta_origen,
        moneda=moneda_origen,
        monto=monto,
        descripcion=descripcion or f"Traslado {cuenta_origen}→{cuenta_destino}",
        usuario=usuario,
    )
    return lid

def registrar_costo_operacion(op_id, tipo_costo, descripcion, monto, moneda):
    """Registra un costo asociado a una operación."""
    t = get_ultima_tasa()
    pat_bs = ((t.get('ban_bs_compra',0) or 0)+(t.get('ban_bs_venta',0) or 0))/2 or 1
    dol_obs = t.get('dolar_obs',1) or 1
    monto_usdt = calcular_usdt_equiv(moneda, monto, pat_bs, dol_obs)

    conn = get_conn()
    conn.execute("""INSERT INTO costos_operacion
        (op_id, tipo_costo, descripcion, monto, moneda, monto_usdt)
        VALUES (?,?,?,?,?,?)""",
        (op_id, tipo_costo, descripcion, monto, moneda, monto_usdt))
    conn.commit(); conn.close()

    # Insertar en ledger
    cuenta_gasto = {
        'FEE_BINANCE': 'FEE_BINANCE',
        'FEE_BANCARIO': 'FEE_BANCARIO',
        'DELIVERY': 'DELIVERY_ENCOMIENDA',
        'CORRESPONSAL': 'GASTOS_OPERATIVOS',
        'FEE_RETIRO_CUCUTA': 'FEE_RETIRO_CUCUTA',
    }.get(tipo_costo, 'GASTOS_OPERATIVOS')

    ledger_insert(
        tipo='GASTO_OPERATIVO',
        cuenta_debe=cuenta_gasto,
        cuenta_haber='USDT_BINANCE',
        moneda=moneda,
        monto=monto,
        ref_id=op_id,
        ref_tipo='operacion',
        descripcion=descripcion,
    )

def calcular_costos_operacion(datos_op):
    """Calcula todos los costos asociados a una operación."""
    costos = []
    usdt_equiv = datos_op.get('usdt_equiv', 0) or 0
    tipo_op = datos_op.get('tipo_op', '')

    # Fee Binance si es operación Binance
    if tipo_op in ('BS→USDT', 'USDT→BS', 'CLP→USDT', 'USDT→CLP'):
        fee_binance = usdt_equiv * 0.0002
        if fee_binance > 0:
            costos.append({
                'tipo': 'FEE_BINANCE',
                'descripcion': 'Fee Binance 0.02%',
                'monto': round(fee_binance, 6),
                'moneda': 'USDT',
            })

    # Comisión corresponsal 2.5%
    if datos_op.get('corresponsal') and usdt_equiv > 0:
        comision = usdt_equiv * 0.025
        costos.append({
            'tipo': 'CORRESPONSAL',
            'descripcion': f"Comisión 2.5% → {datos_op['corresponsal']}",
            'monto': round(comision, 4),
            'moneda': 'USDT',
        })

    # Fee retiro Cúcuta
    if datos_op.get('fee_retiro_cucuta', 0) > 0:
        costos.append({
            'tipo': 'FEE_RETIRO_CUCUTA',
            'descripcion': 'Fee retiro efectivo Cúcuta (2,000 COP/millón)',
            'monto': datos_op['fee_retiro_cucuta'],
            'moneda': 'COP',
        })

    # Delivery/encomienda
    if datos_op.get('encomienda_cop', 0) > 0:
        costos.append({
            'tipo': 'DELIVERY',
            'descripcion': 'Costo encomienda/delivery',
            'monto': datos_op['encomienda_cop'],
            'moneda': 'COP',
        })

    return costos

def ejecutar_cierre_diario(usuario='sistema', forzar=False):
    """Ejecuta el cierre diario del sistema."""
    conn = get_conn()
    hoy = str(today_local())

    # Verificar si ya existe cierre hoy
    cierre_exist = conn.execute(
        "SELECT id, estado FROM cierres_diarios WHERE fecha=?", (hoy,)
    ).fetchone()

    if cierre_exist and cierre_exist['estado'] == 'CERRADO' and not forzar:
        conn.close()
        return False, "Ya existe un cierre para hoy."

    # Verificar C2C importado
    c2c_hoy = conn.execute("""
        SELECT COUNT(*) as cnt FROM operaciones
        WHERE fecha=? AND metodo LIKE '%Binance%'
    """, (hoy,)).fetchone()['cnt']

    # Obtener datos del día
    ops = conn.execute("""
        SELECT COUNT(*) as cnt,
               COALESCE(SUM(usdt_equiv),0) as vol,
               COALESCE(SUM(gan_comercial_usdt),0) as gan_com,
               COALESCE(SUM(gan_financiera_usdt),0) as gan_fin
        FROM operaciones WHERE fecha=? AND estado='Completada'
    """, (hoy,)).fetchone()

    gastos = conn.execute("""
        SELECT COALESCE(SUM(monto_usdt),0) as total
        FROM costos_operacion WHERE fecha=?
    """, (hoy,)).fetchone()['total'] or 0

    cxc = conn.execute("""
        SELECT COALESCE(SUM(usdt_equiv),0) as total
        FROM cuentas_pendientes WHERE tipo='CXC' AND estado='Pendiente'
    """).fetchone()['total'] or 0

    cxp = conn.execute("""
        SELECT COALESCE(SUM(usdt_equiv),0) as total
        FROM cuentas_pendientes WHERE tipo='CXP' AND estado='Pendiente'
    """).fetchone()['total'] or 0

    # Calcular patrimonio actual
    saldos = get_saldos()
    t = get_ultima_tasa()
    pat_bs = ((t.get('ban_bs_compra',0) or 0)+(t.get('ban_bs_venta',0) or 0))/2 or 1
    dol_obs = t.get('dolar_obs',1) or 1
    patrimonio = sum(
        calcular_usdt_equiv(info.get('moneda',''), info.get('saldo',0) or 0, pat_bs, dol_obs)
        for info in saldos.values()
    )

    inv = get_inventario()
    gan_com = ops['gan_com'] or 0
    gan_fin = ops['gan_fin'] or 0
    gan_total = gan_com + gan_fin
    utilidad_neta = gan_total - gastos

    import json
    saldos_json = json.dumps({k: {'saldo': v.get('saldo',0), 'moneda': v.get('moneda','')}
                               for k,v in saldos.items()})
    tasas_json = json.dumps({
        'ban_bs_compra': t.get('ban_bs_compra'),
        'ban_bs_venta': t.get('ban_bs_venta'),
        'dolar_obs': t.get('dolar_obs'),
        'trm': t.get('trm'),
        'western': t.get('western'),
    })

    # Insertar o actualizar cierre
    if cierre_exist:
        conn.execute("""UPDATE cierres_diarios SET
            estado='CERRADO', patrimonio_usdt=?, ganancia_comercial_usdt=?,
            ganancia_financiera_usdt=?, ganancia_total_usdt=?, gastos_usdt=?,
            utilidad_neta_usdt=?, cxc_total=?, cxp_total=?,
            inventario_usdt=?, cpp_bs=?, saldos_json=?, tasas_cierre_json=?,
            ops_total=?, volumen_usdt=?, usuario_cierre=?,
            fecha_cierre=CURRENT_TIMESTAMP
            WHERE fecha=?""",
            (patrimonio, gan_com, gan_fin, gan_total, gastos, utilidad_neta,
             cxc, cxp, inv['cantidad'], inv['cpp_bs'],
             saldos_json, tasas_json, ops['cnt'], ops['vol'], usuario, hoy))
    else:
        conn.execute("""INSERT INTO cierres_diarios
            (fecha, estado, patrimonio_usdt, ganancia_comercial_usdt,
             ganancia_financiera_usdt, ganancia_total_usdt, gastos_usdt,
             utilidad_neta_usdt, cxc_total, cxp_total, inventario_usdt,
             cpp_bs, saldos_json, tasas_cierre_json, ops_total, volumen_usdt,
             usuario_cierre)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (hoy, 'CERRADO', patrimonio, gan_com, gan_fin, gan_total, gastos,
             utilidad_neta, cxc, cxp, inv['cantidad'], inv['cpp_bs'],
             saldos_json, tasas_json, ops['cnt'], ops['vol'], usuario))

    # Insertar en ledger
    ledger_insert(
        tipo='CIERRE_DIARIO',
        cuenta_debe='APERTURA',
        cuenta_haber='APERTURA',
        moneda='USDT',
        monto=patrimonio,
        descripcion=f"Cierre diario {hoy} — Utilidad: {utilidad_neta:.4f} USDT",
        usuario=usuario,
    )

    conn.commit(); conn.close()

    alerta_c2c = "" if c2c_hoy > 0 else "\n⚠️ C2C Binance no importado hoy"
    return True, f"""✅ *CIERRE DIARIO — {hoy}*
━━━━━━━━━━━━━━━━━━━━
Operaciones: `{ops['cnt']}` | Vol: `{ops['vol']:.4f} USDT`
Ganancia comercial: `{gan_com:.4f} USDT`
Ganancia financiera: `{gan_fin:.4f} USDT`
Gastos: `{gastos:.4f} USDT`
━━━━━━━━━━━━━━━━━━━━
💰 *Utilidad neta: `{utilidad_neta:.4f} USDT`*
🏛️ Patrimonio: `{patrimonio:.4f} USDT`
📋 CXC pendiente: `{cxc:.4f} USDT`
📋 CXP pendiente: `{cxp:.4f} USDT`
📦 Inventario: `{inv['cantidad']:.4f} USDT`{alerta_c2c}"""

def reconstruir_dia(fecha_str):
    """Reconstruye el estado completo de un día desde el ledger."""
    conn = get_conn()
    # Buscar cierre de ese día
    cierre = conn.execute(
        "SELECT * FROM cierres_diarios WHERE fecha=?", (fecha_str,)
    ).fetchone()

    if not cierre:
        conn.close()
        return f"Sin cierre registrado para {fecha_str}"

    import json
    try:
        saldos = json.loads(cierre['saldos_json'])
        tasas = json.loads(cierre['tasas_cierre_json'])
    except: saldos = {}; tasas = {}

    m = f"📅 *RECONSTRUCCIÓN — {fecha_str}*\n━━━━━━━━━━━━━━━━━━━━\n\n"
    m += f"*P&L DEL DÍA:*\n"
    m += f"  Ganancia comercial: `{cierre['ganancia_comercial_usdt']:.4f} USDT`\n"
    m += f"  Ganancia financiera: `{cierre['ganancia_financiera_usdt']:.4f} USDT`\n"
    m += f"  Gastos: `{cierre['gastos_usdt']:.4f} USDT`\n"
    m += f"  Utilidad neta: `{cierre['utilidad_neta_usdt']:.4f} USDT`\n\n"
    m += f"*PATRIMONIO AL CIERRE:*\n"
    m += f"  Total: `{cierre['patrimonio_usdt']:.4f} USDT`\n"
    m += f"  Inventario: `{cierre['inventario_usdt']:.4f} USDT`\n\n"
    m += f"*PENDIENTES AL CIERRE:*\n"
    m += f"  CXC: `{cierre['cxc_total']:.4f} USDT`\n"
    m += f"  CXP: `{cierre['cxp_total']:.4f} USDT`\n\n"
    m += f"*OPERACIONES:*\n"
    m += f"  Total: `{cierre['ops_total']}` | Vol: `{cierre['volumen_usdt']:.4f} USDT`\n"
    m += f"  Cerrado por: `{cierre['usuario_cierre']}`\n"
    conn.close()
    return m


def segundos_hasta_proximo_en_punto():
    """Calcula segundos hasta el próximo :00 o :30."""
    ahora = now_local()
    minuto = ahora.minute
    segundo = ahora.second
    if minuto < 30:
        faltan_min = 30 - minuto
    else:
        faltan_min = 60 - minuto
    return faltan_min * 60 - segundo

def loop_tasas():
    """Tasas sincronizadas al reloj: cada :00 y :30. Especiales a las 9AM y 1PM."""
    global ultimo_datos
    print("▶ Loop tasas sincronizado iniciado")

    # Primera espera hasta el próximo :00 o :30
    espera = segundos_hasta_proximo_en_punto()
    print(f"⏳ Esperando {espera}s para sincronizar tasas al reloj...")
    time.sleep(espera)

    while True:
        try:
            ahora = now_local()
            hora_actual = ahora.hour
            minuto_actual = ahora.minute
            es_especial = (hora_actual == 9 and minuto_actual < 5) or \
                          (hora_actual == 13 and minuto_actual < 5)

            print(f"\n⏰ Tasas — {ahora.strftime('%H:%M:%S')} {'★ ESPECIAL' if es_especial else ''}")
            datos = consultar_y_guardar(western_rate)
            ultimo_datos = datos
            send(TELEGRAM_CHAT_ID, construir_mensaje(datos, es_especial=es_especial))
        except Exception as e:
            print(f"Error tasas: {e}")

        # Esperar exactamente hasta el próximo :00 o :30
        time.sleep(segundos_hasta_proximo_en_punto())

def loop_mercado():
    """Monitorea precios reales BS y CLP cada 5 min. Guarda historial y envía alertas."""
    global ultimo_precio_bs_venta, ultimo_precio_clp
    print("▶ Loop mercado iniciado")
    time.sleep(30)  # Espera inicial

    while True:
        try:
            # FUENTE ÚNICA: get_binance_banco_promedio para spread y alertas
            ban_c, ban_v, ban_s = get_binance_banco_promedio("Banesco")
            mer_c, mer_v, mer_s = get_binance_banco_promedio("Mercantil")

            # Mejor banco = el de mayor spread
            if (mer_s or 0) > (ban_s or 0):
                mejor_banco = "Mercantil"
                spread_maker = mer_s or 0
                precio_compra_maker = mer_c or 0
                precio_venta_maker  = mer_v or 0
            else:
                mejor_banco = "Banesco"
                spread_maker = ban_s or 0
                precio_compra_maker = ban_c or 0
                precio_venta_maker  = ban_v or 0

            # Anuncios reales top 2 para mostrar en alertas y guardar historial
            compras_bs, ventas_bs = get_top_anuncios_bs(min_trans_ves=1000)
            clp_ads = get_top_anuncios_clp()
            cop_ads = get_top_anuncios_cop()

            # Guardar historial con precios reales
            compras_clp_h, ventas_clp_h = clp_ads if isinstance(clp_ads, tuple) else ([], clp_ads)
            if compras_bs or ventas_bs:
                guardar_precio_historico(compras_bs, ventas_bs, compras_clp_h, cop_ads)

            # Construir mk_c y mk_v desde el banco ganador para mostrar en alertas
            # Usando precio promedio pero mostrando anuncios reales del mejor banco
            mk_c = compras_bs  # anuncios reales compra
            mk_v = ventas_bs   # anuncios reales venta

            tendencia = analizar_tendencia_spread(spread_maker) if spread_maker > 0 else None
            print(f"[mercado] {mejor_banco} compra={precio_compra_maker} venta={precio_venta_maker} spread={spread_maker:.2f} umbral={SPREAD_MIN_ALERTA}")

            if spread_maker > 0:
                precio_venta_mk = precio_venta_maker
                # Alerta cada 5 min si spread >= 10 Bs y cambio >= 0.5 Bs
                spread_cambio = abs(spread_maker - ultimo_precio_bs_venta)
                if spread_maker >= SPREAD_MIN_ALERTA and spread_cambio >= 0.5:
                    send(TELEGRAM_CHAT_ID, msg_alerta_bs(mk_c, mk_v, spread_maker, precio_compra_maker, precio_venta_maker))
                    ultimo_precio_bs_venta = spread_maker
                    msg_opt = msg_momento_optimo(mk_c, mk_v, spread_maker, tendencia)
                    if msg_opt:
                        send(TELEGRAM_CHAT_ID, msg_opt)

                    # Detector ventana excepcional
                    msg_exc = msg_detector_ventana_excepcional(spread_maker)
                    if msg_exc:
                        send(TELEGRAM_CHAT_ID, msg_exc)

                    # Calculadora con meta
                    msg_calc = msg_calculadora_con_meta(mk_c, mk_v, spread_maker)
                    if msg_calc:
                        send(TELEGRAM_CHAT_ID, msg_calc)
                if spread_maker < SPREAD_MIN_ALERTA and ultimo_precio_bs_venta >= SPREAD_MIN_ALERTA:
                    alerta_cayendo = msg_alerta_spread_cayendo(spread_maker)
                    if alerta_cayendo:
                        send(TELEGRAM_CHAT_ID, alerta_cayendo)
                    ultimo_precio_bs_venta = spread_maker

                    # Verificar arbitraje triangular si hay datos CLP
                    compras_clp_tri = clp_ads[0] if isinstance(clp_ads, tuple) else clp_ads
                    if compras_clp_tri and ultimo_datos:
                        alerta_tri = msg_alerta_triangular(500000, compras_clp_tri, ventas_bs, ultimo_datos)
                        if alerta_tri:
                            send(TELEGRAM_CHAT_ID, alerta_tri)
                        else:
                            # Calcular si habría sido rentable pero no hay capital
                            try:
                                cap = analizar_capital()
                                clp_disponible = cap['clp_total']
                                if clp_disponible < 100000 and clp_ads:
                                    precio_clp = clp_ads[0]['precio']
                                    precio_venta_bs = ventas_bs[0]['precio'] if ventas_bs else 0
                                    t = get_ultima_tasa()
                                    pat_bs = ((t.get('ban_bs_compra',0) or 0)+(t.get('ban_bs_venta',0) or 0))/2 or 1
                                    dol_obs = t.get('dolar_obs',1) or 1
                                    monto_ref = 500000
                                    usdt_ref = monto_ref / precio_clp if precio_clp else 0
                                    bs_ref = usdt_ref * precio_venta_bs * (1-FEE_USDT_BS) if precio_venta_bs else 0
                                    clp_ref = bs_ref / pat_bs * dol_obs if pat_bs else 0
                                    ganancia_ref = clp_ref - monto_ref
                                    if ganancia_ref > 0:
                                        ganancia_usdt = ganancia_ref / dol_obs if dol_obs else 0
                                        registrar_oportunidad_perdida(
                                            tipo='Triangular CLP→USDT→BS→CLP',
                                            descripcion=f'Spread BS {spread:.1f} Bs — triangular rentable pero sin capital CLP',
                                            monto_requerido=monto_ref,
                                            moneda='CLP',
                                            ganancia_estimada=round(ganancia_usdt, 4),
                                            razon=f'Solo {clp_disponible:,.0f} CLP disponible, se necesitan {monto_ref:,.0f} CLP',
                                        )
                            except Exception as _oe:
                                print(f"Error registrando oportunidad: {_oe}")

            # Alerta CLP si cambió >= 5 CLP
            compras_clp, ventas_clp = clp_ads if isinstance(clp_ads, tuple) else ([], clp_ads)
            if ventas_clp:
                # Usamos precio de venta (donde tú vendes USDT) como referencia
                precio_ref = ventas_clp[0]['precio']
                if abs(precio_ref - ultimo_precio_clp) >= CAMBIO_MIN_CLP:
                    send(TELEGRAM_CHAT_ID, msg_alerta_clp(compras_clp, ventas_clp))
                    ultimo_precio_clp = precio_ref
            elif compras_clp:
                precio_ref = compras_clp[0]['precio']
                if abs(precio_ref - ultimo_precio_clp) >= CAMBIO_MIN_CLP:
                    send(TELEGRAM_CHAT_ID, msg_alerta_clp(compras_clp, ventas_clp))
                    ultimo_precio_clp = precio_ref

            # Verificar ranking si hay sesión activa
            if sesion_activa and compras_bs and ventas_bs:
                alerta_ranking = verificar_ranking_bs(compras_bs, ventas_bs)
                if alerta_ranking:
                    send(TELEGRAM_CHAT_ID, alerta_ranking)

        except Exception as e:
            print(f"Error loop_mercado: {e}")

        time.sleep(INTERVALO_MERCADO_SEG)

def loop_western_reminder():
    """Recuerda actualizar Western Union a las 10 AM si no fue actualizado."""
    print("▶ Loop western reminder iniciado")
    western_alertado_hoy = None

    while True:
        try:
            ahora = now_local()
            hoy = str(ahora.date())
            hora = ahora.hour

            if hora == 10 and western_alertado_hoy != hoy:
                western_actualizado = get_config('western_actualizado_hoy', '')
                if western_actualizado != hoy:
                    send(TELEGRAM_CHAT_ID,
                         "⚠️ *WESTERN UNION SIN ACTUALIZAR*\n\n"
                         "No has registrado la tasa Western de hoy.\n"
                         "Usa: `/western 0.0042`\n\n"
                         "_La tasa Western afecta los límites CLP/COP_")
                    western_alertado_hoy = hoy
        except Exception as e:
            print(f"Error western reminder: {e}")

        time.sleep(300)  # Revisar cada 5 min

def loop_reporte_diario():
    print("▶ Loop reporte diario + cierre iniciado")
    reporte_enviado_hoy = None
    cierre_ejecutado_hoy = None
    while True:
        try:
            ahora = now_local()
            hoy = str(ahora.date())
            hora = ahora.hour
            if hora == 23 and reporte_enviado_hoy != hoy:
                send(TELEGRAM_CHAT_ID, generar_reporte_diario())
                reporte_enviado_hoy = hoy
            if hora == 0 and cierre_ejecutado_hoy != hoy:
                conn_c = get_conn()
                c2c_cnt = conn_c.execute("SELECT COUNT(*) as c FROM operaciones WHERE fecha=? AND metodo LIKE '%Binance%'", (hoy,)).fetchone()['c']
                conn_c.close()
                if c2c_cnt == 0:
                    send(TELEGRAM_CHAT_ID, f"⚠️ *CIERRE PENDIENTE {hoy}*\nNo se importo C2C hoy.\n/cierredia forzar → cierra sin C2C")
                else:
                    ok, msg = ejecutar_cierre_diario(usuario='automatico')
                    send(TELEGRAM_CHAT_ID, msg)
                cierre_ejecutado_hoy = hoy
        except Exception as e:
            print(f"Error loop diario: {e}")
        time.sleep(300)

def loop_reporte_semanal():
    """Envía reporte semanal los domingos a las 8 PM."""
    print("▶ Loop reporte semanal iniciado")
    reporte_enviado_semana = None

    while True:
        try:
            ahora = now_local()
            semana = ahora.strftime('%Y-W%W')
            hora = ahora.hour
            dia_semana = ahora.weekday()  # 6 = domingo

            if dia_semana == 6 and hora == 20 and reporte_enviado_semana != semana:
                send(TELEGRAM_CHAT_ID, generar_reporte_semanal())
                reporte_enviado_semana = semana
        except Exception as e:
            print(f"Error reporte semanal: {e}")

        time.sleep(600)

def loop_gestor_capital():
    """Revisa capital y alerta si hay desbalance o liquidez crítica cada 15 min."""
    print("▶ Loop gestor de capital iniciado")
    time.sleep(60)

    while True:
        try:
            analisis = analizar_capital()
            alertas_criticas = [a for a in analisis['alertas'] if a['tipo'] == 'umbral']

            if alertas_criticas:
                m = "🚨 *ALERTA DE LIQUIDEZ*\n\n"
                for a in alertas_criticas:
                    m += f"🔴 *{a['cuenta']}*\n"
                    m += f"   Saldo: `{a['saldo']:,.2f} {a['moneda']}`\n"
                    m += f"   Mínimo: `{a['umbral']:,.2f} {a['moneda']}`\n\n"
                m += "_Usa /capital para análisis completo_"
                send(TELEGRAM_CHAT_ID, m)
        except Exception as e:
            print(f"Error gestor capital: {e}")

        time.sleep(INTERVALO_LIQUIDEZ_SEG)

def loop_trm_anticipacion():
    """Verifica TRM del dia siguiente todos los dias a las 5PM."""
    print("▶ Loop TRM anticipación iniciado")
    trm_alertado_hoy = None
    while True:
        try:
            ahora = now_local()
            hoy = str(ahora.date())
            hora = ahora.hour
            if hora >= 17 and trm_alertado_hoy != hoy:
                msg_trm = anticipar_trm()
                if msg_trm:
                    send(TELEGRAM_CHAT_ID, msg_trm)
                trm_alertado_hoy = hoy
        except Exception as e:
            print(f"Error TRM anticipacion: {e}")
        time.sleep(600)

def loop_monitor_gsa():
    """Monitorea el anuncio de GSA_Cambios en Binance cada 5 min."""
    print("▶ Loop monitor GSA_Cambios iniciado")
    time.sleep(60)  # Espera inicial
    while True:
        try:
            alertas = verificar_anuncio_gsa()
            for alerta in alertas:
                send(TELEGRAM_CHAT_ID, alerta['msg'])
        except Exception as e:
            print(f"Error monitor GSA: {e}")
        time.sleep(INTERVALO_MERCADO_SEG)

def loop_csv():
    print("▶ Loop CSV iniciado")
    while True:
        try: exportar_csv()
        except Exception as e: print(f"Error CSV: {e}")
        time.sleep(INTERVALO_CSV_SEG)

# ══════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════
def main():
    global ultimo_offset
    print("="*50)
    print("   GSA CAMBIOS — BOT v6.0 INICIANDO")
    print("="*50)
    print(f"DB: {DB_PATH}")
    print(f"Supabase: {'ACTIVADO ✅' if USE_SUPABASE else 'DESACTIVADO ❌'}")
    print(f"Zona horaria: UTC{UTC_OFFSET}")

    init_db(); init_saldos()

    if not TELEGRAM_BOT_TOKEN:
        print("❌ TELEGRAM_BOT_TOKEN no configurado"); return

    # Iniciar todos los loops
    threading.Thread(target=loop_tasas,          daemon=True).start()
    threading.Thread(target=loop_mercado,         daemon=True).start()
    threading.Thread(target=loop_western_reminder,daemon=True).start()
    threading.Thread(target=loop_reporte_diario,  daemon=True).start()
    threading.Thread(target=loop_reporte_semanal, daemon=True).start()
    threading.Thread(target=loop_gestor_capital,  daemon=True).start()
    threading.Thread(target=loop_csv,             daemon=True).start()
    threading.Thread(target=loop_trm_anticipacion,daemon=True).start()
    threading.Thread(target=loop_monitor_gsa,    daemon=True).start()

    supa_msg = "✅ Supabase conectado" if USE_SUPABASE else "⚠️ Supabase no configurado"
    send(TELEGRAM_CHAT_ID,
         f"✅ *GSA Cambios Bot v6.0 iniciado*\n"
         f"{supa_msg}\n"
         f"📡 Monitoreo de mercado activo\n"
         f"📊 Historial de precios activado\n"
         f"💼 Gestor de capital activo\n\n"
         f"Usa /ayuda para ver los comandos.")

    print("\n✅ Bot v6.0 corriendo...\n")

    while True:
        try:
            updates = get_updates(ultimo_offset)
            for update in updates:
                ultimo_offset = update["update_id"] + 1
                if "message" in update:
                    msg = update["message"]
                    chat_id = str(msg["chat"]["id"])
                    texto = msg.get("text", "")
                    if "document" in msg:
                        doc = msg["document"]
                        procesar_documento(chat_id, doc.get("file_id"), doc.get("file_name","archivo"))
                    elif texto:
                        procesar(chat_id, texto)
        except Exception as e:
            print(f"Error main: {e}"); time.sleep(5)

if __name__ == "__main__":
    main()
